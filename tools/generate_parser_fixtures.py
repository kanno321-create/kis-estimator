#!/usr/bin/env python3
"""
합성 파서 샘플 생성기 (Synthetic Parser Fixture Generator)
Zero-Mock 준수: 실제 xlsx/csv 파일을 생성·저장하여 I/O 테스트

생성 케이스:
- 탭 2개 (6개): 일반, MCC 포함, 다중 분전반
- 탭 3개 이상 (4개): 고압반 제외, 1+3번 탭만
- OCR 오탈자 2건: '소게', '합게'
- 공백 변이: ±1~2행
"""

import argparse
import csv
import os
from pathlib import Path
from typing import List, Dict, Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
except ImportError:
    print("ERROR: openpyxl not installed")
    print("FIX: pip install openpyxl")
    exit(1)


class FixtureGenerator:
    """합성 샘플 생성기"""

    def __init__(self, output_dir: str):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate_all(self, count: int = 10) -> List[str]:
        """모든 합성 샘플 생성"""
        generated = []

        # 탭 2개 케이스 (6개)
        generated.append(self._generate_2tab_simple())
        generated.append(self._generate_2tab_with_mcc())
        generated.append(self._generate_2tab_multi_panel())
        generated.append(self._generate_2tab_typo_sogae())
        generated.append(self._generate_2tab_spacing_variance())
        generated.append(self._generate_2tab_csv())

        # 탭 3개 이상 케이스 (4개)
        generated.append(self._generate_3tab_highvolt_excluded())
        generated.append(self._generate_4tab_complex())
        generated.append(self._generate_3tab_typo_hapge())
        generated.append(self._generate_3tab_csv())

        return generated[:count]

    def _create_workbook_2tabs(self, filename: str, tab1_data: List[List], tab2_data: List[List]) -> str:
        """2탭 xlsx 생성"""
        wb = Workbook()

        # 탭 1
        ws1 = wb.active
        ws1.title = "저압반1"
        for row in tab1_data:
            ws1.append(row)

        # 탭 2
        ws2 = wb.create_sheet("저압반2")
        for row in tab2_data:
            ws2.append(row)

        filepath = self.output_dir / filename
        wb.save(filepath)
        return str(filepath)

    def _create_workbook_multi_tabs(self, filename: str, tabs_data: Dict[str, List[List]]) -> str:
        """다중 탭 xlsx 생성"""
        wb = Workbook()

        for idx, (title, data) in enumerate(tabs_data.items()):
            if idx == 0:
                ws = wb.active
                ws.title = title
            else:
                ws = wb.create_sheet(title)

            for row in data:
                ws.append(row)

        filepath = self.output_dir / filename
        wb.save(filepath)
        return str(filepath)

    def _create_csv(self, filename: str, data: List[List]) -> str:
        """CSV 생성"""
        filepath = self.output_dir / filename
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerows(data)
        return str(filepath)

    # === 탭 2개 케이스 ===

    def _generate_2tab_simple(self) -> str:
        """케이스1: 탭 2개 - 단순 분전반"""
        tab1 = [
            ["번호", "품명", "규격", "단위", "수량", "단가", "금액"],
            ["1", "배선용차단기", "3P 100A", "EA", "2", "50000", "100000"],
            ["2", "배선용차단기", "3P 50A", "EA", "4", "30000", "120000"],
            ["", "", "", "", "", "소계", "220000"],
        ]

        tab2 = [
            ["번호", "품명", "규격", "단위", "수량", "단가", "금액"],
            ["1", "누전차단기", "3P 60A", "EA", "3", "80000", "240000"],
            ["", "", "", "", "", "소계", "240000"],
        ]

        return self._create_workbook_2tabs("01_2tab_simple.xlsx", tab1, tab2)

    def _generate_2tab_with_mcc(self) -> str:
        """케이스2: 탭 2개 - MCC 포함"""
        tab1 = [
            ["MCC-01 (Motor Control Center)"],
            ["번호", "품명", "규격", "단위", "수량", "단가", "금액"],
            ["1", "전자접촉기", "LS MC-32a", "EA", "5", "45000", "225000"],
            ["2", "열동계전기", "TH-22", "EA", "5", "25000", "125000"],
            ["", "", "", "", "", "소계", "350000"],
        ]

        tab2 = [
            ["번호", "품명", "규격", "단위", "수량", "단가", "금액"],
            ["1", "배선용차단기", "2P 30A", "EA", "10", "20000", "200000"],
            ["", "", "", "", "", "합계", "200000"],
        ]

        return self._create_workbook_2tabs("02_2tab_mcc.xlsx", tab1, tab2)

    def _generate_2tab_multi_panel(self) -> str:
        """케이스3: 탭 2개 - 다중 분전반 (소계 후 공백+새 분전반)"""
        tab1 = [
            ["분전반 A"],
            ["번호", "품명", "규격", "단위", "수량", "단가", "금액"],
            ["1", "차단기", "3P 100A", "EA", "2", "50000", "100000"],
            ["", "", "", "", "", "소계", "100000"],
            [""],  # 공백행
            ["분전반 B"],
            ["번호", "품명", "규격", "단위", "수량", "단가", "금액"],
            ["1", "차단기", "3P 50A", "EA", "3", "30000", "90000"],
            ["", "", "", "", "", "합계", "90000"],
        ]

        tab2 = [
            ["번호", "품명", "규격", "단위", "수량", "단가", "금액"],
            ["1", "누전차단기", "2P 30A", "EA", "5", "40000", "200000"],
            ["", "", "", "", "", "소계", "200000"],
        ]

        return self._create_workbook_2tabs("03_2tab_multi_panel.xlsx", tab1, tab2)

    def _generate_2tab_typo_sogae(self) -> str:
        """케이스4: 탭 2개 - OCR 오탈자 '소게' (편집거리 1)"""
        tab1 = [
            ["번호", "품명", "규격", "단위", "수량", "단가", "금액"],
            ["1", "차단기", "3P 75A", "EA", "3", "40000", "120000"],
            ["", "", "", "", "", "소게", "120000"],  # 오탈자: 소계→소게
        ]

        tab2 = [
            ["번호", "품명", "규격", "단위", "수량", "단가", "금액"],
            ["1", "전선관", "25mm", "M", "50", "3000", "150000"],
            ["", "", "", "", "", "소계", "150000"],
        ]

        return self._create_workbook_2tabs("04_2tab_typo_sogae.xlsx", tab1, tab2)

    def _generate_2tab_spacing_variance(self) -> str:
        """케이스5: 탭 2개 - 공백 변이 (±2행)"""
        tab1 = [
            ["분전반 A"],
            ["번호", "품명", "규격", "단위", "수량", "단가", "금액"],
            ["1", "차단기", "2P 50A", "EA", "4", "25000", "100000"],
            ["", "", "", "", "", "소계", "100000"],
            [""],  # 공백 1행
            [""],  # 공백 2행 (±2행 변이)
            ["분전반 B"],
            ["번호", "품명", "규격", "단위", "수량", "단가", "금액"],
            ["1", "전선", "6mm²", "M", "100", "1500", "150000"],
            ["", "", "", "", "", "합계", "150000"],
        ]

        tab2 = [
            ["번호", "품명", "규격", "단위", "수량", "단가", "금액"],
            ["1", "접지봉", "φ14x1500", "EA", "10", "15000", "150000"],
            ["", "", "", "", "", "소계", "150000"],
        ]

        return self._create_workbook_2tabs("05_2tab_spacing.xlsx", tab1, tab2)

    def _generate_2tab_csv(self) -> str:
        """케이스6: 탭 2개 상당 - CSV (단일 파일)"""
        data = [
            ["[탭1: 저압반1]"],
            ["번호", "품명", "규격", "단위", "수량", "단가", "금액"],
            ["1", "차단기", "3P 100A", "EA", "2", "50000", "100000"],
            ["", "", "", "", "", "소계", "100000"],
            [""],
            ["[탭2: 저압반2]"],
            ["번호", "품명", "규격", "단위", "수량", "단가", "금액"],
            ["1", "누전차단기", "3P 60A", "EA", "3", "80000", "240000"],
            ["", "", "", "", "", "합계", "240000"],
        ]

        return self._create_csv("06_2tab_equiv.csv", data)

    # === 탭 3개 이상 케이스 ===

    def _generate_3tab_highvolt_excluded(self) -> str:
        """케이스7: 탭 3개 - 2번=고압반 제외, 1+3번만 분석"""
        tabs = {
            "저압반1": [
                ["번호", "품명", "규격", "단위", "수량", "단가", "금액"],
                ["1", "차단기", "3P 200A", "EA", "1", "150000", "150000"],
                ["", "", "", "", "", "소계", "150000"],
            ],
            "고압반_SKIP": [  # 2번 탭: 고압반 (무시됨)
                ["고압 차단기"],
                ["번호", "품명", "규격", "단위", "수량", "단가", "금액"],
                ["1", "VCB", "22.9kV", "EA", "1", "5000000", "5000000"],
                ["", "", "", "", "", "소계", "5000000"],
            ],
            "저압반3": [
                ["번호", "품명", "규격", "단위", "수량", "단가", "금액"],
                ["1", "누전차단기", "2P 30A", "EA", "8", "35000", "280000"],
                ["", "", "", "", "", "합계", "280000"],
            ],
        }

        return self._create_workbook_multi_tabs("07_3tab_highvolt_skip.xlsx", tabs)

    def _generate_4tab_complex(self) -> str:
        """케이스8: 탭 4개 - 2번 제외, 1+3+4번 분석"""
        tabs = {
            "분전반1": [
                ["번호", "품명", "규격", "단위", "수량", "단가", "금액"],
                ["1", "차단기", "3P 100A", "EA", "3", "50000", "150000"],
                ["", "", "", "", "", "소계", "150000"],
            ],
            "고압반": [  # 2번 탭: 제외
                ["고압 설비"],
                ["VCB", "22.9kV", "1식", "5000000"],
            ],
            "분전반3": [
                ["번호", "품명", "규격", "단위", "수량", "단가", "금액"],
                ["1", "전선", "16mm²", "M", "200", "2500", "500000"],
                ["", "", "", "", "", "소계", "500000"],
            ],
            "분전반4": [
                ["번호", "품명", "규격", "단위", "수량", "단가", "금액"],
                ["1", "접지", "D16", "EA", "15", "20000", "300000"],
                ["", "", "", "", "", "합계", "300000"],
            ],
        }

        return self._create_workbook_multi_tabs("08_4tab_complex.xlsx", tabs)

    def _generate_3tab_typo_hapge(self) -> str:
        """케이스9: 탭 3개 - OCR 오탈자 '합게' (편집거리 1)"""
        tabs = {
            "저압반1": [
                ["번호", "품명", "규격", "단위", "수량", "단가", "금액"],
                ["1", "차단기", "2P 50A", "EA", "5", "30000", "150000"],
                ["", "", "", "", "", "합게", "150000"],  # 오탈자: 합계→합게
            ],
            "고압반": [  # 2번 탭: 제외
                ["고압설비", "22.9kV"],
            ],
            "저압반3": [
                ["번호", "품명", "규격", "단위", "수량", "단가", "금액"],
                ["1", "전선관", "32mm", "M", "80", "4000", "320000"],
                ["", "", "", "", "", "소계", "320000"],
            ],
        }

        return self._create_workbook_multi_tabs("09_3tab_typo_hapge.xlsx", tabs)

    def _generate_3tab_csv(self) -> str:
        """케이스10: 탭 3개 상당 - CSV"""
        data = [
            ["[탭1: 저압반1]"],
            ["번호", "품명", "규격", "단위", "수량", "단가", "금액"],
            ["1", "차단기", "3P 150A", "EA", "2", "100000", "200000"],
            ["", "", "", "", "", "소계", "200000"],
            [""],
            ["[탭2: 고압반_SKIP]"],
            ["고압차단기", "22.9kV", "1식", "5000000"],
            [""],
            ["[탭3: 저압반3]"],
            ["번호", "품명", "규격", "단위", "수량", "단가", "금액"],
            ["1", "누전차단기", "3P 75A", "EA", "4", "70000", "280000"],
            ["", "", "", "", "", "합계", "280000"],
        ]

        return self._create_csv("10_3tab_equiv.csv", data)


def main():
    parser = argparse.ArgumentParser(description="합성 파서 샘플 생성기")
    parser.add_argument("--out", default="tests/parser/fixtures", help="출력 디렉토리")
    parser.add_argument("--n", type=int, default=10, help="생성할 샘플 개수 (최대 10)")
    args = parser.parse_args()

    generator = FixtureGenerator(args.out)

    print(f"합성 샘플 생성 중... (출력: {args.out})")
    generated = generator.generate_all(args.n)

    print(f"\n[OK] {len(generated)} samples generated:")
    for idx, filepath in enumerate(generated, 1):
        print(f"  {idx}. {filepath}")

    print(f"\n[INFO] Zero-Mock compliant: Real files created (xlsx 6, csv 4)")
    print(f"[INFO] Production gate: Replace with 60 real samples when available")


if __name__ == "__main__":
    main()