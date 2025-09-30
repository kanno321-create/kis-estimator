"""
파서 서비스 (Parser Service)
Excel/CSV 파일 파싱 및 분전반 분석
"""

import csv
import logging
import re
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime, timezone
import uuid

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

from api.services.parser_rules import ParserRules

logger = logging.getLogger(__name__)


class ParserService:
    """파서 서비스"""

    def __init__(self):
        self.rules_engine = ParserRules()

    def parse_file(
        self,
        file_path: str,
        trace_id: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        파일 파싱 메인 함수

        Args:
            file_path: 파일 경로 (xlsx/csv)
            trace_id: 추적 ID

        Returns:
            {
                panels: [{tab, panel_id, items}],
                evidence: {traceId, rules_applied, warnings}
            }
        """
        start_time = datetime.now(timezone.utc)
        trace_id = trace_id or str(uuid.uuid4())

        self.rules_engine.clear_rules()

        try:
            # 파일 확장자별 처리
            file_ext = Path(file_path).suffix.lower()

            if file_ext == ".xlsx":
                tabs_data = self._parse_xlsx(file_path)
            elif file_ext == ".csv":
                tabs_data = self._parse_csv(file_path)
            else:
                raise ValueError(f"Unsupported file format: {file_ext}")

            # 탭 전략 결정
            tab_count = len(tabs_data)
            tabs_to_analyze, tab_rule = self.rules_engine.detect_tab_count_and_strategy(tab_count)

            logger.info(
                f"[{trace_id}] Parsing file: {file_path}, "
                f"tabs={tab_count}, analyze={tabs_to_analyze}, rule={tab_rule}"
            )

            # 각 탭별 분전반 분석
            panels = []
            warnings = []

            for tab_idx in tabs_to_analyze:
                if tab_idx >= len(tabs_data):
                    warnings.append(f"Tab index {tab_idx} out of range")
                    continue

                tab_name, rows = tabs_data[tab_idx]

                # 분전반 경계 탐지
                panel_boundaries = self.rules_engine.find_panel_boundaries(rows)

                for panel_num, (start, end) in enumerate(panel_boundaries, 1):
                    panel_id = f"TAB{tab_idx + 1}_PANEL{panel_num}"
                    panel_rows = rows[start:end + 1]

                    # 아이템 추출
                    items = self._extract_items(panel_rows)

                    panels.append({
                        "tab": tab_name,
                        "tab_index": tab_idx,
                        "panel_id": panel_id,
                        "rows_span": [start, end],
                        "items": items
                    })

            # Duration
            duration_ms = int((datetime.now(timezone.utc) - start_time).total_seconds() * 1000)

            # Evidence 수집
            evidence = {
                "traceId": trace_id,
                "rules_applied": self.rules_engine.get_rules_applied(),
                "warnings": warnings,
                "tabs_detected": tab_count,
                "tabs_analyzed": tabs_to_analyze,
                "panels_count": len(panels),
                "duration_ms": duration_ms
            }

            logger.info(
                f"[{trace_id}] Parsing complete: "
                f"panels={len(panels)}, duration={duration_ms}ms"
            )

            return {
                "panels": panels,
                "evidence": evidence
            }

        except Exception as e:
            duration_ms = int((datetime.now(timezone.utc) - start_time).total_seconds() * 1000)

            logger.error(
                f"[{trace_id}] Parsing failed: {str(e)}, duration={duration_ms}ms",
                exc_info=True
            )

            raise Exception(f"PARSE_FAILED: {str(e)}")

    def _parse_xlsx(self, file_path: str) -> List[Tuple[str, List[List[str]]]]:
        """Excel 파일 파싱"""
        if load_workbook is None:
            raise ImportError("openpyxl not installed")

        wb = load_workbook(file_path, data_only=True)
        tabs_data = []

        for sheet in wb.worksheets:
            rows = []
            for row in sheet.iter_rows(values_only=True):
                rows.append([str(cell) if cell is not None else "" for cell in row])

            tabs_data.append((sheet.title, rows))

        return tabs_data

    def _parse_csv(self, file_path: str) -> List[Tuple[str, List[List[str]]]]:
        """
        CSV 파일 파싱
        [탭N: 이름] 마커로 가상 탭 구분
        """
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            all_rows = list(reader)

        # 가상 탭 분리
        tabs_data = []
        current_tab_name = "Tab1"
        current_rows = []

        for row in all_rows:
            row_text = " ".join(row)

            # [탭N: 이름] 패턴 탐지
            import re
            tab_marker = re.match(r'\[탭(\d+):\s*(.+)\]', row_text)

            if tab_marker:
                # 이전 탭 저장
                if current_rows:
                    tabs_data.append((current_tab_name, current_rows))

                # 새 탭 시작
                tab_num = tab_marker.group(1)
                tab_name = tab_marker.group(2).strip()
                current_tab_name = f"Tab{tab_num}_{tab_name}"
                current_rows = []
            else:
                current_rows.append(row)

        # 마지막 탭 저장
        if current_rows:
            tabs_data.append((current_tab_name, current_rows))

        # 탭 마커 없으면 단일 탭
        if not tabs_data:
            tabs_data.append(("Tab1", all_rows))

        return tabs_data

    def _extract_items(self, rows: List[List[str]]) -> List[Dict[str, Any]]:
        """
        행에서 아이템 추출
        간단한 헤더 기반 매핑 (번호, 품명, 규격, 단위, 수량, 단가, 금액)
        """
        items = []

        # 헤더 탐지 (첫 2행 내에서)
        header_row_idx = None
        for i in range(min(3, len(rows))):
            row_text = " ".join(rows[i]).lower()
            if "번호" in row_text and "품명" in row_text:
                header_row_idx = i
                break

        if header_row_idx is None:
            # 헤더 없으면 전체를 raw 데이터로
            return [{"raw": row} for row in rows]

        # 헤더 이후 데이터 행 추출
        for row in rows[header_row_idx + 1:]:
            # 빈 행 또는 소계/합계 행 제외
            row_text = " ".join(row).lower()
            if not any(row) or "소계" in row_text or "합계" in row_text:
                continue

            # 아이템 생성 (간단한 매핑)
            item = {
                "no": row[0] if len(row) > 0 else "",
                "name": row[1] if len(row) > 1 else "",
                "spec": row[2] if len(row) > 2 else "",
                "unit": row[3] if len(row) > 3 else "",
                "qty": row[4] if len(row) > 4 else "",
                "price": row[5] if len(row) > 5 else "",
                "amount": row[6] if len(row) > 6 else "",
            }

            # 유효한 아이템만 추가 (품명이 있는 경우)
            if item["name"].strip():
                items.append(item)

        return items


# Singleton 인스턴스
parser_service = ParserService()