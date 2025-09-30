# improved_quickquote_excel.py  
# 개선된 Excel 견적서 생성기 - 정확한 셀 위치 매핑
# 목적: quotes/quote_*.json 을 읽어서 "견적서양식.xlsx" 템플릿에 정확히 값을 채워 저장

import os, sys, glob, json, datetime, re
from typing import Optional, Tuple, Dict, List
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

TEMPLATE_NAME = "견적서양식.xlsx"   
EXPORT_DIR = os.path.join(".", "exports")

def find_latest_quote(quotes_dir="quotes"):
    paths = sorted(glob.glob(os.path.join(quotes_dir, "quote_*.json")))
    return paths[-1] if paths else None

def load_payload(path=None):
    if path is None:
        path = find_latest_quote()
        if not path:
            raise FileNotFoundError("quotes\\quote_*.json 파일이 없습니다.")
    with open(path, "r", encoding="utf-8") as f:
        payload = json.load(f)
    return path, payload

def vat_10(amount: int) -> Tuple[int, int]:
    vat = int(round(amount * 0.10))
    total = int(amount + vat)
    return vat, total

def number_to_korean(num):
    """숫자를 한글로 변환"""
    if num == 0:
        return "영원정"
    
    # 한글 숫자 매핑
    digits = ["", "일", "이", "삼", "사", "오", "육", "칠", "팔", "구"]
    units1 = ["", "십", "백", "천"]
    units2 = ["", "만", "억", "조"]
    
    def convert_under_10000(n):
        result = ""
        for i in range(4):
            digit = n % 10
            if digit != 0:
                if digit == 1 and i > 0:  # 십, 백, 천의 경우 일 생략
                    result = units1[i] + result
                else:
                    result = digits[digit] + units1[i] + result
            n //= 10
        return result
    
    result = ""
    unit_idx = 0
    
    while num > 0:
        part = num % 10000
        if part != 0:
            part_korean = convert_under_10000(part)
            if unit_idx > 0:
                result = part_korean + units2[unit_idx] + result
            else:
                result = part_korean + result
        num //= 10000
        unit_idx += 1
    
    return result + "원정"

def safe_set_cell_value(ws, cell_address, value):
    """병합셀 안전 처리하여 값 설정"""
    try:
        cell = ws[cell_address]
        if isinstance(cell, MergedCell):
            # 병합셀이면 병합 범위의 첫 번째 셀에 값 설정
            for merged_range in ws.merged_cells.ranges:
                if cell_address in merged_range:
                    first_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    first_cell.value = value
                    return
        else:
            # 일반 셀이면 바로 설정
            cell.value = value
    except:
        # 실패하면 셀 좌표로 직접 접근
        try:
            from openpyxl.utils import coordinate_from_string
            col, row = coordinate_from_string(cell_address)
            ws.cell(row=row, column=col, value=value)
        except:
            print(f"  ⚠️ 셀 {cell_address} 설정 실패")

def fill_first_tab_precise(ws, payload: dict):
    """정확한 셀 위치에 맞춰 첫 번째 탭 채우기"""
    print("\n📝 견적서 첫 번째 탭 작성 중 (정확한 위치)")
    
    # 데이터 준비
    meta = payload.get("meta", {})
    enc = payload.get("enclosure", {})
    costs = payload.get("costs", {})
    inputs = payload.get("inputs", {})
    grand = int(payload.get("grand_total", 0))
    
    vat, total_vat_inc = vat_10(grand)
    
    # 기본 정보
    customer = meta.get("customer", "")
    company = "한국산업"
    project = meta.get("project", "")
    notes = meta.get("notes", "")
    
    print("💼 기본 정보 입력:")
    
    # A1:K1 행 높이 100% 늘리기
    ws.row_dimensions[1].height = ws.row_dimensions[1].height * 2 if ws.row_dimensions[1].height else 30
    print("  ✍️ A1:K1 행 높이 100% 증가")
    
    # 열 너비 조정
    # B17 열 너비 23% 줄이기
    current_width_B = ws.column_dimensions['B'].width or 10
    ws.column_dimensions['B'].width = current_width_B * 0.77
    
    # D17:E17 열 너비 15% 늘리기
    current_width_D = ws.column_dimensions['D'].width or 10
    current_width_E = ws.column_dimensions['E'].width or 10
    ws.column_dimensions['D'].width = current_width_D * 1.15
    ws.column_dimensions['E'].width = current_width_E * 1.15
    print("  📏 B17 열 너비 23% 감소, D17:E17 열 너비 15% 증가")
    
    # G2:K3 영역에 회사명 삽입 (이미지 대신)
    try:
        from openpyxl.styles import Font, Alignment
        
        # G2:K3 병합
        ws.merge_cells('G2:K3')
        
        # 회사명 입력
        company_cell = ws['G2']
        company_cell.value = "(주) 한국산업"
        
        # 스타일 적용 (글자크기 20, 굵게)
        company_cell.font = Font(size=20, bold=True)
        company_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        print("  ✍️ G2:K3 (주) 한국산업 텍스트 삽입 완료")
    except Exception as e:
        print(f"  ⚠️ 회사명 스타일 적용 실패: {e}")
        # 기본 텍스트는 입력
        safe_set_cell_value(ws, 'G2', "(주) 한국산업")
    
    # A3:D7 영역 정렬 개선 (오와열 맞춤)
    from datetime import datetime
    current_date = datetime.now().strftime("%Y년 %m월 %d일")
    
    # A3:D3 - 일자 정보 (정렬 개선)
    date_text = f"일       자 :         {current_date}"
    safe_set_cell_value(ws, 'A3', date_text)
    print(f"  ✍️ A3:D3 일자: {date_text}")
    
    # A5:D5 - 수신 정보 (정렬 개선)
    reception_text = f"수       신 :         {customer}"
    safe_set_cell_value(ws, 'A5', reception_text)
    print(f"  ✍️ A5:D5 수신: {reception_text}")
    
    # A7:D7 - 건명 정보 (정렬 개선)  
    project_text = f"건       명 :         {project}"
    safe_set_cell_value(ws, 'A7', project_text)
    print(f"  ✍️ A7:D7 건명: {project_text}")
    
    # A9:D9 - 빈 칸으로 유지
    print("  ✍️ A9:D9 공란 유지")
    
    # G7:K7 - FAX (고정값)
    fax_text = "FAX   031-431-1419"
    safe_set_cell_value(ws, 'G7', fax_text)
    print(f"  ✍️ G7:K7 팩스: {fax_text}")
    
    # G8:K8 - 이메일 (고정값)
    email_text = "E-MAIL  master@hkkor.com"
    safe_set_cell_value(ws, 'G8', email_text)
    print(f"  ✍️ G8:K8 이메일: {email_text}")
    
    # G11:K11 - 공란으로 유지
    print("  ✍️ G11:K11 공란 유지")
    
    # G13:K13 - 담당자 정보 (고정값)
    manager_text = "담당자 이예지 대리 031-431-1413"
    safe_set_cell_value(ws, 'G13', manager_text)
    print(f"  ✍️ G13:K13 담당자: {manager_text}")
    
    # 제품 정보 - 여러 제품 지원
    print("\n📊 제품 정보 입력:")
    
    # 제품 데이터 준비
    products = payload.get("products", [])
    if not isinstance(products, list) or len(products) == 0:
        # 기본 제품 하나 생성
        main_info = payload.get("main", {})
        product_name = f"MCCB {inputs.get('main_poles_text','4P')} {main_info.get('amp',inputs.get('main_amp_text',''))} 분전반"
        products = [{
            "name": product_name,
            "qty": 1,
            "unit_price": grand,
            "amount": grand
        }]
    
    # 제품별 입력 (15행부터 시작)
    start_row = 15
    total_amount = 0
    
    for i, product in enumerate(products):
        current_row = start_row + i
        
        product_name = product.get('name', f'제품 {i+1}')
        qty = product.get('qty', 1)
        unit_price = product.get('unit_price', 0)
        
        # A열 - 제품 번호
        safe_set_cell_value(ws, f'A{current_row}', f"{i+1}.")
        
        # B열 - 제품명
        safe_set_cell_value(ws, f'B{current_row}', product_name)
        
        # G열 - 수량
        safe_set_cell_value(ws, f'G{current_row}', qty)
        
        # H열 - 단가
        safe_set_cell_value(ws, f'H{current_row}', unit_price)
        
        # I열 - 금액 (엑셀 수식으로)
        safe_set_cell_value(ws, f'I{current_row}', f"=G{current_row}*H{current_row}")
        
        total_amount += qty * unit_price
        print(f"  📦 제품 {i+1}: {product_name} x{qty} = {qty * unit_price:,}원")
    
    # 합계 행 (제품 목록 다음 행)
    total_row = start_row + len(products)
    safe_set_cell_value(ws, f'A{total_row}', "합계")
    safe_set_cell_value(ws, f'I{total_row}', total_amount)
    print(f"  📦 합계: {total_amount:,}원")
    
    print("\n💰 합계 정보 입력:")
    
    # C15:K15 - 공급가액의 한글 표기 (정확한 금액으로)
    korean_amount = number_to_korean(grand)
    safe_set_cell_value(ws, 'C15', korean_amount)
    print(f"  💰 C15:K15 한글금액: {korean_amount}")
    
    # I17 - 실제 금액 (첫 번째 제품의 수식)
    safe_set_cell_value(ws, 'I17', f"=G17*H17")
    print(f"  💰 I17 첫 번째 제품 금액: =G17*H17")
    
    # I18 - 공급가액 합계 (모든 제품 합계)
    safe_set_cell_value(ws, 'I18', total_amount)
    print(f"  💰 I18 공급가액: {total_amount:,}원")
    
    # I19 - 부가세 포함 금액 (I18 * 1.1)
    safe_set_cell_value(ws, 'I19', f"=I18*1.1")
    print(f"  💰 I19 부가세포함: =I18*1.1")
    
    print("  💰 I20, I22 공란 유지 (추가 제품용)")

def fill_second_tab_precise(ws, payload: dict):
    """정확한 셀 위치에 맞춰 두 번째 탭(내역서) 채우기"""
    print("\n📋 내역서 탭 작성 중 (정확한 위치)")
    
    costs = payload.get("costs", {})
    enc = payload.get("enclosure", {})
    
    # 내역서 항목들 준비
    items = []
    
    # 1. 외함 정보
    enclosure_type = "옥내노출 STEEL 1.6T"  # 기본값
    enclosure_spec = f"{enc.get('W',500)}*{enc.get('H',600)}*{enc.get('D',150)}"
    items.append({
        "name": enclosure_type,
        "spec": enclosure_spec,
        "unit": "면",
        "qty": 1,
        "amount": enc.get('cost', 0)
    })
    
    # 2. 인건비
    labor_sum = costs.get("labor_sum", 0)
    if labor_sum > 0:
        items.append({
            "name": "조립/ET/마그네트 인건비", 
            "spec": "조립+전선정리+마그네트 설치",
            "unit": "식",
            "qty": 1, 
            "amount": labor_sum
        })
    
    # 3. 기타 자재들
    other_meta = costs.get("other_meta", {})
    if isinstance(other_meta, dict):
        material_items = {
            "NT": {"name": "N.T (중성선 단자대)", "unit": "개"},
            "NP_3T_40_200": {"name": "N.P / 3T*40*200", "unit": "개"}, 
            "NP_CARD_HOLDER": {"name": "N.P 카드홀더", "unit": "개"},
            "COATING": {"name": "코팅 처리", "unit": "식"},
            "ELB_SUPPORT": {"name": "ELB 지지대", "unit": "개"},
            "INSULATOR": {"name": "인슐레이터", "unit": "개"}
        }
        
        for key, material_info in material_items.items():
            item_data = other_meta.get(key)
            if item_data and item_data.get("cost", 0) > 0:
                cost = int(item_data.get("cost", 0))
                qty = item_data.get("meta", {}).get("qty", 1) if isinstance(item_data.get("meta"), dict) else 1
                items.append({
                    "name": material_info["name"],
                    "spec": "",
                    "unit": material_info["unit"],
                    "qty": qty,
                    "amount": cost
                })
    
    # 4. 잡자재비
    consumables = costs.get("consumables_sum", 0)
    if consumables > 0:
        items.append({
            "name": "잡자재비",
            "spec": "배선부품, 소모품 등", 
            "unit": "식",
            "qty": 1,
            "amount": consumables
        })
    
    print(f"  📝 {len(items)}개 항목 입력:")
    
    # 데이터 입력 (A3부터 시작)
    start_row = 3
    for i, item in enumerate(items):
        row = start_row + i
        
        # A열 - 번호
        ws.cell(row=row, column=1, value=i + 1)
        
        # B열 - 품명/모델명
        ws.cell(row=row, column=2, value=item["name"])
        
        # C열 - 규격/사양
        ws.cell(row=row, column=3, value=item["spec"])
        
        # D열 - 단위
        ws.cell(row=row, column=4, value=item["unit"])
        
        # E열 - 수량  
        ws.cell(row=row, column=5, value=item["qty"])
        
        # F열 - 단가
        unit_price = item["amount"] // item["qty"] if item["qty"] > 0 else item["amount"]
        ws.cell(row=row, column=6, value=unit_price)
        
        # G열 - 금액
        ws.cell(row=row, column=7, value=item["amount"])
        
        print(f"    📦 {i+1}. {item['name']}: {item['amount']:,}원")

def ensure_workbook(template_path: str) -> Tuple[Workbook, bool]:
    """템플릿 로드 또는 기본 워크북 생성"""
    if os.path.exists(template_path):
        print(f"📄 템플릿 로드: {template_path}")
        wb = load_workbook(template_path)
        return wb, True
    
    print("📄 기본 템플릿 생성")
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "견적서"
    
    # 기본 견적서 레이아웃 생성
    ws1["A1"] = "견적서"
    
    # 내역서 탭
    ws2 = wb.create_sheet("내역서")
    ws2["A1"] = "내역서"
    detail_headers = ["NO", "품명", "규격", "단위", "수량", "단가", "금액"]
    for col, header in enumerate(detail_headers, 1):
        ws2.cell(row=2, column=col, value=header)
    
    return wb, False

def main():
    """메인 실행 함수"""
    print("🚀 정확한 위치 매핑 Excel 견적서 생성기 시작")
    
    # 입력 파일 결정
    input_path = sys.argv[1] if len(sys.argv) > 1 else None
    json_path, payload = load_payload(input_path)
    
    print(f"📥 입력: {os.path.basename(json_path)}")
    
    # 템플릿 로드
    wb, used_template = ensure_workbook(TEMPLATE_NAME)
    
    # 탭 확보
    ws1 = wb.worksheets[0]
    ws2 = wb.worksheets[1] if len(wb.worksheets) > 1 else wb.create_sheet("내역서")
    
    # 각 탭 채우기
    try:
        fill_first_tab_precise(ws1, payload)
        fill_second_tab_precise(ws2, payload)
        
        # 저장
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = os.path.splitext(os.path.basename(json_path))[0]
        os.makedirs(EXPORT_DIR, exist_ok=True)
        output_path = os.path.join(EXPORT_DIR, f"{base_name}_견적서.xlsx")
        
        wb.save(output_path)
        
        print("\n✅ Excel 생성 완료!")
        print(f"📥 입력: {os.path.abspath(json_path)}")
        print(f"📄 템플릿: {'사용' if used_template else '기본 생성'}")
        print(f"📤 출력: {os.path.abspath(output_path)}")
        
    except Exception as e:
        print(f"\n❌ 오류 발생: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()