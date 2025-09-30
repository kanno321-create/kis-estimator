# --- PATH SAFETY: ensure project root on sys.path (for kisan_mapping_v2.py) ---
import os, sys
_BASE_DIR = os.path.dirname(os.path.abspath(__file__))
if _BASE_DIR not in sys.path:
    sys.path.insert(0, _BASE_DIR)
# -------------------------------------------------------------------------------

# system_estimate_ui.py
import os, json, time, re, traceback, threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from estimate_engine import explain_to_mini_report, get_last_explain_safe

# === [자동 대화/행동 로그] 연결 시작 ===
# [SAFE IMPORT: auto_logger]
try:
    from kisan_convo.auto_logger import get_logger  # 표준 경로
except Exception:
    try:
        from auto_logger import get_logger          # 로컬 파일 폴백
    except Exception:
        def get_logger(name: str):
            # 최소 동작: 콘솔만 찍는 더미 로거
            class _Dummy:
                def log(self, *a, **k): print(f"[{name}]", a[0] if a else "", k.get("data", ""))
                def close(self): pass
            return _Dummy()

# === filename helper (Windows-safe, 한글 유지) ===============================
import re
from datetime import datetime

def _slug_kor(name: str) -> str:
    """
    파일명에 쓸 문자열 정리:
    - Windows 금지문자 제거: <>:\"/\\|?*
    - 앞뒤 공백/점 제거
    - 공백을 '_'로
    - 한글은 그대로 유지
    """
    if not name:
        return ""
    s = re.sub(r'[<>:"/\\|?*]+', '', str(name))
    s = s.strip().strip('.')
    s = re.sub(r'\s+', '_', s)
    return s

def _now_tag() -> str:
    return datetime.now().strftime("%y%m%d_%H%M%S")

def make_initial_filename(prefix: str, client_name: str, ext: str) -> str:
    """
    예: make_initial_filename("견적", "한국산업(주)", ".xlsx")
        -> "견적_한국산업(주)_250829_235959.xlsx"
    """
    safe = _slug_kor(client_name) or "NONAME"
    ext = "." + ext.lstrip(".")
    return f"{prefix}_{safe}_{_now_tag()}{ext}"
# ============================================================================ 

# ======== [ADD] RAG 동적 로더 (파일 상단 import 아래에 추가) ========
def _resolve_rag_search():
    """
    query_api.rag_search → 실패 시 assistant_agent.rag_search 로 폴백.
    VS Code/Pylance가 모듈을 못 찾을 때도 런타임에서 안전하게 동작하게 한다.
    반환: callable 또는 None
    """
    import importlib, sys, pathlib
    here = pathlib.Path(__file__).resolve().parent

    if str(MAPPER_DIR) not in sys.path:
        sys.path.append(str(MAPPER_DIR))

    # 1) 현재 디렉토리 우선 탐색 보장
    if str(here) not in sys.path:
        sys.path.append(str(here))

    # 2) query_api.rag_search 시도
    try:
        mod = importlib.import_module("query_api")
        fn = getattr(mod, "rag_search", None)
        if callable(fn):
            return fn
    except Exception:
        fn = None

    # 3) assistant_agent.rag_search 폴백
    try:
        mod = importlib.import_module("assistant_agent")
        fn = getattr(mod, "rag_search", None)
        if callable(fn):
            return fn
    except Exception:
        fn = None

    return None
# ======== [/ADD] ========


# 이 파일(UI) 전용 세션 로거 (파일명 예: logs/YYYYMMDD/HHMMSS_estimate_ui.jsonl)
APP_LOG = get_logger("estimate_ui")

def _log(role: str, data: dict, msg: str = ""):
    """
    role: "user" | "assistant" | "event" 등
    data: dict로 행동/입력/결과를 실어 보냄
    msg : 사람이 읽기 쉬운 한 줄 요약(없으면 data['action'] 시도)
    """
    try:
        text = msg or (data.get("action") if isinstance(data, dict) else "")
        APP_LOG.log(role, text or "", data if isinstance(data, dict) else {"raw": str(data)})
    except Exception as e:
        print("[LOG][FAIL]", e)

def _flush_log(note: str = ""):
    """
    프로그램 종료 직전 마지막 줄 기록 + 파일 핸들 닫기
    (세션 중 로그를 한 줄도 안 남겼더라도 이때 파일이 생성됨)
    """
    try:
        APP_LOG.log("system", f"[FLUSH] {note}")
        APP_LOG.close()
    except Exception as e:
        print("[LOG][CLOSE-FAIL]", e)

    # --- RAG: auto reindex after flush (graceful skip) ---
    try:
        import os
        # 환경 변수로 RAG 활성 제어 (없으면 비활성)
        enabled = os.getenv("RAG_ENABLED", "0").strip().lower() in ("1", "true", "yes", "on")

        # 모듈 유무 확인: AssistantAgent 경로로 시도
        AgentCls = None
        try:
            from assistant_agent import AssistantAgent as AgentCls  # 우선 권장 경로
        except Exception:
            AgentCls = None

        if not enabled or AgentCls is None:
            # 비활성/모듈없음은 에러 아님: 정상 스킵
            print("[RAG] auto reindex after flush: {'ok': True, 'docs': 0, 'skipped': True, 'reason': 'disabled_or_no_module'}")
            return

        # 활성 + 모듈존재 시에만 실제 재색인
        try:
            agent = AgentCls()
        except Exception as e:
            print("[RAG] auto reindex after flush: {'ok': True, 'docs': 0, 'skipped': True, 'reason': 'agent_init_fail'}")
            return

        # 새 인덱싱 경로: estimates/*.json → RAG 인덱스
        res = agent.rebuild_index_from_estimates()

        # 기존/혼합 리턴 타입 호환: int 또는 dict
        if isinstance(res, int):
            docs = res
        elif isinstance(res, dict):
            try:
                docs = int(res.get("docs", 0) or 0)
            except Exception:
                docs = 0
        else:
            try:
                docs = int(getattr(res, "docs", 0) or 0)
            except Exception:
                docs = 0

        print(f"[RAG] auto reindex after flush: {{'ok': True, 'docs': {docs}, 'skipped': False}}")

    except Exception:
        # 예외도 사용자 경험상 에러로 띄우지 않고 '정상 스킵'으로 기록
        print("[RAG] auto reindex after flush: {'ok': True, 'docs': 0, 'skipped': True, 'reason': 'exception_soft_skip'}")


# === [자동 대화/행동 로그] 연결 끝 ===

# ---- RAG 모듈 임포트 (Pylance와 런타임 동시 만족) ----
from typing import TYPE_CHECKING
if TYPE_CHECKING:
    import utils.rag as R  # type: ignore[reportMissingImports]

try:
    import utils.rag as R  # type: ignore[reportMissingImports]
except Exception:
    # VS Code가 폴더 루트로 열리지 않았거나, 경로가 다른 경우 대비
    import sys, pathlib
    sys.path.append(str(pathlib.Path(__file__).resolve().parent))
    try:
        import utils.rag as R  # type: ignore[reportMissingImports]
    except Exception:
        R = None  # 런타임 안전장치
# ------------------------------------------------------


# === RAG 재색인: Tk에서 바로 쓰는 도우미 ===
def _do_rag_reindex():
    """
    AssistantAgent.rebuild_index_from_estimates()를 직접 호출하여
    견적 JSON → RAG 인덱스 갱신을 수행한다.
    - 성공: "RAG 재색인 완료 (entries=숫자)" 안내
    - 비활성/무자료/예외: 사용자 혼란 방지를 위해 '정상 스킵' 또는 정보성 메시지로 처리
    """
    try:
        from assistant_agent import AssistantAgent
        agent = AssistantAgent()

        # 새 방식: dict 반환 표준에 맞춰 안전 처리
        res = None
        try:
            # base_dir, index_dir는 기본값을 사용 (내부에서 안전 처리)
            res = agent.rebuild_index_from_estimates()  # ← 핵심 변경점
        except Exception as _e:
            res = {"ok": False, "docs": 0, "error": str(_e)}

        # 결과 정규화
        ok = False
        docs = 0
        reason = None
        if isinstance(res, dict):
            ok = bool(res.get("ok"))
            docs = int(res.get("docs") or 0)
            reason = res.get("reason") or res.get("error")
        elif isinstance(res, int):
            ok = (res >= 0)
            docs = max(0, res)
        else:
            ok, docs, reason = False, 0, "unknown_result_type"

        # 사용자 안내 메시지 구성
        if ok and docs >= 0:
            msg = f"RAG 재색인 완료 (entries={docs})"
            level = "info"
        else:
            # 비활성/무자료/모듈없음 등은 '정상적으로 스킵'
            msg = "RAG 비활성/데이터 없음 — 정상적으로 스킵했습니다."
            if reason:
                msg += f" ({reason})"
            level = "info"

    except Exception as e:
        # 예외도 사용자 경험상 '정보성'으로만 안내
        msg = f"RAG 스킵(예외 처리): {e}"
        level = "info"

    # UI 팝업(가능하면): 항상 정보성으로만
    try:
        from tkinter import messagebox as _mb
        _mb.showinfo("RAG 재색인", msg)
    except Exception:
        pass

    print("[RAG]", msg)



# === /RAG 재색인 도우미 끝 ===


# === RAG 검색: 간단 질의 → 상위 매치 요약 ===

def _do_rag_query():
    """
    RAG 검색 도우미
    - R 모듈(rag 또는 utils.rag)이 없으면 '정보성 안내' 후 조용히 스킵(에러/경고 금지)
    - 검색 실패도 에러가 아닌 정보성으로만 안내
    """
    try:
        from tkinter import simpledialog as _sd, messagebox as _mb
    except Exception:
        _sd = None; _mb = None  # UI 팝업이 불가해도 조용히 동작

    # RAG 모듈 로드(우선 rag, 실패 시 utils.rag, 둘 다 없으면 None)
    R = None
    try:
        import rag as R  # 표준 위치
    except Exception:
        try:
            import utils.rag as R  # 레거시/대체 위치
        except Exception:
            R = None

    if R is None:
        # 이전: showerror("RAG 검색", f"모듈 로드 실패: {e}") → 변경: 정보성 안내
        msg = "RAG 비활성/데이터 없음 — 검색을 스킵합니다."
        try:
            if _mb: _mb.showinfo("RAG 검색", msg)
        except Exception:
            pass
        print("[RAG]", msg)
        return

    # 질의 입력
    q = None
    try:
        if _sd:
            q = _sd.askstring("RAG 검색", "키워드를 입력하세요 (예: 'LS 30AF 3P'):")
    except Exception:
        pass
    if not q:
        # 입력 취소 시 조용히 종료
        print("[RAG] 검색 취소 또는 빈 질의")
        return

    # 실제 검색
    try:
        # R 모듈에 따라 search API 명이 다를 수 있어 유연하게 호출
        # 우선순위: rag_search(q) → search(q) → query(q)
        res = None
        if hasattr(R, "rag_search"):
            res = R.rag_search(q)
        elif hasattr(R, "search"):
            res = R.search(q)
        elif hasattr(R, "query"):
            res = R.query(q)
        else:
            raise RuntimeError("검색 API(rag_search/search/query)가 없습니다.")

        # 결과 요약
        items = []
        if isinstance(res, dict) and "items" in res:
            items = res.get("items") or []
        elif isinstance(res, (list, tuple)):
            items = list(res)
        # string만 오더라도 안전 처리
        elif isinstance(res, str):
            items = [res]

        if not items:
            msg = "결과가 없습니다."
        else:
            # 상위 5개만 요약
            lines = []
            for i, t in enumerate(items[:5], 1):
                t = str(t)
                if len(t) > 120:
                    t = t[:120] + "…"
                lines.append(f"{i}. {t}")
            msg = "상위 결과:\n" + "\n".join(lines)

        if _mb:
            _mb.showinfo("RAG 검색 결과", msg)
        print("[RAG] query:", q, "→", msg.replace("\n", " | "))

    except Exception as e:
        # 이전: showerror("검색 실패: ...") → 변경: 정보성 안내
        try:
            if _mb: _mb.showinfo("RAG 검색", f"RAG 스킵(예외 처리): {e}")
        except Exception:
            pass
        print("[RAG] 검색 스킵(예외 처리):", e)
# === /RAG 검색 도우미 끝 ===


# === Explainable Estimate: "왜 이 사이즈?" 팝업 ===
def safe_int(val, default=0):
    try:
        return int(str(val).strip())
    except Exception:
        return default

def _collect_enclosure_inputs(self):
    """UI에서 외함 사이즈 입력을 읽어와 안전하게 정수로 변환"""
    W = safe_int(getattr(self, "ent_W", None).get() if hasattr(self, "ent_W") else "")
    H = safe_int(getattr(self, "ent_H", None).get() if hasattr(self, "ent_H") else "")
    D = safe_int(getattr(self, "ent_D", None).get() if hasattr(self, "ent_D") else "")
    return {"W": W, "H": H, "D": D}

def _estimate_with_trace(spec: dict) -> dict:
    """
    estimate_engine 쪽에 explain 기능이 있으면 사용하고,
    없으면 최소한의 규칙으로 트레이스를 생성(임시).
    """
    try:
        import estimate_engine as EE
        # 우선순위 1: 공식 함수가 존재하는 경우
        if hasattr(EE, "estimate_with_trace"):
            return EE.estimate_with_trace(spec)
        if hasattr(EE, "estimate"):
            out = EE.estimate(spec)
            return {"ok": True, "result": out, "trace": {"source": "EE.estimate", "spec": spec}}
    except Exception as e:
        # 엔진이 없어도 임시 규칙으로 설명 생성
        pass

    # 최소 규칙(임시): 폭/높이 보정 + 근거 텍스트
    W, H, D = spec.get("W", 0), spec.get("H", 0), spec.get("D", 0)
    W2 = max(600, (W + 49)//50*50)  # 50mm 라운딩
    H2 = max(900, (H + 49)//50*50)
    D2 = max(250, (D + 49)//50*50)
    explain = {
        "W": {"in": W, "out": W2, "rule": "최소 600 / 50mm 단위 보정"},
        "H": {"in": H, "out": H2, "rule": "최소 900 / 50mm 단위 보정"},
        "D": {"in": D, "out": D2, "rule": "최소 250 / 50mm 단위 보정"},
        "notes": ["임시 규칙: 실제 엔진 미탑재 시 안전 보정값 사용"]
    }
    return {"ok": True, "result": {"W": W2, "H": H2, "D": D2}, "trace": explain}

def _json_pretty(d: dict) -> str:
    import json
    try:
        return json.dumps(d, ensure_ascii=False, indent=2)
    except Exception:
        return str(d)

def _show_info(title: str, msg: str):
    try:
        from tkinter import messagebox as _mb
        _mb.showinfo(title, msg)
    except Exception:
        print(f"[INFO] {title}: {msg}")

def _show_error(title: str, msg: str):
    try:
        from tkinter import messagebox as _mb
        _mb.showerror(title, msg)
    except Exception:
        print(f"[ERROR] {title}: {msg}")

def _trace_popup(title: str, payload: dict):
    """
    사람이 보기 쉬운 '요약' 탭 + 원본(JSON) 탭을 제공하는 팝업.
    - 요약 탭: W/H/D를 큼직하게, 규칙/근거를 표 형태로
    - JSON 탭: 원본 전체 트레이스 열람
    - 카드 저장: 정책/예외 카드로 바로 저장(원하면)
    """
    try:
        import tkinter as tk
        from tkinter import ttk, scrolledtext, messagebox
        # conv_ledger가 없어도 동작하도록 안전 임포트
        try:
            from conv_ledger import add_card as _add_card
        except Exception:
            _add_card = None

        # 안전 파싱
        spec   = (payload or {}).get("spec")   or {}
        result = (payload or {}).get("result") or {}
        trace  = (payload or {}).get("trace")  or {}

        W = result.get("W") if isinstance(result, dict) else None
        H = result.get("H") if isinstance(result, dict) else None
        D = result.get("D") if isinstance(result, dict) else None

        win = tk.Toplevel()
        win.title(title)
        win.geometry("860x620")

        # 상단: 카드 저장 영역 (제목/종류)
        top = ttk.Frame(win); top.pack(fill="x", padx=10, pady=(10,6))
        ttk.Label(top, text="카드 제목").pack(side="left")
        ent_title = ttk.Entry(top, width=42); ent_title.pack(side="left", padx=(6,12))
        ent_title.insert(0, "외함 사이즈 결정 근거")

        var_kind = tk.StringVar(value="policy")
        ttk.Radiobutton(top, text="정책(policy)", variable=var_kind, value="policy").pack(side="left")
        ttk.Radiobutton(top, text="예외(exception)", variable=var_kind, value="exception").pack(side="left", padx=(8,0))

        def _json_pretty_local(d: dict) -> str:
            import json
            try:
                return json.dumps(d, ensure_ascii=False, indent=2)
            except Exception:
                return str(d)

        # 노트북(요약/JSON)
        nb = ttk.Notebook(win); nb.pack(fill="both", expand=True, padx=10, pady=(0,10))

        # === [탭1] 요약 ===
        tab_summary = ttk.Frame(nb)
        nb.add(tab_summary, text="요약 보기")

        # 결과 박스: W x H x D 큼직하게
        box = ttk.Frame(tab_summary); box.pack(fill="x", padx=6, pady=(10,6))
        big = ttk.Label(box, text=f"W×H×D = {W or '-'} × {H or '-'} × {D or '-'} (mm)", font=("맑은 고딕", 16, "bold"))
        big.pack(anchor="w")

        # 입력 스펙 요약
        frm_spec = ttk.LabelFrame(tab_summary, text="입력 스펙", padding=8)
        frm_spec.pack(fill="x", padx=6, pady=(4,6))
        def _add_row(frame, k, v, r):
            ttk.Label(frame, text=k).grid(row=r, column=0, sticky="e", padx=(0,6), pady=2)
            ttk.Label(frame, text=str(v)).grid(row=r, column=1, sticky="w", pady=2)

        r = 0
        enc = (spec or {}).get("enclosure") or {}
        if enc:
            ttk.Label(frm_spec, text=f"외함: {enc.get('외함 종류','-')} / 재질: {enc.get('외함 재질','-')} / 베이스: {enc.get('베이스 유무','-')}").grid(row=r, column=0, columnspan=2, sticky="w"); r+=1
        client = (spec or {}).get("client") or ""
        project = (spec or {}).get("project") or ""
        if client: _add_row(frm_spec, "고객", client, r); r+=1
        if project: _add_row(frm_spec, "프로젝트", project, r); r+=1
        br = (spec or {}).get("branch_specs") or []
        if br:
            _add_row(frm_spec, "분기 요약", ", ".join(map(str, br)), r); r+=1

        # 규칙/근거 테이블 (trace)
        frm_rules = ttk.LabelFrame(tab_summary, text="규칙/근거", padding=8)
        frm_rules.pack(fill="both", expand=True, padx=6, pady=(0,8))

        # trace 포맷 ①: {"W":{"in":..,"out":..,"rule":".."}, "H":{..}, "D":{..}, "notes":[..]}
        # trace 포맷 ②: 기타 딕셔너리 → key: value로 평면화
        tree = ttk.Treeview(frm_rules, columns=("col1","col2","col3"), show="headings", height=10)
        tree.heading("col1", text="항목")
        tree.heading("col2", text="입력→결정")
        tree.heading("col3", text="근거/규칙")
        tree.column("col1", width=120, anchor="center")
        tree.column("col2", width=380)
        tree.column("col3", width=280)
        tree.pack(fill="both", expand=True)

        def _push_rule_row(name, d):
            if not isinstance(d, dict):
                tree.insert("", "end", values=(name, "-", str(d))); return
            _in  = d.get("in",  "-")
            _out = d.get("out", "-")
            _rule= d.get("rule","-")
            io = f"{_in} → {_out}"
            tree.insert("", "end", values=(name, io, _rule))

        if isinstance(trace, dict) and any(k in trace for k in ("W","H","D")):
            for k in ("W","H","D"):
                if k in trace: _push_rule_row(k, trace.get(k))
            # 나머지 키들도 보여주기
            for k,v in trace.items():
                if k in ("W","H","D"): continue
                if k=="notes" and isinstance(v, (list,tuple)):
                    for i, note in enumerate(v, 1):
                        tree.insert("", "end", values=(f"참고{i}", "-", str(note)))
                else:
                    _push_rule_row(k, v)
        elif isinstance(trace, dict):
            for k,v in trace.items():
                _push_rule_row(str(k), v)
        else:
            tree.insert("", "end", values=("trace", "-", str(trace)))

        # === [탭2] 원본(JSON) ===
        tab_json = ttk.Frame(nb)
        nb.add(tab_json, text="원본(JSON)")
        txt = scrolledtext.ScrolledText(tab_json, wrap="word")
        txt.pack(fill="both", expand=True, padx=6, pady=6)
        txt.insert("1.0", _json_pretty_local(payload))
        txt.configure(state="disabled")

        # 하단 버튼: 카드 저장 / 닫기
        bar = ttk.Frame(win); bar.pack(fill="x", padx=10, pady=(0,10))
        def _save_card():
            if _add_card is None:
                messagebox.showwarning("카드 저장", "conv_ledger 모듈을 찾을 수 없어 카드 저장을 건너뜁니다.")
                return
            title_s = (ent_title.get().strip() or "무제")
            try:
                _add_card(kind=var_kind.get(), title=title_s, body=_json_pretty_local(payload))
                messagebox.showinfo("카드 저장", f"[{var_kind.get()}] '{title_s}' 저장 완료")
            except Exception as e:
                messagebox.showerror("카드 저장 실패", str(e))

        ttk.Button(bar, text="카드 저장", command=_save_card).pack(side="left")
        ttk.Button(bar, text="닫기", command=win.destroy).pack(side="right")

    except Exception as e:
        # 팝업 실패 시 정보창으로 폴백
        try:
            from tkinter import messagebox as _mb
            _mb.showinfo(title, f"(요약 생성 실패) 원본 표시:\n\n{_json_pretty_local(payload)}")
        except Exception:
            print("[TRACE POPUP ERROR]", e)


def _collect_branch_brief(self) -> list:
    """분기 차단기 간단 요약 (필드명이 다르면 이 부분은 건너뛰어도 됨)"""
    out = []
    for name in dir(self):
        if name.startswith("ent_branch_") and name.endswith("_spec"):
            try:
                ent = getattr(self, name)
                val = ent.get()
                if val:
                    out.append(val)
            except Exception:
                pass
    return out

def _build_spec_from_ui(self) -> dict:
    enc = _collect_enclosure_inputs(self)
    # 엔진이 요구하는 최소 필드 채우기 (없으면 기본값)
    enc_full = {
        "외함 종류": "표준형",
        "외함 재질": "철제",
        "W": enc.get("W", 0),
        "H": enc.get("H", 0),
        "D": enc.get("D", 0),
    }
    branches = _collect_branch_brief(self)
    return {
        "enclosure": enc_full,
        "branch_specs": branches,
        "client": getattr(self, "ent_client_name", None).get() if hasattr(self, "ent_client_name") else "",
        "project": getattr(self, "ent_project_name", None).get() if hasattr(self, "ent_project_name") else ""
    }


def _ensure_conv_ledger_note():
    """외함 분류 관련 정책카드 기록(shim). 없어도 무시."""
    try:
        from conv_ledger import add_card
        add_card(kind="policy", title="외함 분류 고정",
                 body="외함은 '경제형/표준형' 구분 자체가 없음. 질문/추정 금지.")
    except Exception:
        pass

def _show_enclosure_explain(self):
    try:
        spec = _build_spec_from_ui(self)
        res = _estimate_with_trace(spec.get("enclosure", {}))
        _ensure_conv_ledger_note()
        if not res or not res.get("ok"):
            _show_error("왜 이 사이즈?", "설명 생성 실패")
            return
        payload = {
            "spec": spec,
            "result": res.get("result"),
            "trace": res.get("trace")
        }
        _trace_popup("왜 이 사이즈? — Explain Trace", payload)
        print("[EXPLAIN]", _json_pretty(payload))
    except Exception as e:
        _show_error("왜 이 사이즈?", f"오류: {e}")
# === /Explainable Estimate 끝 ===

# ===== 기존 엔진 유지 사용 =====
from estimate_engine import compute_enclosure_size, build_estimate_lines, get_last_explain

# ===== 선택 의존성 =====
_PIL_OK=_TESS_OK=_PDFPLUMBER_OK=_DOCX_OK=_PDF2IMG_OK=_DND_OK=True
try:
    from PIL import Image
except: _PIL_OK=False
# >>> BEGIN: external I/O imports (guarded)
# 안전 가드 임포트: pytesseract, pdfplumber, python-docx, pdf2image, tkinterdnd2

# pytesseract (OCR)
try:
    import pytesseract  # pip: pytesseract
    _IMPORT_ERROR_PYTESS = None
    # 필요시 경로 지정 예시:
    # pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
except Exception as e:
    pytesseract = None
    _IMPORT_ERROR_PYTESS = e

# pdfplumber (PDF 텍스트/테이블)
try:
    import pdfplumber  # pip: pdfplumber
    _IMPORT_ERROR_PDFPLUMBER = None
except Exception as e:
    pdfplumber = None
    _IMPORT_ERROR_PDFPLUMBER = e

# python-docx (DOCX 파싱) — 모듈 별칭과 Document 둘 다 준비
try:
    import docx as _docx            # 모듈 접근용 (docx.Document)
    from docx import Document       # 바로 호출용 (Document)
    docx = _docx                    # 모듈 이름 노출
    _IMPORT_ERROR_DOCX = None
except Exception as e:
    docx = None
    Document = None
    _IMPORT_ERROR_DOCX = e

# pdf2image (PDF→이미지). Windows는 Poppler 설치 필요
try:
    from pdf2image import convert_from_path  # pip: pdf2image
    _IMPORT_ERROR_PDF2IMG = None
except Exception as e:
    convert_from_path = None
    _IMPORT_ERROR_PDF2IMG = e

# tkinterdnd2 (드래그앤드롭)
try:
    from tkinterdnd2 import TkinterDnD, DND_FILES
    _HAS_TKINTERDND2 = True
    _IMPORT_ERROR_TKDND = None
except Exception as e:
    TkinterDnD = None
    DND_FILES = None
    _HAS_TKINTERDND2 = False
    _IMPORT_ERROR_TKDND = e

def _require_or_hint(obj, name: str, extra_hint: str = ""):
    """필수 모듈이 없을 때 사용자에게 설치/설정 힌트 제공"""
    if obj is None:
        msg = f"[모듈 누락] '{name}'가 필요합니다. " + (extra_hint or "")
        raise RuntimeError(msg)

def _guard_check_before_ocr():
    if pytesseract is None:
        raise RuntimeError(
            "[OCR 사용 불가] pytesseract 미설치 또는 불러오기 실패.\n"
            "- pip install pytesseract\n"
            "- Windows: Tesseract-OCR 설치 및 경로 확인\n"
            f"- import error: {_IMPORT_ERROR_PYTESS}"
        )

def _guard_check_before_pdf_text():
    if pdfplumber is None:
        raise RuntimeError(
            "[PDF 추출 불가] pdfplumber 미설치 또는 불러오기 실패.\n"
            "- pip install pdfplumber\n"
            f"- import error: {_IMPORT_ERROR_PDFPLUMBER}"
        )

def _guard_check_before_docx():
    if Document is None or docx is None:
        raise RuntimeError(
            "[DOCX 추출 불가] python-docx 미설치 또는 불러오기 실패.\n"
            "- pip install python-docx\n"
            f"- import error: {_IMPORT_ERROR_DOCX}"
        )

def _guard_check_before_pdf2image():
    if convert_from_path is None:
        raise RuntimeError(
            "[PDF→이미지 변환 불가] pdf2image 미설치 또는 불러오기 실패.\n"
            "- pip install pdf2image\n"
            "- Windows: Poppler 설치 및 bin 폴더 PATH 추가\n"
            f"- import error: {_IMPORT_ERROR_PDF2IMG}"
        )

def _guard_check_before_tkdnd():
    if not _HAS_TKINTERDND2:
        raise RuntimeError(
            "[드래그앤드롭 비활성] tkinterdnd2 미설치 또는 불러오기 실패.\n"
            "- pip install tkinterdnd2\n"
            f"- import error: {_IMPORT_ERROR_TKDND}"
        )
# >>> END: external I/O imports (guarded)


# --- Excel(.xlsx) 파싱 옵션 ---
_OPENPYXL_OK = True
try:
    from openpyxl import load_workbook
    from openpyxl.cell.cell import MergedCell
except Exception:
    _OPENPYXL_OK = False
    load_workbook = None
    MergedCell = None

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

from assistant_agent import AssistantAgent  # 에이전트

# === 표준 메일 라이브러리
import imaplib, email
from email.header import decode_header
from email.utils import parsedate_to_datetime

APP_TITLE = "한국산업 견적 도우미"
BASE_DIR = os.path.dirname(os.path.dirname(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
EST_DIR  = os.path.join(DATA_DIR, "estimates")
CFG_PATH = os.path.join(DATA_DIR, "settings.json")
CLIENTS_JSON = os.path.join(DATA_DIR, "clients.json")
PRICES_DIR = os.path.join(DATA_DIR, "prices")
for d in (DATA_DIR, EST_DIR, PRICES_DIR):
    os.makedirs(d, exist_ok=True)

# === MAPPER 디렉토리(견적서 맵핑/공식 스키마) 등록 ===
import sys, glob, json, importlib
from pathlib import Path

MAPPER_DIR = Path(__file__).resolve().parent / "MAPPER"
# 폴더가 이미 있으면 통과, 없으면 생성(있어도 안전)
try:
    MAPPER_DIR.mkdir(parents=True, exist_ok=True)
except Exception:
    pass

# 파이썬 모듈 탐색 경로에 추가 (kisan_mapping_v2.py 임포트용)
mp = str(MAPPER_DIR)
if mp not in sys.path:
    sys.path.append(mp)

# 전역 캐시
__KISAN_MAPPER__ = None
__OFFICIAL_SCHEMA__ = None
__OFFICIAL_SCHEMA_PATH__ = None

def _import_kisan_mapper():
    """MAPPER/kisan_mapping_v2.py 동적 임포트 (없으면 None)"""
    global __KISAN_MAPPER__
    if __KISAN_MAPPER__ is not None:
        return __KISAN_MAPPER__
    try:
        mod = importlib.import_module("kisan_mapping_v2")
        __KISAN_MAPPER__ = mod
        print(f"[MAPPER] kisan_mapping_v2 loaded from: {getattr(mod, '__file__', '(memory)')}")
        return mod
    except Exception as e:
        print("[MAPPER] kisan_mapping_v2 not found / load failed:", e)
        __KISAN_MAPPER__ = None
        return None

def _pick_latest_official_yaml():
    """
    MAPPER 폴더에서 kisan_official*.yaml 중 '가장 최신'을 고른다.
    우선순위: v7.1 > v7 > v6 > ... > 기본 kisan_official.yaml
    """
    # 버전 패턴 파일 모두 탐색
    patterns = [
        str(MAPPER_DIR / "kisan_official.v*.yaml"),
        str(MAPPER_DIR / "kisan_official.yaml"),
    ]
    candidates = []
    for pat in patterns:
        candidates.extend(glob.glob(pat))

    if not candidates:
        return None

    def _version_key(p):
        # 예: kisan_official.v7.1.yaml → (7,1)
        name = Path(p).name
        import re
        m = re.search(r"\.v(\d+)(?:\.(\d+))?\.yaml$", name)
        if not m:
            # 버전 미표기 기본 파일은 가장 낮은 우선순위
            return (-1, -1)
        major = int(m.group(1))
        minor = int(m.group(2) or 0)
        return (major, minor)

    # 가장 높은 버전 우선으로 정렬
    candidates.sort(key=_version_key, reverse=True)
    return candidates[0]

def _load_official_schema():
    """공식 스키마(YAML) 로드 → dict (없으면 빈 dict)"""
    import yaml
    global __OFFICIAL_SCHEMA__, __OFFICIAL_SCHEMA_PATH__
    if __OFFICIAL_SCHEMA__ is not None:
        return __OFFICIAL_SCHEMA__
    ypath = _pick_latest_official_yaml()
    if not ypath:
        print("[MAPPER] no kisan_official*.yaml found")
        __OFFICIAL_SCHEMA__ = {}
        __OFFICIAL_SCHEMA_PATH__ = None
        return __OFFICIAL_SCHEMA__
    try:
        with open(ypath, "r", encoding="utf-8") as f:
            data = yaml.safe_load(f) or {}
        __OFFICIAL_SCHEMA__ = data
        __OFFICIAL_SCHEMA_PATH__ = ypath
        print(f"[MAPPER] official schema loaded: {Path(ypath).name}")
        return data
    except Exception as e:
        print("[MAPPER] official schema load failed:", e)
        __OFFICIAL_SCHEMA__ = {}
        __OFFICIAL_SCHEMA_PATH__ = None
        return __OFFICIAL_SCHEMA__

def get_mapper():
    """
    외부에서 사용하는 진입점.
    - kisan_mapping_v2 모듈 객체 또는 None
    - 모듈 내에서 필요한 경우 _load_official_schema()를 호출해 쓸 수 있음
    """
    return _import_kisan_mapper()

def get_official_schema():
    """외부에서 사용하는 진입점: 공식 스키마(dict). 없으면 {}"""
    return _load_official_schema()

def mapper_sanity_check(verbose: bool = True) -> dict:
    """
    맵퍼/스키마 준비상태 점검.
    반환 예: {"mapper": True, "schema": True, "yaml": "kisan_official.v7.1.yaml"}
    """
    m = get_mapper()
    s = get_official_schema()
    ok = {
        "mapper": bool(m),
        "schema": bool(s),
        "yaml": (Path(__OFFICIAL_SCHEMA_PATH__).name if __OFFICIAL_SCHEMA_PATH__ else None),
    }
    if verbose:
        print(f"[MAPPER] ready: mapper={ok['mapper']} schema={ok['schema']} yaml={ok['yaml']}")
    return ok

# 모듈 로드만 미리 시도(있으면 준비 로그 출력)
try:
    mapper_sanity_check(verbose=True)
except Exception as _e:
    print("[MAPPER] sanity check error:", _e)


# === IMAP 서버 고정값 (네이버웍스/네이버 공용)
IMAP_HOST = "imap.naver.com"
IMAP_PORT = 993  # SSL

def save_json(path, data):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def load_json(path, default=None):
    if not os.path.exists(path): return default
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def now_tag(): return time.strftime("%y%m%d_%H%M%S")
# system_estimate_ui.py
def slug(s):
    import re
    s = str(s or "")
    # 한글/영문/숫자와 공백/하이픈/언더바만 남김
    s = re.sub(r"[^0-9a-zA-Z가-힣 _-]", "", s)
    s = re.sub(r"\s+", "_", s.strip())
    return s


# -------------------------
# OCR/문서 파서
# -------------------------
AMP_CHOICES = [15,20,30,40,50,60,75,100,125,150,175,200,225,250,300,350,400,500,600,630,700,800]
def _norm(t): return re.sub(r"[ \t]+"," ",t or "").replace("\r","\n")
def _extract_email(t):
    m=re.search(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", t or "")
    return m.group(0) if m else ""
def _extract_phone(t):
    m=re.search(r"(01[0-9]-?\d{3,4}-?\d{4}|0\d{1,2}-?\d{3,4}-?\d{4})", t or "")
    return m.group(0) if m else ""
def _extract_company(t):
    for key in ["업체명","회사","상호","법인명","거래처"]:
        m=re.search(key+r"\s*[:：]\s*([^\n]+)", t or "")
        if m:
            name=re.sub(r"[^가-힣A-Za-z0-9()_\- ]+","",m.group(1)).strip()
            return name[:60]
    mail=_extract_email(t or "")
    if mail:
        guess=re.sub(r"[\._\-0-9]+$","", mail.split("@")[0])
        if guess: return guess[:60]
    return ""

# === 안전한 OCR 함수 (파일에 딱 1개만 존재!) ===
def _ocr_image(path, tesseract_path=""):
    try:
        from PIL import Image
        import pytesseract
    except Exception as e:
        _show_error("OCR 모듈 없음", str(e))
        return ""

    try:
        if tesseract_path:
            pytesseract.pytesseract.tesseract_cmd = tesseract_path
    except Exception as e:
        _show_error("Tesseract 경로 설정 실패", str(e))
        return ""

    try:
        with Image.open(path) as im:
            img = im.copy()
        text = pytesseract.image_to_string(img, lang="kor+eng")
        return text
    except Exception as e:
        _show_error("OCR 실패", str(e))
        return ""
# === /끝 ===


def _pdf_to_text(path, tesseract_path=""):
    texts=[]
    if _PDFPLUMBER_OK:
        try:
            with pdfplumber.open(path) as pdf:
                for p in pdf.pages:
                    t=p.extract_text() or ""
                    if t.strip(): texts.append(t)
        except: pass
    joined="\n".join(texts)
    if joined.strip(): return joined
    if _PDF2IMG_OK and _PIL_OK and _TESS_OK:
        try:
            if tesseract_path and os.path.exists(tesseract_path):
                pytesseract.pytesseract.tesseract_cmd=tesseract_path
            imgs=convert_from_path(path, dpi=200)
            buf=[]
            for im in imgs[:10]:
                buf.append(pytesseract.image_to_string(im, lang="kor+eng"))
            return "\n".join(buf)
        except: return ""
    return ""

def _scan_specs(text):
    text = text or ""
    lines=[s.strip() for s in text.splitlines() if s.strip()]
    specs=[]
    for ln in lines:
        kind="MCCB" if re.search(r"\bMCCB\b", ln, re.I) else ("ELB" if (re.search(r"\bELB\b", ln, re.I) or "누전" in ln) else None)
        poles=None; amp=None; qty=1
        m=re.search(r"\b([234]P)\b", ln, re.I);     poles=m.group(1).upper() if m else None
        m2=re.search(r"(\d{2,3})\s*A\b", ln, re.I)
        if m2:
            av=safe_int(m2.group(1)); amp=f"{min(AMP_CHOICES,key=lambda x:abs(x-av))}A"
        mqty=re.search(r"(?:x|X|수량\s*)(\d+)\b|(\d+)\s*EA\b", ln, re.I)
        if mqty: qty=safe_int([g for g in mqty.groups() if g][0],1)
        if kind and poles and amp:
            specs.append({"종류":kind,"극수":poles,"용량":amp,"수량":str(qty)})
    # 2-line 조합
    for i in range(len(lines)-1):
        ln1,ln2=lines[i],lines[i+1]
        kind="MCCB" if (re.search(r"\bMCCB\b",ln1+ln2,re.I)) else ("ELB" if (re.search(r"\bELB\b",ln1+ln2,re.I) or "누전" in (ln1+ln2)) else None)
        poles=None; amp=None; qty=1
        m=re.search(r"\b([234]P)\b", ln1+" "+ln2, re.I); poles=m.group(1).upper() if m else None
        m2=re.search(r"(\d{2,3})\s*A\b", ln1+" "+ln2, re.I)
        if m2:
            av=safe_int(m2.group(1)); amp=f"{min(AMP_CHOICES,key=lambda x:abs(x-av))}A"
        mqty=re.search(r"(?:x|X|수량\s*)(\d+)\b|(\d+)\s*EA\b", ln1+" "+ln2, re.I)
        if mqty: qty=safe_int([g for g in mqty.groups() if g][0],1)
        if kind and poles and amp:
            specs.append({"종류":kind,"극수":poles,"용량":amp,"수량":str(qty)})
    return specs

def _split_main_branch(specs, text):
    text = text or ""
    for s in specs:
        patt=rf"{s['극수']}.*{s['용량']}"
        if re.search(r"메인|MAIN|주차단기", text, re.I) and re.search(patt, text, re.I):
            return s, [x for x in specs if x is not s]
    def score(sp):
        amp=safe_int(sp['용량'].replace("A",""),0)
        pole_rank={"4P":3,"3P":2,"2P":1}.get(sp['극수'],0)
        return (amp, pole_rank)
    if specs:
        main=sorted(specs, key=score, reverse=True)[0]
        b=specs.copy(); b.remove(main)
        return main, b
    return None, []

def extract_info_from_files(paths, tesseract_path=""):
    """
    첨부 분석 → 탭1 자동 채움:
      - 이미지/PDF/DOCX/TXT 기존 처리 + XLSX 직파싱 추가
      - XLSX가 하나라도 들어오면: 구조화(main/branches/client) 우선 사용,
        그 외 파일 텍스트는 full_text에 합쳐서 RAG 근거로만 활용
    반환: (client, main, branches, full_text)
    """
    texts = []
    xlsx_client = None
    xlsx_main = None
    xlsx_branches = None
    xlsx_text = ""

    # ========== 보조 유틸 ==========
    def _digits(s) -> str:
        return "".join(ch for ch in str(s) if ch.isdigit())

    AMP_TARGETS = [15,20,25,30,40,50,60,75,100,125,150,175,200,225,250,300,350,400,500,630,800]

    def _parse_amp(val, default=None):
        d = _digits(val)
        if not d:
            return default
        n = int(d)
        return min(AMP_TARGETS, key=lambda t: abs(t - n))

    def _parse_poles(val, default=None):
        s = str(val).strip().upper()
        if not s:
            return default
        d = _digits(s)
        if not d:
            return default
        return f"{int(d)}P"

    def _normlower(s: str) -> str:
        return (s or "").strip().lower().replace(" ", "")

    # 병합셀 읽을 때 좌상단 값으로 평탄화
    def _flatten_cell(ws, r, c):
        cell = ws.cell(row=r, column=c)
        if _OPENPYXL_OK and MergedCell is not None and isinstance(cell, MergedCell):
            for rng in ws.merged_cells.ranges:
                if (rng.min_row <= r <= rng.max_row) and (rng.min_col <= c <= rng.max_col):
                    return ws.cell(row=rng.min_row, column=rng.min_col).value
        return cell.value

    def _sheet_to_rows(ws):
        rows = []
        for r in range(1, ws.max_row + 1):
            row = []
            for c in range(1, ws.max_column + 1):
                row.append(_flatten_cell(ws, r, c))
            # 뒤쪽 완전 공백 제거
            while row and (row[-1] in (None, "", " ")):
                row.pop()
            rows.append(row)
        # 완전 빈줄 제거
        while rows and not any(v not in (None, "", " ") for v in rows[-1]):
            rows.pop()
        return rows

    # 헤더 자동 매핑
    def _find_header_indexes(header_vals):
        def find_any(aliases):
            al = [_normlower(a) for a in aliases]
            for idx, v in enumerate(header_vals, start=1):
                if isinstance(v, str) and _normlower(v) in al:
                    return idx
            return None
        return {
            "client": find_any(["업체명","거래처","회사","상호","customer","client"]),
            "phone":  find_any(["연락처","전화","tel","phone"]),
            "email":  find_any(["이메일","email","mail"]),
            "item":   find_any(["품명","항목","item","product","name"]),
            "spec":   find_any(["규격","규격/설명","spec","description","desc"]),
            "poles":  find_any(["극수","poles","pole"]),
            "amp":    find_any(["용량","전류","amp","current","정격"]),
            "qty":    find_any(["수량","qty","quantity","ea","개수"]),
            "remark": find_any(["비고","remark","메모","memo"]),
        }

    def _guess_kind(txt: str) -> str | None:
        t = (txt or "").upper()
        if "ELB" in t or "ELCB" in t or "누전" in t:
            return "ELB"
        if "MCCB" in t or "NFB" in t or "VCB" in t:
            return "MCCB"
        return None

    def _select_main_and_branches(items):
        """
        items 원소 예:
          {"품명":..,"규격":..,"극수":..,"용량":..,"수량":..,"remark":..,"종류":..}
        규칙:
          1) '메인' 키워드(품명/비고)에 있으면 우선
          2) 없으면 (용량 근사값, 극수랭크) 최댓값을 메인
        """
        if not items:
            return {"종류":"","극수":"","용량":"","수량":1}, []

        def score(it):
            amp = _parse_amp(it.get("용량"))
            poles = int(_digits(it.get("극수") or "") or 0)
            return (amp or 0, poles)

        mains = [it for it in items if ("메인" in str(it.get("품명","")) or
                                        "MAIN" in str(it.get("품명","")).upper() or
                                        "메인" in str(it.get("remark","")))]
        if not mains:
            mains = items[:]
        main_item = max(mains, key=score)

        branches = []
        for it in items:
            if it is main_item:
                continue
            branches.append({
                "종류": it.get("종류") or (_guess_kind(it.get("품명")) or _guess_kind(it.get("규격")) or "MCCB"),
                "극수": _parse_poles(it.get("극수")) or "4P",
                "용량": f"{_parse_amp(it.get('용량') or 0) or ''}A",
                "수량": int(_digits(it.get("수량") or "1") or 1),
            })

        main = {
            "종류": main_item.get("종류") or (_guess_kind(main_item.get("품명")) or _guess_kind(main_item.get("규격")) or "MCCB"),
            "극수": _parse_poles(main_item.get("극수")) or "4P",
            "용량": f"{_parse_amp(main_item.get('용량') or 0) or ''}A",
            "수량": int(_digits(main_item.get("수량") or "1") or 1),
        }
        return main, branches

    # ========== 파일 순회 ==========
    for p in paths:
        ext = os.path.splitext(p)[1].lower()
        t = ""

        if ext in [".png",".jpg",".jpeg",".bmp",".tif",".tiff"]:
            t = _ocr_image(p, tesseract_path)

        elif ext == ".pdf":
            t = _pdf_to_text(p, tesseract_path)

        elif ext == ".docx" and _DOCX_OK:
            try:
                doc = docx.Document(p)
                t = "\n".join([pa.text for pa in doc.paragraphs])
            except:
                t = ""

        elif ext == ".xlsx":
            # === 엑셀 구조 파싱 ===
            if not _OPENPYXL_OK:
                texts.append("[xlsx 파싱 실패: openpyxl 미설치]")
            else:
                try:
                    wb = load_workbook(p, data_only=True)
                    items = []
                    xlsx_text_chunks = []
                    # 기본값: 엑셀에서 고객정보 헤더가 없을 수도 있으니 우선 빈값
                    x_client = {"업체명":"", "연락처":"", "이메일":""}

                    for ws in wb.worksheets:
                        rows = _sheet_to_rows(ws)
                        # full_text 구성
                        for rv in rows:
                            if not rv:
                                continue
                            xlsx_text_chunks.append("\t".join(str(x) for x in rv if x not in (None,""," ")))

                        # 헤더 후보(상위 10행)에서 매핑
                        hdr_row = None
                        hdr_idx = None
                        for i in range(min(10, len(rows))):
                            vals = rows[i]
                            if not vals:
                                continue
                            idx = _find_header_indexes([str(v) if v is not None else "" for v in vals])
                            # 한 개라도 매칭되면 헤더로 잡고 진행
                            if any(idx.values()):
                                hdr_row, hdr_idx = i, idx
                                break
                        if hdr_row is None:
                            continue

                        # 고객 정보(헤더행에 client/phone/email 열이 있으면 1행 아래에서 한 줄만 훑어본다)
                        if hdr_idx.get("client"):
                            rr = hdr_row + 1
                            if rr < len(rows):
                                row = rows[rr]
                                def _get(ci):
                                    return (row[ci-1] if ci and (ci-1) < len(row) else "")
                                c = str(_get(hdr_idx.get("client")) or "").strip()
                                if c: x_client["업체명"] = c
                                ph = str(_get(hdr_idx.get("phone")) or "").strip()
                                if ph: x_client["연락처"] = ph
                                em = str(_get(hdr_idx.get("email")) or "").strip()
                                if em: x_client["이메일"] = em

                        # 데이터 영역: 품목 테이블
                        for r in range(hdr_row + 1, len(rows)):
                            row = rows[r]
                            if not row or not any(row):
                                continue
                            def getcol(key):
                                ci = hdr_idx.get(key)
                                if not ci:
                                    return ""
                                return row[ci-1] if (ci-1) < len(row) else ""

                            item_txt = str(getcol("item") or "")
                            spec_txt = str(getcol("spec") or "")
                            poles    = getcol("poles")
                            amp      = getcol("amp")
                            qty      = getcol("qty")
                            remark   = str(getcol("remark") or "")

                            nm = _normlower(item_txt)
                            if nm in ("합계","total","sum"):
                                continue
                            if not any([item_txt, spec_txt, poles, amp, qty, remark]):
                                continue

                            k_guess = _guess_kind(item_txt) or _guess_kind(spec_txt) or _guess_kind(remark)
                            items.append({
                                "품명": item_txt or spec_txt,
                                "규격": spec_txt,
                                "극수": poles,
                                "용량": amp,
                                "수량": qty or 1,
                                "remark": remark,
                                "종류": k_guess or "",
                            })

                    # 엑셀에서 뽑은 것 보관
                    xlsx_client = x_client
                    xlsx_main, xlsx_branches = _select_main_and_branches(items)
                    xlsx_text = "\n".join(xlsx_text_chunks)

                except Exception as e:
                    texts.append(f"[xlsx 파싱 오류] {e}")

        else:
            # 일반 텍스트
            try:
                with open(p, "r", encoding="utf-8") as f:
                    t = f.read()
            except:
                try:
                    with open(p, "r", encoding="cp949") as f:
                        t = f.read()
                except:
                    t = os.path.basename(p)

        if t:
            texts.append(t)

    # full_text(엑셀에서 모은 텍스트도 함께)
    full = _norm("\n".join(texts + ([xlsx_text] if xlsx_text else [])))

    # ===== 우선순위: 엑셀 구조가 있으면 그걸 사용 =====
    if xlsx_main or xlsx_branches:
        client = xlsx_client or {"업체명":"","연락처":"","이메일":""}
        main = xlsx_main or {"종류":"","극수":"","용량":"","수량":1}
        branches = xlsx_branches or []
        return client, main, branches, full

    # ===== 엑셀이 없거나 실패 → 기존 텍스트 파이프라인 =====
    client = {"업체명": _extract_company(full), "연락처": _extract_phone(full), "이메일": _extract_email(full)}
    specs  = _scan_specs(full)
    main, branches = _split_main_branch(specs, full)
    return client, main, branches, full


# -------------------------
# Tk Base(드래그앤드롭 지원 여부)
# -------------------------
BaseTk = TkinterDnD.Tk if _DND_OK else tk.Tk

def _dh(s):
    parts = decode_header(s or "")
    out=[]
    for txt, enc in parts:
        if isinstance(txt, bytes):
            out.append(txt.decode(enc or "utf-8", errors="ignore"))
        else:
            out.append(txt)
    return "".join(out).strip()

def _msg_summary(msg):
    subj=_dh(msg.get("Subject",""))
    frm=_dh(msg.get("From",""))
    try:
        dt=parsedate_to_datetime(msg.get("Date"))
        date_s = dt.strftime("%Y-%m-%d %H:%M")
    except:
        date_s = msg.get("Date","")
    atts=0
    for part in msg.walk():
        if (part.get_content_disposition() or "").lower() == "attachment":
            atts+=1
    return date_s, frm, subj, atts


class App(BaseTk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1350x900")
        self.cfg=load_json(CFG_PATH, default={
            "imap": {"host":"imap.naver.com","port":993,"user":"","pass":"","ssl":True,"ok":False},
            "smtp":{"host":"","port":587,"user":"","pass":"","sender":""},
            "naver_works":{"id":"","api_key":""},
            "chatgpt":{"url":"https://api.openai.com/v1/chat/completions","api_key":"","ok":False,"model":"gpt-4o"},
            "claude":{"url":"https://api.anthropic.com/v1/messages","api_key":"","ok":False,"model":"claude-3-7-sonnet-20250219"},
            "tesseract_path":""
            })

        self._init_state()
        self.agent = AssistantAgent(base_dir=BASE_DIR, data_dir=DATA_DIR)

        self.style = ttk.Style()
        try: self.style.theme_use("clam")
        except: pass

        # === 상단 공용 툴바 ===
        topbar = ttk.Frame(self); topbar.pack(fill="x", pady=(4,2))
        ttk.Button(topbar, text="RAG 검색",   command=_do_rag_query).pack(side="left", padx=4)
        ttk.Button(topbar, text="RAG 재색인", command=_do_rag_reindex).pack(side="left", padx=4)

        # 항상 보이는 RAG 검색 입력창 + Enter/버튼 실행 + 단축키(Ctrl+K)
        self.var_rag_query_top = tk.StringVar(value="")
        self.ent_rag_query_top = ttk.Entry(topbar, textvariable=self.var_rag_query_top, width=24)
        self.ent_rag_query_top.pack(side="left", padx=4)
        self.ent_rag_query_top.bind("<Return>", lambda _e: self._rag_search_from_entry())

        # 전역 단축키: Ctrl+K → 상단 입력창 포커스
        self.bind_all("<Control-k>", lambda e: (self._focus_rag_entry(), "break"))
        self.bind_all("<Control-K>", lambda e: (self._focus_rag_entry(), "break"))
        for _cls in ("Text", "TEntry", "Treeview", "Listbox"):
            self.bind_class(_cls, "<Control-k>", lambda e: (self._focus_rag_entry(), "break"))
            self.bind_class(_cls, "<Control-K>", lambda e: (self._focus_rag_entry(), "break"))

        ttk.Button(topbar, text="검색 실행", command=self._rag_search_from_entry).pack(side="left", padx=4)
        ttk.Label(topbar, text="※ 전체 자료 색인을 다시 구축합니다").pack(side="left", padx=6)




        self.nb = ttk.Notebook(self); self.nb.pack(fill="both", expand=True)
        self.protocol("WM_DELETE_WINDOW", self._on_close)
        self._build_tab_estimate()
        self._build_tab_ai()
        self._build_tab_email()
        self._build_tab_repo()
        self._build_tab_prices()
        self._build_tab_clients()
        self._build_tab_shipping()
        self._build_tab_reports()
        self._build_tab_settings()
        self.nb.select(0)

    def _on_close(self):
        """
        창 닫기(X) 또는 종료 시도 시 호출되는 안전 종료 핸들러.
        - 예약된 after() 콜백, 백그라운드 작업, 임시 창 등을 정리한 뒤,
        최종적으로 self.destroy()로 모든 Tk 리소스를 제거합니다.
        """
        try:
            # 1) after() 예약 작업 취소 (우리가 예약 ID들을 리스트에 모아두는 경우)
            if hasattr(self, "_after_ids"):
                for aid in list(getattr(self, "_after_ids") or []):
                    try: self.after_cancel(aid)
                    except Exception: pass

            # 2) 백그라운드 스레드/워커 정지 (있다면)
            if hasattr(self, "_bg_workers"):
                for w in list(getattr(self, "_bg_workers") or []):
                    try:
                        if hasattr(w, "stop"): w.stop()
                    except Exception: pass
                    try:
                        if hasattr(w, "join"): w.join(timeout=0.5)
                    except Exception: pass

            # 3) 떠 있는 보조 창(Toplevel) 정리 (있다면)
            try:
                import tkinter as tk
                for child in list(self.winfo_children()):
                    try:
                        if isinstance(child, tk.Toplevel):
                            child.destroy()
                    except Exception:
                        pass
            except Exception:
                pass

        except Exception:
            # 종료 중 에러는 조용히 무시(사용자 경험 보호)
            pass
        finally:
            try:
                # Tk 리소스 파괴 -> mainloop() 반환
                self.destroy()
            except Exception:
                pass

    def _init_state(self):
        self.branches=[]
        self.accessories=[]
        self.lines=[]
        self.last_enclosure=None
        self.last_main=None
        self.last_client=None
        # 메일 캐시
        self.mail_cache = {}

    # ========= [탭1] 견적 =========
    def _build_tab_estimate(self):
        tab = ttk.Frame(self.nb, padding=8); self.nb.add(tab, text="견적")
        # (추가) 계산 근거 팝업 버튼
        self.btn_explain = ttk.Button(tab, text="왜 이 사이즈?", command=lambda s=self: _show_enclosure_explain(s))
        self.btn_explain.pack(anchor="ne", padx=8, pady=4)
        try:
            self.btn_explain.state(['disabled'])  # 계산 전에는 비활성화
        except Exception:
            pass
        left = ttk.Frame(tab); left.pack(side="left", fill="y")

        # 고객
        g0=ttk.LabelFrame(left, text="고객 정보", padding=8); g0.pack(fill="x", pady=(0,8))
        ttk.Label(g0, text="업체명").grid(row=0,column=0,sticky="w")
        ttk.Label(g0, text="연락처").grid(row=1,column=0,sticky="w")
        ttk.Label(g0, text="이메일").grid(row=2,column=0,sticky="w")
        self.ent_client_name=ttk.Entry(g0,width=26); self.ent_client_name.grid(row=0,column=1,padx=4)
        self.ent_client_phone=ttk.Entry(g0,width=26); self.ent_client_phone.grid(row=1,column=1,padx=4)
        self.ent_client_email=ttk.Entry(g0,width=26); self.ent_client_email.grid(row=2,column=1,padx=4)
        ttk.Button(g0,text="이메일 확인 후 불러오기", command=self._fill_client_from_selected_email).grid(row=0,column=2,rowspan=3,padx=(8,0))

        # 외함 (개선)
        g1=ttk.LabelFrame(left, text="외함", padding=8); g1.pack(fill="x", pady=(0,8))
        ttk.Label(g1,text="설치").grid(row=0,column=0,sticky="e")
        self.cbo_enc_place=ttk.Combobox(g1, values=["옥내","옥외","옥내자립","옥외자립","전주부착형"],
                                        width=10, state="readonly"); self.cbo_enc_place.set("옥내")
        self.cbo_enc_place.grid(row=0,column=1,padx=4,sticky="w")

        ttk.Label(g1,text="함종류").grid(row=1,column=0,sticky="e")
        self.cbo_enc_kind=ttk.Combobox(g1, values=["기성함","주문제작함","계량기함","CT계량기함","FRP 박스","하이박스"],
                                       width=12, state="readonly"); self.cbo_enc_kind.set("기성함")
        self.cbo_enc_kind.grid(row=1,column=1,padx=4,sticky="w")

        ttk.Label(g1,text="재질").grid(row=2,column=0,sticky="e")
        self.cbo_enc_mat=ttk.Combobox(g1, values=["STEEL 1.0T","STEEL 1.6T","STEEL 2.0T","SUS201 1.0T","SUS201 1.2T","SUS201 1.5T","SUS304 1.2T","SUS304 1.5T","SUS304 2.0T"],
                                      width=16, state="readonly"); self.cbo_enc_mat.set("STEEL 1.6T")
        self.cbo_enc_mat.grid(row=2,column=1,padx=4,sticky="w")

        ttk.Label(g1,text="기타요청").grid(row=3,column=0,sticky="e")
        self.ent_enc_misc=ttk.Entry(g1, width=26)
        self.ent_enc_misc.grid(row=3,column=1,padx=4,sticky="w")

        ttk.Label(g1,text="주문제작 단가").grid(row=4,column=0,sticky="e")
        self.ent_enc_custom_price=ttk.Entry(g1, width=12)
        self.ent_enc_custom_price.grid(row=4,column=1,padx=4,sticky="w")
        ttk.Label(g1,text="(함종류가 '주문제작함'일 때만 적용)").grid(row=4,column=2,sticky="w")

        self.lbl_base_auto = ttk.Label(g1, text="※ 자립형(옥내/옥외자립)은 베이스 포함으로 자동 처리")
        self.lbl_base_auto.grid(row=5, column=0, columnspan=3, sticky="w", pady=(4,0))

        # 메인
        POLES=["2P","3P","4P"]
        AMPS=["15A","20A","30A","40A","50A","60A","75A","100A","125A","150A","175A","200A","225A","250A","300A","350A","400A","500A","600A","630A","700A","800A"]
        g2=ttk.LabelFrame(left, text="메인 차단기", padding=8); g2.pack(fill="x", pady=(0,8))
        ttk.Label(g2,text="종류").grid(row=0,column=0,sticky="w")
        ttk.Label(g2,text="극수").grid(row=1,column=0,sticky="w")
        ttk.Label(g2,text="용량").grid(row=2,column=0,sticky="w")
        ttk.Label(g2,text="수량").grid(row=3,column=0,sticky="w")
        self.cbo_main_kind =ttk.Combobox(g2, values=["MCCB","ELB"], state="readonly", width=8); self.cbo_main_kind.set("MCCB"); self.cbo_main_kind.grid(row=0,column=1,padx=4)
        self.cbo_main_poles=ttk.Combobox(g2, values=POLES, state="readonly", width=8); self.cbo_main_poles.set("4P"); self.cbo_main_poles.grid(row=1,column=1,padx=4)
        self.cbo_main_amp  =ttk.Combobox(g2, values=AMPS, state="readonly", width=8); self.cbo_main_amp.set("100A"); self.cbo_main_amp.grid(row=2,column=1,padx=4)
        self.ent_main_qty  =ttk.Entry(g2, width=10); self.ent_main_qty.insert(0,"1"); self.ent_main_qty.grid(row=3,column=1,padx=4)

        # --- (추가) 메인 차단기 브랜드 선택 ---
        ttk.Label(g2, text="브랜드").grid(row=0, column=2, sticky="w")
        self.cbo_main_brand = ttk.Combobox(
            g2,
            values=["상도", "LS", "대륙", "비츠로", "현대"],
            state="readonly",
            width=8
        )
        self.cbo_main_brand.set("상도")
        self.cbo_main_brand.grid(row=0, column=3, padx=4)
        # --- /브랜드 선택 끝 ---

        # --- 여기서 SPD 체크박스용 상태 변수 초기화 ---
        self.var_spd = tk.BooleanVar(value=False)
        ttk.Checkbutton(g2, text="SPD 포함", variable=self.var_spd).grid(row=4, column=0, columnspan=2, sticky="w")
        # ------------------------------------------------

        # 분기
        g3=ttk.LabelFrame(left, text="분기 차단기", padding=8); g3.pack(fill="x", pady=(0,8))
        self.cbo_b_kind =ttk.Combobox(g3, values=["MCCB","ELB"], state="readonly", width=8); self.cbo_b_kind.set("ELB"); self.cbo_b_kind.grid(row=0,column=0,padx=2)
        self.cbo_b_poles=ttk.Combobox(g3, values=POLES, state="readonly", width=8); self.cbo_b_poles.set("2P"); self.cbo_b_poles.grid(row=0,column=1,padx=2)
        self.cbo_b_amp  =ttk.Combobox(g3, values=AMPS, state="readonly", width=8); self.cbo_b_amp.set("30A"); self.cbo_b_amp.grid(row=0,column=2,padx=2)
        self.ent_b_qty  =ttk.Entry(g3, width=6); self.ent_b_qty.insert(0,"4"); self.ent_b_qty.grid(row=0,column=3,padx=2)
        ttk.Button(g3, text="분기 추가", command=self._add_branch).grid(row=0,column=4,padx=4)
        ttk.Button(g3, text="선택 삭제", command=self._remove_branch).grid(row=0,column=5,padx=2)
        self.lst_branches=tk.Listbox(g3, height=8, width=45); self.lst_branches.grid(row=1,column=0,columnspan=6,sticky="we",pady=(6,0))

        # 부속자재
        gAcc=ttk.LabelFrame(left, text="부속자재", padding=8); gAcc.pack(fill="x", pady=(0,8))
        MAGNETS=["MC-22","MC-32","MC-40","MC-50","MC-65","MC-75","MC-80"]
        ttk.Label(gAcc,text="마그네트").grid(row=0,column=0,sticky="w")
        self.cbo_mag=ttk.Combobox(gAcc, values=MAGNETS, width=12, state="readonly"); self.cbo_mag.grid(row=0,column=1,padx=2)
        self.ent_mag_qty=ttk.Entry(gAcc, width=6); self.ent_mag_qty.insert(0,"1"); self.ent_mag_qty.grid(row=0,column=2,padx=2)
        ttk.Button(gAcc,text="추가",command=self._add_acc_mag).grid(row=0,column=3,padx=4)

        CAP_VOLT=["220V","380V"]; CAP_SIZE=["10","15","20","30","40","50","60","75","100","150","175","200","250","300","400","500","1000"]
        ttk.Label(gAcc,text="콘덴서").grid(row=1,column=0,sticky="w")
        self.cbo_cap_v=ttk.Combobox(gAcc, values=CAP_VOLT, width=6, state="readonly"); self.cbo_cap_v.grid(row=1,column=1,padx=2,sticky="w")
        self.cbo_cap_s=ttk.Combobox(gAcc, values=CAP_SIZE, width=6, state="readonly"); self.cbo_cap_s.grid(row=1,column=2,padx=2,sticky="w")
        self.ent_cap_qty=ttk.Entry(gAcc, width=6); self.ent_cap_qty.insert(0,"1"); self.ent_cap_qty.grid(row=1,column=3,padx=2)
        ttk.Button(gAcc,text="추가",command=self._add_acc_cap).grid(row=1,column=4,padx=4)

        METERS=["V/A-meter","3상계량기(피에스텍)","단상계량기(피에스텍)","CT계량기(피에스텍)","3상계량기(LS)","단상계량기(LS)","CT계량기(LS)"]
        ttk.Label(gAcc,text="계측기").grid(row=2,column=0,sticky="w")
        self.cbo_meter=ttk.Combobox(gAcc, values=METERS, width=24, state="readonly"); self.cbo_meter.grid(row=2,column=1,columnspan=2,padx=2,sticky="w")
        self.ent_meter_qty=ttk.Entry(gAcc, width=6); self.ent_meter_qty.insert(0,"1"); self.ent_meter_qty.grid(row=2,column=3,padx=2)
        ttk.Button(gAcc,text="추가",command=self._add_acc_meter).grid(row=2,column=4,padx=4)

        ttk.Label(gAcc,text="기타부자재").grid(row=3,column=0,sticky="w")
        self.ent_misc_name=ttk.Entry(gAcc, width=24); self.ent_misc_name.grid(row=3,column=1,columnspan=2,padx=2,sticky="w")
        self.ent_misc_qty=ttk.Entry(gAcc, width=6); self.ent_misc_qty.insert(0,"1"); self.ent_misc_qty.grid(row=3,column=3,padx=2)
        ttk.Button(gAcc,text="추가",command=self._add_acc_misc).grid(row=3,column=4,padx=4)

        self.lst_acc=tk.Listbox(gAcc, height=7, width=52); self.lst_acc.grid(row=4,column=0,columnspan=5,sticky="we",pady=(6,2))
        ttk.Button(gAcc, text="선택 삭제", command=self._del_acc).grid(row=5,column=0,pady=(2,0))
        ttk.Button(gAcc, text="전체 초기화", command=self._clear_acc).grid(row=5,column=1,pady=(2,0))

        # 가운데 버튼
        mid=ttk.Frame(tab,padding=8); mid.pack(side="left", fill="y")
        ttk.Button(mid,text="시스템 견적 생성", command=self._make_system_estimate).pack(fill="x",pady=(0,6))
        ttk.Button(mid,text="AI 견적 생성", command=self._make_ai_estimate).pack(fill="x",pady=(0,6))
        ttk.Button(mid,text="화면 지우기", command=self._clear_output).pack(fill="x")

        # 우측 출력(표 + 합계)
        right=ttk.Frame(tab,padding=8); right.pack(side="left", fill="both", expand=True)

        sumbar = ttk.Frame(right); sumbar.pack(fill="x")
        self.var_est_title = tk.StringVar(value="(견적 없음)")
        ttk.Label(sumbar, textvariable=self.var_est_title).pack(side="left")
        self.var_est_total = tk.StringVar(value="합계: 0 원")
        ttk.Label(sumbar, textvariable=self.var_est_total).pack(side="right")

        cols=("no","name","spec","unit","qty","price","amount")
        self.grid_est = ttk.Treeview(right, columns=cols, show="headings", height=24)
        headers = ["No","품명","규격","단위","수량","단가","금액"]
        widths  = [50,260,300,60,70,120,140]
        for c,h,w in zip(cols,headers,widths):
            self.grid_est.heading(c, text=h)
            self.grid_est.column(c, width=w, anchor=("e" if c in ("qty","price","amount","no") else "w"))
        self.grid_est.pack(fill="both", expand=True, pady=(6,0))

        btns = ttk.Frame(right); btns.pack(fill="x", pady=(6,0))
        ttk.Button(btns, text="엑셀 저장(회사 양식)", command=self._save_to_company_excel).pack(side="left")
        ttk.Button(btns, text="CSV 저장", command=self._save_current_table_csv).pack(side="left", padx=6)
        ttk.Button(btns, text="미니 리포트", command=self._on_show_mini_report).pack(side="left", padx=6)
        ttk.Button(btns, text="미해결 티켓", command=self._on_show_dlq_list).pack(side="left")
        
    def _append(self,msg):
        # (이제 텍스트 영역 사용 안 함) – 남겨두되 호출 안함
        pass

    def _clear_output(self):
        # 표 초기화
        for iid in self.grid_est.get_children():
            self.grid_est.delete(iid)
        self.var_est_title.set("(견적 없음)")
        self.var_est_total.set("합계: 0 원")

    def _fill_client_from_selected_email(self):
        # 이메일 탭에서 선택된 메일의 본문/헤더에서 업체명/연락처/이메일 추정
        try:
            sel = getattr(self, "tree_mail").selection()
        except Exception:
            sel = None
        if not sel:
            messagebox.showwarning("안내","이메일 탭에서 먼저 메일을 선택하세요.")
            return
        uid = sel[0]
        entry = self.mail_cache.get(uid)
        if not entry:
            messagebox.showwarning("안내","선택된 메일 본문을 먼저 열어주세요.")
            return
        msg = entry.get("msg")
        body = self._extract_best_body_text(msg)
        # 헤더 + 본문 통합 텍스트에서 추출
        header_txt = f"From: {_dh(msg.get('From',''))}\nSubject: {_dh(msg.get('Subject',''))}\n"
        full = header_txt + "\n" + (body or "")
        client={"업체명":_extract_company(full),"연락처":_extract_phone(full),"이메일":_extract_email(full)}
        if client.get("업체명"): self.ent_client_name.delete(0,tk.END); self.ent_client_name.insert(0,client["업체명"])
        if client.get("연락처"): self.ent_client_phone.delete(0,tk.END); self.ent_client_phone.insert(0,client["연락처"])
        if client.get("이메일"): self.ent_client_email.delete(0,tk.END); self.ent_client_email.insert(0,client["이메일"])
        messagebox.showinfo("완료","메일에서 고객 정보를 불러왔습니다.")

    # 분기/부자재
    def _add_branch(self):
        k=self.cbo_b_kind.get().strip(); p=self.cbo_b_poles.get().strip(); a=self.cbo_b_amp.get().strip(); q=self.ent_b_qty.get().strip() or "1"
        try:
            if int(q)<=0: raise ValueError
        except:
            messagebox.showwarning("수량 오류","1 이상 숫자"); return
        self.branches.append({"종류":k,"극수":p,"용량":a,"수량":q})
        self.lst_branches.insert(tk.END, f"{k} | {p} {a} | 수량 {q}")
        
    def _remove_branch(self):
        sel=self.lst_branches.curselection()
        if not sel: return
        idx=sel[0]; self.lst_branches.delete(idx); del self.branches[idx]
        
    def _push_acc(self,name,qty):
        try:
            q=int(str(qty).strip() or "1")
            if q<=0: raise ValueError
        except:
            messagebox.showwarning("수량 오류","1 이상 숫자"); return
        self.accessories.append({"name":name,"qty":q})
        self.lst_acc.insert(tk.END, f"{name} x {q}")
        
    def _add_acc_mag(self):
        n=self.cbo_mag.get().strip()
        if n: self._push_acc(f"마그네트 {n}", self.ent_mag_qty.get())
        
    def _add_acc_cap(self):
        v=self.cbo_cap_v.get().strip(); s=self.cbo_cap_s.get().strip()
        if v and s: self._push_acc(f"콘덴서 {v} {s}μF", self.ent_cap_qty.get())
        
    def _add_acc_meter(self):
        m=self.cbo_meter.get().strip()
        if m: self._push_acc(m, self.ent_meter_qty.get())
        
    def _add_acc_misc(self):
        t=self.ent_misc_name.get().strip()
        if not t: messagebox.showwarning("입력 필요","이름 입력"); return
        self._push_acc(t, self.ent_misc_qty.get()); self.ent_misc_name.delete(0,tk.END)
        
    def _del_acc(self):
        sel=self.lst_acc.curselection()
        if not sel: return
        idx=sel[0]; self.lst_acc.delete(idx); del self.accessories[idx]
        
    def _clear_acc(self):
        self.accessories.clear(); self.lst_acc.delete(0,tk.END)

    def _render_estimate_table(self, lines: list):
        for iid in self.grid_est.get_children():
            self.grid_est.delete(iid)
        total = 0
        for L in lines:
            try:
                amt = int(float(L.get("금액",0)))
            except:
                amt = 0
            total += amt
            self.grid_est.insert("", "end", values=(
                L.get("no",""),
                L.get("품명",""),
                L.get("규격",""),
                L.get("단위",""),
                L.get("수량",""),
                L.get("단가",""),
                L.get("금액",""),
            ))
        self.var_est_total.set(f"합계: {total:,} 원")

    def _render_estimate_table(self, lines: list):
        for iid in self.grid_est.get_children():
            self.grid_est.delete(iid)
        total = 0
        for L in lines:
            try:
                amt = int(float(L.get("금액",0)))
            except:
                amt = 0
            total += amt
            self.grid_est.insert("", "end", values=(
                L.get("no",""),
                L.get("품명",""),
                L.get("규격",""),
                L.get("단위",""),
                L.get("수량",""),
                L.get("단가",""),
                L.get("금액",""),
            ))
        self.var_est_total.set(f"합계: {total:,} 원")

    # [ADD] Explain → 미니 리포트 팝업
    # === KISAN-AI PATCH: 외함 선정 사유 미니 리포트 (복사/저장 지원) ===
    def _on_show_mini_report(self):
        """
        마지막 explain을 정규화하여 A/B/C/D/E/H/정책을 한눈에 보여주고,
        바로 복사/저장할 수 있는 전용 팝업을 띄운다.
        - 공식/엔진 로직에는 개입하지 않음(읽기만).
        """
        import os, time
        try:
            from estimate_engine import get_last_explain_safe, explain_to_mini_report
        except Exception as e:
            messagebox.showerror("미니 리포트", f"엔진 모듈 임포트 실패: {e}")
            return

        try:
            ex = get_last_explain_safe()
            text = explain_to_mini_report(ex)  # 카드형 평문
        except Exception as e:
            messagebox.showerror("미니 리포트", f"리포트 생성 오류: {e}")
            return

        # 전용 팝업
        win = tk.Toplevel(self)
        win.title("외함 사이즈 선정 사유 (Mini Report)")
        win.geometry("820x580")

        frm = ttk.Frame(win, padding=10)
        frm.pack(fill="both", expand=True)

        # 읽기 전용 텍스트 박스
        txt = scrolledtext.ScrolledText(frm, wrap="word")
        txt.pack(fill="both", expand=True)
        txt.insert("1.0", text)
        txt.configure(state="disabled")

        # 하단 버튼 바: 복사 / 저장 / 닫기
        bar = ttk.Frame(frm)
        bar.pack(fill="x", pady=(10,0))

        def _copy():
            try:
                self.clipboard_clear()
                self.clipboard_append(text)
                messagebox.showinfo("복사", "미니 리포트가 클립보드에 복사되었습니다.")
            except Exception as e:
                messagebox.showerror("복사 실패", str(e))

        def _save():
            try:
                out_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "out")
                os.makedirs(out_dir, exist_ok=True)
                ts = time.strftime("%Y%m%d-%H%M%S")
                path = os.path.join(out_dir, f"mini_report_{ts}.txt")
                with open(path, "w", encoding="utf-8") as f:
                    f.write(text)
                messagebox.showinfo("저장", f"저장 완료:\n{path}")
            except Exception as e:
                messagebox.showerror("저장 실패", str(e))

        ttk.Button(bar, text="복사", command=_copy).pack(side="left")
        ttk.Button(bar, text="저장", command=_save).pack(side="left", padx=(6,0))
        ttk.Button(bar, text="닫기", command=win.destroy).pack(side="right")

    # === KISAN-AI PATCH: DLQ 미해결 티켓 리스트 팝업 ===
    def _on_show_dlq_list(self):
        """
        out/dlq/dlq_index.json을 읽어 오픈 티켓 목록을 표로 표시.
        DLQ가 없거나 비어있으면 안내 메시지.
        """
        try:
            dlq_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "out", "dlq")
            index_path = os.path.join(dlq_dir, "dlq_index.json")
            if not os.path.exists(index_path):
                from tkinter import messagebox
                messagebox.showinfo("미해결 티켓", "오픈 티켓이 없습니다. (인덱스 파일 없음)")
                return
            with open(index_path, "r", encoding="utf-8") as f:
                idx = json.load(f)
            opened = (idx or {}).get("open") or {}
            if not opened:
                from tkinter import messagebox
                messagebox.showinfo("미해결 티켓", "오픈 티켓이 없습니다.")
                return

            # 팝업 윈도우
            win = tk.Toplevel(self)
            win.title("미해결 티켓 (DLQ)")
            frm = ttk.Frame(win, padding=10); frm.pack(fill="both", expand=True)

            tree = ttk.Treeview(frm, columns=("id","case","fail","path","error"), show="headings", height=12)
            for c, title, w in (
                ("id","ID",120),
                ("case","케이스",220),
                ("fail","실패누적",80),
                ("path","경로",420),
                ("error","오류요약",420),
            ):
                tree.heading(c, text=title); tree.column(c, width=w, anchor="w")
            # 데이터 주입
            for tid, meta in opened.items():
                fc = meta.get("fail_count", 1)
                case = meta.get("case","")
                path = meta.get("path","")
                err  = meta.get("error","")
                if isinstance(err, str) and len(err) > 120:
                    err = err[:120] + "…"
                tree.insert("", "end", values=(tid, case, fc, path, err))
            tree.pack(fill="both", expand=True)

            # 하단 버튼: 닫기
            bar = ttk.Frame(frm); bar.pack(fill="x", pady=(8,0))
            ttk.Button(bar, text="닫기", command=win.destroy).pack(side="right")

        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("미해결 티켓", f"표시 중 오류: {e}")


    def _build_and_show_estimate(self, mode="system"):
        from tkinter import messagebox
        # === UI 값 읽기 ===
        place = self.cbo_enc_place.get()
        kind  = self.cbo_enc_kind.get()
        mat   = self.cbo_enc_mat.get()
        misc  = self.ent_enc_misc.get().strip()
        custom_price = self.ent_enc_custom_price.get().strip()
        has_base = place in ("옥내자립", "옥외자립")
        # ...이하 기존 코드 계속...


    def _build_and_show_estimate(self, mode="system"):
        from tkinter import messagebox

        # === UI 값 읽기 ===
        place = self.cbo_enc_place.get()
        kind  = self.cbo_enc_kind.get()
        mat   = self.cbo_enc_mat.get()
        misc  = self.ent_enc_misc.get().strip()
        custom_price = self.ent_enc_custom_price.get().strip()
        has_base = place in ("옥내자립", "옥외자립")

        # 화면 표시에 쓰일 원본 외함 정보
        enclosure_ui = {
            "설치": place,
            "함종류": kind,
            "외함 재질": mat,
            "베이스 유무": ("있음" if has_base else "없음"),
            "기타요청": misc,
        }
        if kind == "주문제작함" and custom_price:
            enclosure_ui["주문제작_단가"] = custom_price

        # 메인 차단기
        main = {
            "종류": self.cbo_main_kind.get(),
            "극수": self.cbo_main_poles.get(),
            "용량": self.cbo_main_amp.get(),
            "브랜드": self.cbo_main_brand.get(),
            "수량": self.ent_main_qty.get().strip() or "1",
        }
        client = {
            "업체명": self.ent_client_name.get().strip(),
            "연락처": self.ent_client_phone.get().strip(),
            "이메일": self.ent_client_email.get().strip(),
        }

        try:
            # 외함 W/H/D 계산
            W, H, D = compute_enclosure_size(
                main["종류"], main["극수"], main["용량"], self.branches, style="경제형"
            )

            # ★ 엔진에 맞는 키로 재구성 (필수)
            enclosure_for_engine = {
                "외함 종류": enclosure_ui.get("함종류", "표준형"),
                "외함 재질": enclosure_ui.get("외함 재질", "철제"),
                "베이스 유무": enclosure_ui.get("베이스 유무", "없음"),
                "W": W, "H": H, "D": D,
            }

            # [AUTOLOG] 입력 로깅
            try:
                _log("user", {
                    "action": "run_estimate",
                    "mode": mode,
                    "client": client,
                    "enclosure": enclosure_for_engine,
                    "main": main,
                    "branches": self.branches,
                    "accessories": self.accessories,
                    "spd": self.var_spd.get() if hasattr(self, "var_spd") else None,
                })
            except Exception as _e:
                print("[AUTOLOG] input log fail:", _e)

            # 견적 라인 생성
            lines, _sz = build_estimate_lines(
                enclosure_for_engine,
                client,
                main,
                self.branches,
                self.accessories,
                self.var_spd.get(),
                style="경제형",
            )

        except Exception as e:
            messagebox.showerror("에러", f"견적 계산 실패:\n{e}")
            traceback.print_exc()
            return

        # 계산 근거 저장 + 버튼 활성화
        self._last_explain = get_last_explain()
        try:
            self.btn_explain.state(['!disabled'])
        except Exception:
            pass

        # 화면/저장용 병합
        self.lines = lines
        self.last_enclosure = {**enclosure_ui, "W": W, "H": H, "D": D}
        self.last_main = main
        self.last_client = client

        title = f"[{mode.upper()}] 외함 {W}x{H}x{D} | 메인 {main['종류']} {main['극수']} {main['용량']} x {main['수량']} | 분기 {len(self.branches)}"
        self.var_est_title.set(title)
        self._render_estimate_table(lines)

        data = {
            "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
            "mode": mode,
            "client": client,
            "enclosure": self.last_enclosure,
            "main": main,
            "branches": self.branches,
            "accessories": self.accessories,
            "spd": self.var_spd.get(),
            "lines": lines,
            "status": {"order": False, "ship": False, "pay": "미입금"},
            "memo": "",
        }
        fname = f"{now_tag()}_{slug(client.get('업체명') or 'NONAME')}.json"
        fpath = os.path.join(EST_DIR, fname)
        save_json(fpath, data)
        self._repo_add_item(data, fpath)

        # [AUTOLOG] 최종 결과 로깅
        try:
            # 합계 금액을 안전하게 계산
            total_amount = 0
            for l in lines:
                try:
                    # '금액' 항목의 값을 가져와서 쉼표를 제거하고, float(소수)으로 변환 후 더함
                    amount_str = str(l.get("금액") or "0").replace(",", "")
                    total_amount += float(amount_str)
                except (ValueError, TypeError):
                    # 숫자로 바꾸지 못하면 그냥 넘어감
                    pass

            _log("assistant", {
                "action": "run_estimate_result",
                "result_path": fpath,
                "num_lines": len(lines),
                "sum": int(total_amount),  # 최종 합계는 int(정수)로 저장
                "size": _sz,
                "preview": lines[:5],
            }, msg=f"견적 결과 {len(lines)}라인 / 합계={total_amount:,.0f}")
        except Exception as _e:
            print("[AUTOLOG] result log fail:", _e)

        # --- NEW: 견적 생성 직후 RAG 자동 재색인(quiet) ---
        try:
            if hasattr(self, "_rag_reindex_quiet"):
                self._rag_reindex_quiet()
            else:
                # 백엔드 직접 호출하되, 팝업 없이 콘솔만
                try:
                    from assistant_agent import rag_reindex as _rag_reindex_backend
                    res = _rag_reindex_backend()
                    # 표준화(quiet 로깅)
                    if isinstance(res, dict):
                        ok  = bool(res.get("ok"))
                        docs = int(res.get("docs") or 0)
                    elif isinstance(res, int):
                        ok  = (res >= 0)
                        docs = max(0, res)
                    else:
                        ok, docs = False, 0
                    if ok:
                        print(f"[RAG] quiet reindex ok: docs={docs}")
                    else:
                        print("[RAG] quiet reindex skipped: no module/data or backend returned <0")
                except Exception as _e:
                    print("[RAG] quiet reindex failed:", _e)
        except Exception as e:
            print("[RAG] auto reindex failed after estimate:", e)



    def _make_system_estimate(self): self._build_and_show_estimate("system")
    def _make_ai_estimate(self): self._build_and_show_estimate("ai")

    def _save_to_company_excel(self):
        try:
            from openpyxl import load_workbook
        except Exception:
            messagebox.showerror("엑셀", "openpyxl이 필요합니다: pip install openpyxl")
            return

        tpl_path = os.path.join(DATA_DIR, "realsample.xlsx")
        if not os.path.exists(tpl_path):
            messagebox.showwarning("양식 없음", f"회사 양식이 없습니다.\n{tpl_path} 위치에 realsample.xlsx를 두세요.")
            return

        if not self.lines:
            messagebox.showwarning("저장 불가", "표에 표시된 견적 라인이 없습니다.")
            return

        # 템플릿 로드
        wb = load_workbook(tpl_path)

        # [시트1] 고객 정보 채우기 (가능하면)
        try:
            ws1 = wb.worksheets[0]
            ws1["A10"] = (self.last_client or {}).get("업체명", "")
            ws1["A11"] = (self.last_client or {}).get("연락처", "")
            ws1["A12"] = (self.last_client or {}).get("이메일", "")
        except Exception:
            pass

        # [시트2] 라인아이템 채우기
        try:
            ws2 = wb.worksheets[1]
        except Exception:
            ws2 = wb.active  # 안전 폴백

        row = 2
        for L in self.lines:
            ws2.cell(row=row, column=1, value=L.get("no"))
            ws2.cell(row=row, column=2, value=L.get("품명"))
            ws2.cell(row=row, column=3, value=L.get("규격"))
            ws2.cell(row=row, column=4, value=L.get("수량"))
            ws2.cell(row=row, column=5, value=L.get("단가"))
            ws2.cell(row=row, column=6, value=L.get("금액"))
            ws2.cell(row=row, column=7, value=L.get("비고", ""))
            row += 1

        # 저장 경로 다이얼로그
        client_name = (self.last_client or {}).get("업체명") or ""
        initial_name = make_initial_filename("견적", client_name, ".xlsx")  # 파일 초반 헬퍼 함수 사용 :contentReference[oaicite:1]{index=1}

        out_path = filedialog.asksaveasfilename(
            title="엑셀로 저장",
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", ".xlsx")],
            initialfile=initial_name,
        )
        if not out_path:
            return  # 사용자가 취소했을 때

        # 저장
        try:
            wb.save(out_path)
            messagebox.showinfo("저장 완료", out_path)
        except Exception as e:
            messagebox.showerror("저장 실패", str(e))


    def _save_current_table_csv(self):
        from tkinter import messagebox, filedialog
        if not getattr(self, "lines", None):
            messagebox.showwarning("저장 불가", "표에 표시된 견적 라인이 없습니다.")
            return
        out = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV", ".csv")],
            initialfile=f"견적_{slug((self.last_client or {}).get('업체명') or '')}_{now_tag()}.csv"
        )
        if not out:
            return
        import csv
        try:
            with open(out, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f)
                w.writerow(["No","품명","규격","단위","수량","단가","금액","비고"])
                for L in self.lines:
                    w.writerow([
                        L.get("no",""), L.get("품명",""), L.get("규격",""),
                        L.get("단위",""), L.get("수량",""), L.get("단가",""),
                        L.get("금액",""), L.get("비고","")
                    ])
            messagebox.showinfo("저장 완료", out)
        except Exception as e:
            messagebox.showerror("저장 실패", str(e))

    # ========= [탭2] AI도우미 =========
    def _build_tab_ai(self):
        tab = ttk.Frame(self.nb, padding=8)
        self.nb.add(tab, text="AI도우미")

        # --- RAG 툴바 (재색인 + 검색창) ---
        bar = ttk.Frame(tab)
        bar.pack(fill="x")
        ttk.Button(bar, text="RAG 재색인", command=_do_rag_reindex).pack(side="left", pady=6)
        ttk.Label(bar, text="  검색어:").pack(side="left", padx=(8,4))
        self.ent_rag_query_ai = ttk.Entry(bar, width=42)   # ← 이름 분리
        self.ent_rag_query_ai.pack(side="left")
        # (전역 Ctrl+K는 이미 상단 입력창으로 포커스. 여기서의 bind_all은 제거해도 무방)
        ttk.Button(bar, text="검색", command=self._rag_query).pack(side="left", padx=6)

        # 검색 결과 영역
        self.txt_rag_result = scrolledtext.ScrolledText(tab, wrap="word", height=12)
        self.txt_rag_result.pack(fill="both", expand=True, pady=(6,8))
        self.txt_rag_result.config(state="disabled")

        # 상단: 모델 선택 + 파일 첨부
        top = ttk.LabelFrame(tab, text="모델/첨부", padding=8)
        top.pack(fill="x")

        self.var_ai_model = tk.StringVar(value="chatgpt")
        ttk.Radiobutton(top, text="ChatGPT", variable=self.var_ai_model, value="chatgpt").grid(row=0, column=0, sticky="w")
        ttk.Radiobutton(top, text="Claude",  variable=self.var_ai_model, value="claude").grid(row=0, column=1, sticky="w")
        ttk.Button(top, text="파일 첨부", command=self._ai_attach_files).grid(row=0, column=2, padx=8)

        self.lst_ai_files = tk.Listbox(top, height=5, width=90)
        self.lst_ai_files.grid(row=1, column=0, columnspan=3, sticky="we", pady=(6, 0))
        if _DND_OK:
            self.lst_ai_files.drop_target_register(DND_FILES)
            self.lst_ai_files.dnd_bind('<<Drop>>', self._on_drop_ai)

        # 중간: 분석 버튼 + 메모
        mid = ttk.LabelFrame(tab, text="분석·반영", padding=8)
        mid.pack(fill="x", pady=(8, 0))
        ttk.Button(mid, text="첨부 분석 → 견적탭 자동 채움", command=self._ai_extract_to_estimate_tab).grid(row=0, column=0, padx=4)
        ttk.Label(mid, text="메모(선택)").grid(row=1, column=0, sticky="w")
        self.ent_ai_note = ttk.Entry(mid, width=90)
        self.ent_ai_note.grid(row=1, column=1, sticky="w", padx=6, pady=(4, 2))

        # 하단: 채팅창 + 입력창 + 전송
        bot = ttk.LabelFrame(tab, text="AI 채팅(업무 컨트롤)", padding=8)
        bot.pack(fill="both", expand=True, pady=(8, 0))

        self.txt_ai_chat = scrolledtext.ScrolledText(bot, wrap="word", height=18)
        self.txt_ai_chat.pack(fill="both", expand=True)

        self.txt_ai_chat.tag_configure('ai', background='#F1F5FB', foreground='#0D3B66',
                                       lmargin1=12, lmargin2=12, rmargin=80, spacing1=4, spacing3=8, justify='left')
        self.txt_ai_chat.tag_configure('user', background='#FFF5D6', foreground='#5A3E00',
                                       lmargin1=80, lmargin2=80, rmargin=12, spacing1=4, spacing3=8, justify='right')
        self.txt_ai_chat.tag_configure('sys', foreground='#666666',
                                       lmargin1=8, lmargin2=8, rmargin=8, spacing1=2, spacing3=6)
        self.txt_ai_chat.tag_config("bold", font=("맑은 고딕", 10, "bold"))

        # 읽기전용(입력 차단) + 우클릭 복사 메뉴
        def _block_keys(e): return "break"
        self.txt_ai_chat.bind("<Key>", _block_keys)
        menu = tk.Menu(self.txt_ai_chat, tearoff=0)
        menu.add_command(label="복사", command=lambda: self.txt_ai_chat.event_generate("<<Copy>>"))
        def _on_context(e):
            try: menu.tk_popup(e.x_root, e.y_root)
            finally: menu.grab_release()
        self.txt_ai_chat.bind("<Button-3>", _on_context)

        frm = ttk.Frame(bot); frm.pack(fill="x", pady=(6, 0))
        frm.grid_columnconfigure(0, weight=4)
        frm.grid_columnconfigure(1, weight=1)

        self.txt_ai_input = tk.Text(frm, height=4)
        self.txt_ai_input.grid(row=0, column=0, sticky="we")
        self.txt_ai_input.bind("<Return>", self._ai_on_enter)
        self.txt_ai_input.bind("<Shift-Return>", self._ai_on_shift_enter)

        self.btn_ai_send = ttk.Button(frm, text="보내기", command=self._ai_chat_send)
        self.btn_ai_send.grid(row=0, column=1, padx=6, sticky="e")

        # 최초 가이드 메시지
        self._ai_chat_sys("예) 'ㅇㅇ업체 견적서' → 견적서관리에서 해당 업체 검색/열람 | 'oo업체 출고건 입금했어?' → RAG에서 상태 답변")

    def _build_context_from_hits(hits_list):
        """
        RAG 검색 결과(hits)를 파일명/요약/규격 중심으로 보기 좋게 문자열로 변환.
        - 각 hit: dict로 {text, score, meta:{path,line}} 형태를 가정(유연 처리)
        - 최대 8개 항목 요약, 없으면 '(관련 근거 없음)' 반환
        """
        import os, json

        def _summarize_doc(doc: dict) -> str:
            client = (doc.get("client") or {}).get("업체명", "")
            proj   = (doc.get("project") or {}).get("건명", "")
            main   = doc.get("main_breaker") or {}
            mb = ""
            if isinstance(main, dict) and main:
                mb = f"메인 {main.get('종류','')} {main.get('극수','')} {main.get('용량','')} x{main.get('수량','1')}"
            bs = []
            for b in (doc.get("branches") or []):
                if not isinstance(b, dict): 
                    continue
                bs.append(f"{b.get('종류','')} {b.get('극수','')} {b.get('용량','')}x{b.get('수량','1')}")
            branches = ("분기: " + ", ".join(bs)) if bs else ""
            parts = [s for s in [client, proj, mb, branches] if s]
            return " | ".join(parts) if parts else "(요약 없음)"

        rows = []
        for h in (hits_list or [])[:8]:
            if isinstance(h, dict):
                txt   = h.get("text", "")
                try:
                    score = float(h.get("score", 0.0))
                except Exception:
                    score = 0.0
                meta  = h.get("meta") or {}
            else:
                txt, score, meta = str(h), 0.0, {}

            path = meta.get("path")
            line = meta.get("line")

            if path and os.path.exists(path):
                try:
                    with open(path, "r", encoding="utf-8") as f:
                        doc = json.load(f)
                except Exception:
                    doc = {}
                summary = _summarize_doc(doc)
                rows.append(f"• {round(score,3)}  파일: {os.path.basename(path)}\n  요약: {summary}")
            elif isinstance(line, dict):
                t = f"{line.get('품명','')} {line.get('규격','')} 수량 {line.get('수량','')} 단가 {line.get('단가','')}"
                rows.append(f"• {round(score,3)}  과거 라인: {t}")
            else:
                rows.append(f"• {round(score,3)}  {txt}")

        return "\n".join(rows) if rows else "(관련 근거 없음)"

    def _rag_search_from_entry(self):
        """
        최상단 입력창 값을 사용해 RAG 검색 실행.
        - query_api.rag_search(q, topk) 표준 인터페이스 사용(가능하면)
        - 실패 시 assistant_agent.rag_search로 폴백
        - 결과는 파일명/요약/규격 중심으로 정리해 표시(외함 W×H×D, 메인/분기 요약)
        - 실패/비활성은 정보성 팝업 + 조용한 로그
        """
        # 0) 질의어 확보
        q = ""
        try:
            q = (self.var_rag_query_top.get() or "").strip()
        except Exception:
            pass
        if not q:
            print("[RAG] 검색 취소 또는 빈 질의")
            try:
                from tkinter import messagebox as _mb
                _mb.showinfo("RAG 검색", "검색어를 입력하세요.")
                self._focus_rag_entry()
            except Exception:
                pass
            return

        # 1) 통일 인터페이스 로드 (query_api → assistant_agent 순차 시도)
        try:
            import importlib, sys, pathlib, os, json
            base_dir = pathlib.Path(__file__).resolve().parent
            if str(base_dir) not in sys.path:
                sys.path.append(str(base_dir))

            _rag_search = None
            try:
                _Q = importlib.import_module("query_api")  # type: ignore[reportMissingImports]
                _rag_search = getattr(_Q, "rag_search", None)
            except Exception:
                _rag_search = None

            if _rag_search is None:
                try:
                    from assistant_agent import rag_search as _rag_search  # type: ignore
                except Exception:
                    _rag_search = None

            if _rag_search is None:
                raise RuntimeError("검색 API를 찾을 수 없습니다(query_api/assistant_agent).")

            raw = _rag_search(q, topk=8) or []
        except Exception as ex:
            print("[RAG] 검색 오류:", ex)
            try:
                from tkinter import messagebox as _mb
                _mb.showinfo("RAG 검색", f"검색 실패: {ex}")
            except Exception:
                pass
            return

        # 2) hits 정규화(list 또는 dict.items)
        if isinstance(raw, dict) and "items" in raw:
            hits = list(raw.get("items") or [])
        elif isinstance(raw, (list, tuple)):
            hits = list(raw)
        else:
            hits = []

        # 3) 요약 빌더
        def _summarize_doc(doc: dict) -> str:
            client = (doc.get("client") or {}).get("업체명", "")
            proj   = (doc.get("project") or {}).get("건명", "")
            enc    = (doc.get("enclosure") or {})
            size_s = ""
            try:
                W, H, D = enc.get("W"), enc.get("H"), enc.get("D")
                if all(isinstance(x, (int, float)) for x in (W, H, D)):
                    size_s = f"외함 {int(W)}×{int(H)}×{int(D)}"
            except Exception:
                pass
            main   = doc.get("main_breaker") or {}
            mb = f"메인 {main.get('종류','')} {main.get('극수','')} {main.get('용량','')} x{main.get('수량','1')}" if main else ""
            bs = []
            for b in doc.get("branches", []) or []:
                bs.append(f"{b.get('종류','')} {b.get('극수','')} {b.get('용량','')}x{b.get('수량','1')}")
            branches = ("분기: " + ", ".join(bs)) if bs else ""
            parts = [s for s in [client, proj, size_s, mb, branches] if s]
            return " | ".join(parts) if parts else "(요약 없음)"

        def _format_hits(hits_list):
            rows = []
            for h in hits_list[:8]:
                # hit 표준 필드 방어적 접근
                if isinstance(h, dict):
                    txt   = h.get("text", "")
                    score = h.get("score", 0.0)
                    meta  = h.get("meta") or {}
                else:
                    txt, score, meta = str(h), 0.0, {}

                path  = meta.get("path")
                line  = meta.get("line")  # 과거 라인(품명/규격/수량/단가) dict일 수 있음

                if path and os.path.exists(path):
                    try:
                        with open(path, "r", encoding="utf-8") as f:
                            doc = json.load(f)
                    except Exception:
                        doc = {}
                    summary = _summarize_doc(doc)
                    rows.append(f"• {round(float(score),3)}  파일: {os.path.basename(path)}\n  요약: {summary}")
                elif isinstance(line, dict):
                    t = f"{line.get('품명','')} {line.get('규격','')} 수량 {line.get('수량','')} 단가 {line.get('단가','')}"
                    rows.append(f"• {round(float(score),3)}  과거 라인: {t}")
                else:
                    # 최소 보장
                    base = txt if txt else "(텍스트 없음)"
                    rows.append(f"• {round(float(score),3)}  {base}")
            return "\n".join(rows) if rows else "(관련 근거 없음)"

        # 4) 출력 구성(요약 컨텍스트 + 상위 매치)
        top_bullets = _format_hits(hits)
        msg = f"[RAG 요약]\n{top_bullets}"

        # 5) UI 반영 + 안내
        try:
            self.txt_rag_result.config(state="normal")
            self.txt_rag_result.delete("1.0", "end")
            self.txt_rag_result.insert("end", msg)
            self.txt_rag_result.config(state="disabled")
        except Exception:
            pass

        try:
            from tkinter import messagebox as _mb
            _mb.showinfo("RAG 검색", "검색 완료. 결과 영역을 확인하세요.")
        except Exception:
            pass

        print("[RAG] search done:", len(hits), "hits")
        return


    def _focus_rag_entry(self):
        """Ctrl+K 단축키: 최상단 RAG 검색 입력창에 포커스"""
        try:
            self.ent_rag_query_top.focus_set()
            try:
                self.ent_rag_query_top.focus_force()
            except Exception:
                pass
            import tkinter as tk
            self.ent_rag_query_top.icursor(tk.END)
        except Exception as e:
            print("[RAG] focus shortcut failed:", e)

    def _build_context_from_hits(self, hits_list):
        """
        RAG hits를 화면 요약 문자열로 구성한다.
        - 파일형 문서: 파일명 + 추출 요약(메인/분기)
        - 라인형 문서: 품명/규격/수량/단가
        - 그 외: 텍스트 + 스코어
        """
        import os, json
        def _summarize_doc(doc: dict) -> str:
            client = (doc.get("client") or {}).get("업체명", "")
            proj   = (doc.get("project") or {}).get("건명", "")
            main   = doc.get("main_breaker") or {}
            mb = f"메인 {main.get('종류','')} {main.get('극수','')} {main.get('용량','')} x{main.get('수량','1')}" if main else ""
            bs = []
            for b in doc.get("branches", []):
                bs.append(f"{b.get('종류','')} {b.get('극수','')} {b.get('용량','')}x{b.get('수량','1')}")
            branches = ("분기: " + ", ".join(bs)) if bs else ""
            parts = [s for s in [client, proj, mb, branches] if s]
            return " | ".join(parts) if parts else "(요약 없음)"

        rows = []
        for h in (hits_list or [])[:8]:
            try:
                txt   = h.get("text", "")
                score = h.get("score", 0.0)
                meta  = h.get("meta") or {}
                path  = meta.get("path")
                line  = meta.get("line")  # 과거 라인 dict 가능

                if path and os.path.exists(path):
                    try:
                        with open(path, "r", encoding="utf-8") as f:
                            doc = json.load(f)
                    except Exception:
                        doc = {}
                    summary = _summarize_doc(doc)
                    rows.append(f"• {round(float(score),3)}  파일: {os.path.basename(path)}\n  요약: {summary}")
                elif isinstance(line, dict):
                    t = f"{line.get('품명','')} {line.get('규격','')} 수량 {line.get('수량','')} 단가 {line.get('단가','')}"
                    rows.append(f"• {round(float(score),3)}  과거 라인: {t}")
                else:
                    rows.append(f"• {round(float(score),3)}  {txt}")
            except Exception:
                # 방어적으로 한 줄이라도 남기기
                rows.append(f"• -  {str(h)[:120]}")
        return "\n".join(rows) if rows else "(관련 근거 없음)"

    def _rag_query(self):
        """
        [AI 탭] 입력창 값으로 RAG 검색 실행.
        - query_api.rag_search(q, topk) 우선
        - 실패 시 assistant_agent.rag_search 폴백
        - 결과는 파일명/요약/규격 중심으로 정리해 표시
        - 정적 임포트 제거(동적 importlib)로 Pylance 경고 제거
        """
        import importlib, sys, pathlib, os, json
        from tkinter import messagebox as _mb

        # 0) 질의어 확보 (AI 탭 전용 입력창)
        q = ""
        try:
            q = (self.ent_rag_query_ai.get() or "").strip()
        except Exception:
            pass
        if not q:
            _mb.showinfo("RAG 검색", "검색어를 입력하세요.")
            try:
                self.ent_rag_query_ai.focus_set()
            except Exception:
                pass
            return

        # 1) 통일 인터페이스 로드 (query_api → assistant_agent 순차 시도, 정적 import 금지)
        hits = []
        try:
            base_dir = pathlib.Path(__file__).resolve().parent
            if str(base_dir) not in sys.path:
                sys.path.append(str(base_dir))

            _rag_search = None
            try:
                _Q = importlib.import_module("query_api")  # type: ignore[reportMissingImports]
                _rag_search = getattr(_Q, "rag_search", None)
            except Exception:
                _rag_search = None

            if _rag_search is None:
                try:
                    _A = importlib.import_module("assistant_agent")  # type: ignore[reportMissingImports]
                    _rag_search = getattr(_A, "rag_search", None)
                except Exception:
                    _rag_search = None

            if _rag_search is None or not callable(_rag_search):
                raise RuntimeError("검색 API를 찾을 수 없습니다(query_api/assistant_agent).")

            hits = _rag_search(q, topk=8) or []

        except Exception as ex:
            print("[RAG] 검색 오류:", ex)
            _mb.showinfo("RAG 검색", f"검색 실패: {ex}")
            return

        # 2) 결과 정리 도우미
        def _summarize_doc(doc: dict) -> str:
            client = (doc.get("client") or {}).get("업체명", "")
            proj   = (doc.get("project") or {}).get("건명", "")
            main   = doc.get("main_breaker") or {}
            mb = f"메인 {main.get('종류','')} {main.get('극수','')} {main.get('용량','')} x{main.get('수량','1')}" if main else ""
            bs = []
            for b in doc.get("branches", []):
                bs.append(f"{b.get('종류','')} {b.get('극수','')} {b.get('용량','')}x{b.get('수량','1')}")
            branches = ("분기: " + ", ".join(bs)) if bs else ""
            parts = [s for s in [client, proj, mb, branches] if s]
            return " | ".join(parts) if parts else "(요약 없음)"

        def _format_hits(hits_list):
            rows = []
            for h in hits_list[:8]:
                txt   = h.get("text", "")
                score = h.get("score", 0.0)
                meta  = h.get("meta") or {}
                path  = meta.get("path")
                line  = meta.get("line")  # 과거 라인(품명/규격/수량/단가) dict일 수 있음

                if path and os.path.exists(path):
                    try:
                        with open(path, "r", encoding="utf-8") as f:
                            doc = json.load(f)
                    except Exception:
                        doc = {}
                    summary = _summarize_doc(doc)
                    rows.append(f"• {round(float(score),3)}  파일: {os.path.basename(path)}\n  요약: {summary}")
                elif isinstance(line, dict):
                    t = f"{line.get('품명','')} {line.get('규격','')} 수량 {line.get('수량','')} 단가 {line.get('단가','')}"
                    rows.append(f"• {round(float(score),3)}  과거 라인: {t}")
                else:
                    rows.append(f"• {round(float(score),3)}  {txt}")
            return "\n".join(rows) if rows else "(관련 근거 없음)"

        # 3) 출력 구성
        top_bullets = _format_hits(hits)
        msg = f"[RAG 요약]\n{top_bullets}"

        # 4) UI 반영
        try:
            self.txt_rag_result.config(state="normal")
            self.txt_rag_result.delete("1.0", "end")
            self.txt_rag_result.insert("end", msg)
            self.txt_rag_result.config(state="disabled")
        except Exception:
            pass

        try:
            _mb.showinfo("RAG 검색", "검색 완료. 결과 영역을 확인하세요.")
        except Exception:
            pass

        print("[RAG] search done:", len(hits), "hits")

    def _rag_search_core(self, query: str, topk: int = 5):
        # 1) agent 우선
        try:
            if hasattr(self, "agent") and self.agent is not None:
                fn = getattr(self.agent, "rag_search", None) or getattr(self.agent, "search", None)
                if callable(fn):
                    return fn(query, topk=topk)
        except Exception as e:
            print("[RAG] agent search fail:", e)
        # 2) assistant_agent
        try:
            from assistant_agent import rag_search as as_rag_search
            return as_rag_search(query, topk=topk)
        except Exception as e:
            print("[RAG] assistant_agent.rag_search not available:", e)
        # 3) utils.rag
        try:
            from utils import rag as R
            if hasattr(R, "query"):
                return R.query(query, topk=topk)
        except Exception as e:
            print("[RAG] utils.rag.query not available:", e)
        return []


    # === 채팅 말풍선/전송 핸들러
    def _ai_sys_clear(self):
        try:
            self.txt_ai_chat.config(state="normal")
            start = "1.0"
            while True:
                idx = self.txt_ai_chat.search("[시스템]", start, tk.END)
                if not idx:
                    break
                line_end = f"{idx} lineend"
                self.txt_ai_chat.delete(idx, line_end+"\n")
                start = idx
            self.txt_ai_chat.config(state="disabled")
        except:
            pass

    def _ai_chat_user(self, t: str):
        try:
            if hasattr(self, "ai_chat_list"):
                try:
                    self.ai_chat_list.insert("end", f"사용자: {t}\n")
                    self.ai_chat_list.see("end")
                except Exception:
                    pass
        finally:
            try:
                ui_smoke_tab_open()  # E1: 탭 열림/입력 포커스 이벤트 대체 기록
            except Exception:
                pass


    def _ai_chat_bot(self, t: str):
        try:
            rows = None
            if hasattr(self, "ai_chat_list"):
                try:
                    self.ai_chat_list.insert("end", f"AI: {t}\n")
                    self.ai_chat_list.see("end")
                    # 결과 테이블/리스트 크기 추정
                    try:
                        rows = self.ai_chat_list.size()
                    except Exception:
                        rows = None
                except Exception:
                    pass
        finally:
            try:
                ui_smoke_table_ok({"rows": rows})
            except Exception:
                pass

    # === KISAN-AI PATCH: Explain Mini Report UI Hook ===
    def _show_explain_card(self, explain: dict | None):
        try:
            if not explain:
                return
            try:
                from explain_report import render_explain_card
            except Exception:
                return
            card = render_explain_card(explain)
            # 텍스트 위젯에 표시(있을 때만)
            if hasattr(self, "ai_chat_list") and self.ai_chat_list:
                try:
                    self.ai_chat_list.insert("end", "\n[외함 선정 리포트]\n")
                    self.ai_chat_list.insert("end", card + "\n")
                    self.ai_chat_list.see("end")
                except Exception:
                    pass
            # 파일로도 저장(백업)
            try:
                from pathlib import Path
                out = Path(__file__).resolve().parent / "out" / "explain_report.txt"
                out.parent.mkdir(parents=True, exist_ok=True)
                out.write_text(card, encoding="utf-8")
            except Exception:
                pass
        except Exception:
            pass
    # === /KISAN-AI PATCH ===


    def _ai_chat_sys(self, t: str):
        self.txt_ai_chat.config(state="normal")
        self.txt_ai_chat.insert(tk.END, f"[시스템] {t}\n\n", ("sys",))
        self.txt_ai_chat.see(tk.END)
        self.txt_ai_chat.config(state="disabled")

    # === KISAN-AI PATCH: _ai_on_enter (간단 로깅) ===
    def _ai_on_enter(self, event):
        try:
            import pathlib, json, time
            out = pathlib.Path(__file__).resolve().parent / "out"
            out.mkdir(parents=True, exist_ok=True)
            (out / "ui_keys_log.jsonl").open("a", encoding="utf-8").write(
                json.dumps({"ts": time.time(), "event": "enter"}, ensure_ascii=False) + "\n"
            )
        except Exception:
            pass
        self._ai_chat_send()
        return "break"
    # === /KISAN-AI PATCH ===



    def _ai_on_shift_enter(self, event):
        try:
            self._ai_chat_send(multiline=True)
        finally:
            try:
                ui_smoke_btn2()  # E3: 버튼2(Shift+Enter) 이벤트 기록
            except Exception:
                pass
        return "break"


    # === KISAN-AI PATCH: _ai_chat_send (입력 로깅 추가) ===
    def _ai_chat_send(self):
        import re, pathlib, json, time, tkinter as tk
        msg = self.txt_ai_input.get("1.0", "end-1c").strip()
        if not msg:
            return
        # 입력 로깅
        try:
            out = pathlib.Path(__file__).resolve().parent / "out"
            out.mkdir(parents=True, exist_ok=True)
            (out / "ui_input_log.jsonl").open("a", encoding="utf-8").write(
                json.dumps({"ts": time.time(), "input": msg}, ensure_ascii=False) + "\n"
            )
        except Exception:
            pass

        self._ai_chat_user(msg)
        self.txt_ai_input.delete("1.0", tk.END)

        if "견적서" in msg:
            name = re.sub(r"(견적서|보여줘|열어줘)", "", msg).strip()
            self.nb.select(3)
            self._repo_search_and_focus(name)
            self._ai_chat_bot(f"'{name}' 관련 견적을 견적서관리에서 표시했어요.")
            return
        if "출고" in msg or "입금" in msg or "발주" in msg:
            ans = "해당 기능은 현재 개발 중입니다."
            self._ai_chat_bot(ans)
            return
        if "견적탭" in msg or "시스템 견적" in msg:
            self.nb.select(0)
            self._ai_chat_bot("견적 탭으로 이동했습니다.")
            return

        # ★ RAG+LLM 백그라운드 호출
        self._set_busy(True)
        self._ai_chat_sys("답변 생성 중…")
        import threading
        threading.Thread(target=self._agent_worker, args=(msg,), daemon=True).start()
    # === /KISAN-AI PATCH ===


    def _set_busy(self, busy: bool):
        def _apply():
            try:
                self.btn_ai_send.configure(state=("disabled" if busy else "normal"))
            except:
                pass
        try:
            self.after(0, _apply)
        except Exception:
            _apply()

    # === KISAN-AI PATCH: _agent_worker (로깅+리포트 훅) ===
    def _agent_worker(self, msg: str):
        def _llm_wrapper(prompt: str) -> str:
            try:
                model = self.var_ai_model.get()
            except Exception:
                model = "chatgpt"
            return self._call_llm(model, prompt) or ""

        try:
            use_llm = bool(self.cfg.get("chatgpt", {}).get("ok") or self.cfg.get("claude", {}).get("ok"))
        except Exception:
            use_llm = False

        try:
            if hasattr(self, "agent") and self.agent is not None:
                resp = self.agent.reply(msg, llm_fn=_llm_wrapper if use_llm else None)
            else:
                resp = _llm_wrapper(msg) if use_llm else "RAG 에이전트가 초기화되지 않았습니다."
        except Exception as e:
            resp = f"처리 중 오류: {e}"

        # 결과 로깅(JSONL) + explain 카드 훅 (dict일 때)
        try:
            import json, pathlib, time
            out = pathlib.Path(__file__).resolve().parent / "out"
            out.mkdir(parents=True, exist_ok=True)
            rec = {
                "ts": time.time(),
                "input": msg,
                "output_type": type(resp).__name__,
                "use_llm": use_llm,
            }
            # explain 카드 표시 시도
            if isinstance(resp, dict):
                rec["has_explain"] = bool(resp.get("explain"))
                try:
                    self._show_explain_card(resp.get("explain"))
                except Exception:
                    pass
            (out / "ui_handler_log.jsonl").open("a", encoding="utf-8").write(json.dumps(rec, ensure_ascii=False) + "\n")
        except Exception:
            pass

        def _deliver():
            self._ai_sys_clear()
            try:
                self._ai_chat_bot(resp if not isinstance(resp, dict) else (resp.get("text") or "[dict 응답]"))
            except Exception:
                self._ai_chat_bot(str(resp) if resp else "규칙/자료 기반 응답이 없습니다.")
            self._set_busy(False)

        try:
            self.after(0, _deliver)
        except Exception:
            try:
                self.txt_ai_chat.after(0, _deliver)
            except Exception:
                _deliver()
    # === /KISAN-AI PATCH ===


    # === [AI 탭] 파일 첨부/드롭/파싱/자동 채움 ===
    def _ai_attach_files(self):
        # 엑셀 / CAD까지 한 번에 보이도록 필터 구성
        paths = filedialog.askopenfilenames(
            title="첨부 파일",
            filetypes=[
                ("엑셀", "*.xlsx"),
                ("CAD 도면", "*.dwg;*.dxf"),
                ("PDF", "*.pdf"),
                ("이미지", "*.png;*.jpg;*.jpeg;*.bmp;*.tif;*.tiff"),
                ("워드", "*.docx"),
                ("텍스트", "*.txt"),
                ("모든 파일", "*.*"),
            ]
       )
        if paths:
            for p in paths:
                self.lst_ai_files.insert(tk.END, p)

    def _on_drop_ai(self, evt):
        for f in self._parse_dnd(evt.data):
            self.lst_ai_files.insert(tk.END,f)

    def _parse_dnd(self, data):
        items=[]
        for it in re.findall(r'\{[^}]+\}|[^\s]+', data):
            items.append(it.strip("{}"))
        return items

    def _ai_extract_to_estimate_tab(self):
        files=list(self.lst_ai_files.get(0,tk.END))
        note_text = self.ent_ai_note.get().strip()
        if not files and not note_text:
            messagebox.showwarning("첨부 필요","파일 첨부 또는 메모 입력"); return
        tmp=None
        if note_text:
            tmp=os.path.join(DATA_DIR,f"_tmp_note_{now_tag()}.txt")
            with open(tmp,"w",encoding="utf-8") as f: f.write(note_text)
            files=files+[tmp]
        tess=self.cfg.get("tesseract_path","")
        try:
            client, main, branches, _full = extract_info_from_files(files, tesseract_path=tess)
        except Exception as e:
            messagebox.showerror("분석 실패", f"{e}\n\n(pytesseract / pdfplumber / docx / pdf2image 중 일부가 없을 수 있습니다)")
            if tmp and os.path.exists(tmp): os.remove(tmp)
            return
        finally:
            if tmp and os.path.exists(tmp): os.remove(tmp)
        # 견적 탭 채움
        if client.get("업체명"): self.ent_client_name.delete(0,tk.END); self.ent_client_name.insert(0,client["업체명"])
        if client.get("연락처"): self.ent_client_phone.delete(0,tk.END); self.ent_client_phone.insert(0,client["연락처"])
        if client.get("이메일"): self.ent_client_email.delete(0,tk.END); self.ent_client_email.insert(0,client["이메일"])
        if main:
            self.cbo_main_kind.set(main["종류"]); self.cbo_main_poles.set(main["극수"]); self.cbo_main_amp.set(main["용량"])
            self.ent_main_qty.delete(0,tk.END); self.ent_main_qty.insert(0, main.get("수량","1"))
        self.branches.clear(); self.lst_branches.delete(0,tk.END)
        merged={}
        for b in branches:
            key=(b["종류"],b["극수"],b["용량"])
            merged[key]=merged.get(key,0)+safe_int(b.get("수량","1"),1)
        for (k,p,a),qty in merged.items():
            self.branches.append({"종류":k,"극수":p,"용량":a,"수량":str(qty)})
            self.lst_branches.insert(tk.END, f"{k} | {p} {a} | 수량 {qty}")
        self._ai_chat_sys("첨부 분석 결과를 견적 탭에 반영했습니다."); self.nb.select(0)

    # ========= [탭3] 이메일(3-Pane) =========
    def _build_tab_email(self):
        tab=ttk.Frame(self.nb, padding=8); self.nb.add(tab, text="이메일")
        # 상단 툴바
        bar=ttk.Frame(tab); bar.pack(fill="x")
        ttk.Label(bar,text="최근(일)").pack(side="left", padx=(0,4))
        self.ent_mail_days = ttk.Entry(bar, width=4); self.ent_mail_days.insert(0,"5"); self.ent_mail_days.pack(side="left")
        ttk.Button(bar,text="메일함 동기화", command=self._mail_sync).pack(side="left", padx=(4,0))
        ttk.Button(bar,text="더 불러오기(+5일)", command=self._mail_load_more).pack(side="left", padx=6)
        ttk.Button(bar,text="첨부 다운로드", command=self._mail_download).pack(side="left", padx=6)
        ttk.Button(bar,text="첨부 → AI도우미", command=self._mail_attach_to_ai).pack(side="left", padx=6)
        ttk.Button(bar,text="첨부 → AI견적", command=self._mail_attach_to_ai_est).pack(side="left", padx=6)
        ttk.Button(bar,text="메일 보내기(모의)", command=self._mail_send).pack(side="left", padx=6)

        # 3 Pane
        main=ttk.Panedwindow(tab, orient="horizontal"); main.pack(fill="both", expand=True, pady=(8,0))

        # 좌: 메일함/필터
        left=ttk.Frame(main, width=220); main.add(left, weight=1)
        ttk.Label(left,text="메일함").pack(anchor="w")
        self.lst_mailboxes=tk.Listbox(left, height=8)
        for b in ["INBOX"]:
            self.lst_mailboxes.insert(tk.END, b)
        self.lst_mailboxes.pack(fill="x")
        ttk.Label(left,text="검색").pack(anchor="w", pady=(8,2))
        self.ent_mail_search=ttk.Entry(left); self.ent_mail_search.pack(fill="x")
        ttk.Button(left,text="검색(모의)", command=self._mail_search).pack(anchor="w", pady=(6,0))

        # 중: 메일 리스트
        mid=ttk.Frame(main, width=420); main.add(mid, weight=2)
        cols=("date","from","subject","attachments")
        self.tree_mail=ttk.Treeview(mid, columns=cols, show="headings", height=20)
        for c,w in zip(cols,["날짜","보낸사람","제목","첨부"]):
            self.tree_mail.heading(c, text=w)
        self.tree_mail.column("date", width=150); self.tree_mail.column("from", width=170); self.tree_mail.column("subject", width=380)
        self.tree_mail.column("attachments", width=80, anchor="center")
        self.tree_mail.pack(fill="both", expand=True)
        self.tree_mail.bind("<<TreeviewSelect>>", self._mail_on_select)

        # 우: 본문/첨부/액션
        right=ttk.Frame(main); main.add(right, weight=3)
        self.txt_mail_body=scrolledtext.ScrolledText(right, wrap="word", height=18); self.txt_mail_body.pack(fill="both",expand=True)
        atta_bar=ttk.Frame(right); atta_bar.pack(fill="x", pady=(6,0))
        ttk.Label(atta_bar,text="첨부파일:").pack(side="left")
        self.lst_mail_attachments=tk.Listbox(atta_bar,height=4); self.lst_mail_attachments.pack(side="left", fill="x", expand=True)
        btns=ttk.Frame(right); btns.pack(fill="x", pady=(6,0))
        ttk.Button(btns,text="선택 첨부 → AI도우미", command=self._attachment_to_ai).pack(side="left")
        ttk.Button(btns,text="선택 첨부 → AI견적", command=self._mail_attach_to_ai_est).pack(side="left", padx=6)

    def _imap_since_query(self, days: int) -> str:
        import datetime
        dt = datetime.datetime.now() - datetime.timedelta(days=max(0, days))
        return dt.strftime("%d-%b-%Y")

    def _mail_load_range(self, days: int):
        try:
            imcfg   = self.cfg.get("imap", {})
            host    = imcfg.get("host", "imap.naver.com")
            port    = int(imcfg.get("port", 993))
            user    = imcfg.get("user", "").strip()
            pw      = imcfg.get("pass", "").strip()
            use_ssl = bool(imcfg.get("ssl", True))
            if not (host and user and pw):
                messagebox.showwarning("IMAP", "설정 탭에서 네이버웍스 IMAP 정보를 저장하세요.")
                return

            M = imaplib.IMAP4_SSL(host, port) if use_ssl else imaplib.IMAP4(host, port)
            M.login(user, pw)
            typ, _ = M.select("INBOX")
            if typ != "OK":
                raise RuntimeError("IMAP 메일함 선택 실패")

            since = self._imap_since_query(days)
            typ, data = M.search(None, f'(SINCE "{since}")')
            if typ != "OK":
                raise RuntimeError("IMAP 검색 실패")
            uids = list(reversed(data[0].split()))[:300]  # 최신 300개 제한

            self.tree_mail.delete(*self.tree_mail.get_children())
            self.mail_cache.clear()

            for uid in uids:
                typ, d = M.fetch(uid, "(RFC822.HEADER)")
                if typ != "OK":
                    continue
                msg = email.message_from_bytes(d[0][1])
                date_s, frm, subj, atts = _msg_summary(msg)
                self.tree_mail.insert("", "end", iid=uid.decode(),
                                      values=(date_s, frm, subj, "O" if atts else ""))
            M.logout()
        except Exception as ex:
            traceback.print_exc()
            messagebox.showerror("IMAP 오류", str(ex))

    def _mail_load_more(self):
        try:
            cur = int(self.ent_mail_days.get().strip() or "5")
        except:
            cur = 5
        cur += 5
        self.ent_mail_days.delete(0,tk.END); self.ent_mail_days.insert(0,str(cur))
        self._mail_load_range(cur)

    def _mail_sync(self):
        try:
            days = int(self.ent_mail_days.get().strip() or "5")
        except:
            days = 5
        self._mail_load_range(days)
        messagebox.showinfo("메일", f"최근 {days}일 메일함 동기화 완료")

    def _html_to_text(self, html_s: str) -> str:
        import re, html as ihtml
        s = html_s or ""
        s = re.sub(r'(?is)<(script|style).*?>.*?</\1>', ' ', s)
        s = re.sub(r'(?i)<br\s*/?>', '\n', s)
        s = re.sub(r'(?i)</p\s*>', '\n', s)
        s = re.sub(r'(?i)<(p|div|tr|h[1-6]|li|table|section|article|header|footer)[^>]*>', '\n', s)
        s = re.sub(r'(?s)<[^>]+>', ' ', s)
        s = ihtml.unescape(s).replace('\xa0', ' ')
        s = re.sub(r'[ \t]+', ' ', s)
        s = re.sub(r'\n[ \t]+', '\n', s)
        s = re.sub(r'\n{3,}', '\n\n', s)
        return s.strip()

    def _extract_best_body_text(self, msg) -> str:
        try:
            ctype = (msg.get_content_type() or "").lower()
            if not msg.is_multipart():
                raw = msg.get_payload(decode=True) or b""
                text = raw.decode(msg.get_content_charset() or "utf-8", errors="ignore")
                return text if ctype == "text/plain" else self._html_to_text(text)
            plains, htmls = [], []
            for part in msg.walk():
                cdisp = (part.get_content_disposition() or "").lower()
                if cdisp == "attachment":
                    continue
                ptype = (part.get_content_type() or "").lower()
                raw = part.get_payload(decode=True) or b""
                text = raw.decode(part.get_content_charset() or "utf-8", errors="ignore")
                if ptype == "text/plain":
                    plains.append(text)
                elif ptype == "text/html":
                    htmls.append(self._html_to_text(text))
            if plains:
                return "\n".join(plains).strip()
            if htmls:
                return "\n".join(htmls).strip()
            raw = msg.get_payload(decode=True) or b""
            return raw.decode(msg.get_content_charset() or "utf-8", errors="ignore") or "(본문이 없습니다.)"
        except Exception:
            return "(본문이 없습니다.)"

    def _mail_on_select(self, _e=None):
        sel = self.tree_mail.selection()
        if not sel:
            return
        uid = sel[0]

        imcfg   = self.cfg.get("imap", {})
        host    = imcfg.get("host", "imap.naver.com")
        port    = int(imcfg.get("port", 993))
        user    = imcfg.get("user", "").strip()
        pw      = imcfg.get("pass", "").strip()
        use_ssl = bool(imcfg.get("ssl", True))
        if not (host and user and pw):
            messagebox.showwarning("IMAP", "설정 탭에서 네이버웍스 IMAP 정보를 저장하세요.")
            return

        try:
            M = imaplib.IMAP4_SSL(host, port) if use_ssl else imaplib.IMAP4(host, port)
            M.login(user, pw)
            M.select("INBOX")

            typ, d = M.fetch(uid.encode(), "(RFC822)")
            if typ != "OK":
                raise RuntimeError("본문 로드 실패")
            msg = email.message_from_bytes(d[0][1])

            body_text = self._extract_best_body_text(msg)
            self.txt_mail_body.delete("1.0", tk.END)
            self.txt_mail_body.insert(tk.END, body_text)

            self.lst_mail_attachments.delete(0, tk.END)
            self.mail_cache[uid] = {"msg": msg}
            for part in msg.walk():
                if (part.get_content_disposition() or "").lower() == "attachment":
                    fn = _dh(part.get_filename() or "attachment.bin")
                    self.lst_mail_attachments.insert(tk.END, fn)

            M.logout()

        except Exception as ex:
            traceback.print_exc()
            messagebox.showerror("IMAP 오류", str(ex))

    def _mail_download(self):
        sel=self.tree_mail.selection()
        if not sel: return
        uid=sel[0]
        entry = self.mail_cache.get(uid)
        if not entry:
            messagebox.showwarning("첨부","먼저 메일을 선택해 본문을 로드해주세요.")
            return
        msg = entry.get("msg")
        outdir = os.path.join(DATA_DIR, "mail_attachments", uid)
        os.makedirs(outdir, exist_ok=True)
        saved=[]
        for part in msg.walk():
            if (part.get_content_disposition() or "").lower()=="attachment":
                fn=_dh(part.get_filename() or f"att_{len(saved)+1}.bin")
                path=os.path.join(outdir, fn)
                with open(path,"wb") as f:
                    f.write(part.get_payload(decode=True))
                saved.append(path)
        messagebox.showinfo("다운로드", f"{len(saved)}개 저장됨\n{outdir}")

    def _mail_send(self): messagebox.showinfo("보내기","메일 발송(모의)")
    def _mail_search(self): messagebox.showinfo("검색","검색 결과 반영(모의)")

    def _mail_attach_to_ai(self):
        sel=self.tree_mail.selection()
        if not sel: return
        uid=sel[0]
        entry = self.mail_cache.get(uid)
        if not entry:
            messagebox.showwarning("첨부","먼저 본문을 로드해주세요."); return
        msg = entry.get("msg")
        names=[]
        outdir = os.path.join(DATA_DIR, "mail_attachments", uid)
        os.makedirs(outdir, exist_ok=True)
        i=0
        for part in msg.walk():
            if (part.get_content_disposition() or "").lower()=="attachment":
                i+=1
                fn=_dh(part.get_filename() or f"att_{i}.bin")
                path=os.path.join(outdir, fn)
                with open(path,"wb") as f: f.write(part.get_payload(decode=True))
                names.append(path)
        self.nb.select(1)
        for p in names: self.lst_ai_files.insert(tk.END, p)
        messagebox.showinfo("전송", f"AI도우미로 {len(names)}개 보냈습니다.")

    def _mail_attach_to_ai_est(self):
        # 선택 첨부 저장 + AI도우미 첨부 리스트에 추가
        self._attachment_to_ai()
        # → 첨부 분석을 통해 견적 탭 자동 채움 시도
        try:
            self.nb.select(1)  # AI도우미 탭
            self._ai_extract_to_estimate_tab()   # 첨부 분석 → 견적탭 채움
            self.nb.select(0)  # 견적 탭
            messagebox.showinfo("완료","첨부를 분석하여 견적 탭에 반영했습니다.")
        except Exception as e:
            messagebox.showerror("실패", f"AI 견적 자동 반영 실패: {e}")

    def _attachment_to_ai(self):
        sels=self.lst_mail_attachments.curselection()
        if not sels: return
        uid = self.tree_mail.selection()[0]
        entry = self.mail_cache.get(uid)
        if not entry:
            messagebox.showwarning("첨부","먼저 본문을 로드해주세요."); return
        msg = entry.get("msg")
        outdir = os.path.join(DATA_DIR, "mail_attachments", uid)
        os.makedirs(outdir, exist_ok=True)
        names=[]
        i=0
        for part in msg.walk():
            if (part.get_content_disposition() or "").lower()=="attachment":
                i+=1
                fn=_dh(part.get_filename() or f"att_{i}.bin")
                if (i-1) in sels:
                    path=os.path.join(outdir, fn)
                    with open(path,"wb") as f: f.write(part.get_payload(decode=True))
                    names.append(path)
        self.nb.select(1)
        for p in names: self.lst_ai_files.insert(tk.END, p)
        messagebox.showinfo("전송","AI도우미로 보냈습니다.")

    # ========= [탭4] 견적서 관리 =========
    def _build_tab_repo(self):
        tab=ttk.Frame(self.nb, padding=8); self.nb.add(tab, text="견적서관리")
        top=ttk.Frame(tab); top.pack(fill="x")
        ttk.Button(top,text="새로고침", command=self._repo_reload).pack(side="left")
        ttk.Button(top,text="Excel 저장(.xlsx)", command=self._repo_save_excel).pack(side="left",padx=6)
        ttk.Button(top,text="CSV 저장(.csv)", command=self._repo_save_csv).pack(side="left")
        ttk.Label(top,text="상태:").pack(side="left", padx=(18,4))
        self.var_repo_order = tk.BooleanVar(value=False)
        self.var_repo_ship  = tk.BooleanVar(value=False)
        self.var_repo_pay   = tk.StringVar(value="미입금")
        ttk.Checkbutton(top,text="발주", variable=self.var_repo_order, command=self._repo_update_status).pack(side="left")
        ttk.Checkbutton(top,text="출고", variable=self.var_repo_ship,  command=self._repo_update_status).pack(side="left",padx=(6,0))
        ttk.Combobox(top, values=["미입금","계약금","중도금","잔금","완납"], textvariable=self.var_repo_pay, width=8, state="readonly").pack(side="left", padx=(6,0))
        ttk.Label(top,text="검색 업체").pack(side="left", padx=(18,4))
        self.ent_repo_search=ttk.Entry(top, width=20); self.ent_repo_search.pack(side="left")
        ttk.Button(top,text="찾기", command=lambda:self._repo_search_and_focus(self.ent_repo_search.get().strip())).pack(side="left", padx=4)

        cols=("time","client","order","ship","pay","memo","path")
        self.tree_repo=ttk.Treeview(tab, columns=cols, show="headings", height=22)
        headers=["생성시각","업체명","발주","출고","입금","비고","(경로숨김)"]
        widths =[140,220,60,60,100,260,10]
        for c,h,w in zip(cols,headers,widths):
            self.tree_repo.heading(c, text=h)
            self.tree_repo.column(c, width=w)
        self.tree_repo.column("path", width=1, stretch=False)
        self.tree_repo.pack(fill="both", expand=True, pady=(8,0))
        self.tree_repo.bind("<<TreeviewSelect>>", self._repo_on_select)
        self._repo_reload()

    def _repo_reload(self):
        """
        data/estimates/*.json 을 읽어 '리포지토리' 탭 그리드에 뿌린다.
        - 최상위가 dict 또는 list 모두 허용
        - 필드 누락/형식 오류는 안전 스킵
        """
        import os, glob, json, time
        from datetime import datetime

        # 경로 고정: data/estimates
        try:
            base = EST_DIR  # estimate_policy.EST_DIR 또는 모듈 상단 EST_DIR
        except Exception:
            base = os.path.join(os.path.dirname(__file__), "data", "estimates")
        os.makedirs(base, exist_ok=True)

        rows = []
        files = sorted(glob.glob(os.path.join(base, "*.json")))
        for fp in files:
            try:
                with open(fp, "r", encoding="utf-8") as f:
                    data = json.load(f)
            except Exception:
                # 파싱 실패 파일은 스킵
                continue

            # 레코드 시퀀스로 정규화
            if isinstance(data, dict):
                records = [data]
            elif isinstance(data, list):
                # list 안의 요소가 dict인 경우만 유효
                records = [x for x in data if isinstance(x, dict)]
                if not records:
                    continue
            else:
                continue

            # 파일 mtime 백업 타임스탬프
            try:
                mtime = os.path.getmtime(fp)
                fallback_ts = datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                fallback_ts = ""

            for rec in records:
                # 안전 추출 유틸
                def _g(d, *keys, default=""):
                    cur = d
                    for k in keys:
                        if not isinstance(cur, dict):
                            return default
                        cur = cur.get(k)
                    return cur if cur is not None else default

                client_name = _g(rec, "client", "업체명") or rec.get("client_name", "") or ""
                project     = rec.get("project", "") or _g(rec, "meta", "project", default="")
                ts          = rec.get("timestamp", "") or fallback_ts

                # 라인/요약(있으면)
                try:
                    lines = rec.get("lines", [])
                    if isinstance(lines, list):
                        line_cnt = len(lines)
                    else:
                        line_cnt = 0
                except Exception:
                    line_cnt = 0

                # 차단기/외함 주요 스펙 요약(선택적)
                main_spec = ""
                enc_spec  = ""
                try:
                    main_spec = _g(rec, "main", "breaker", default="") or _g(rec, "main", "model", default="")
                except Exception:
                    pass
                try:
                    enc_spec = _g(rec, "enclosure", "size", default="") or _g(rec, "enclosure", "spec", default="")
                except Exception:
                    pass

                rows.append({
                    "file": os.path.basename(fp),
                    "client": client_name,
                    "project": project,
                    "timestamp": ts,
                    "lines": line_cnt,
                    "main": main_spec,
                    "enclosure": enc_spec,
                    "path": fp,
                })

        # 테이블 바인딩(기존 위젯 변수에 맞춰 세팅)
        # 예: self.repo_table.clear(); self.repo_table.insert(...)
        # 아래는 최소 공통 인터페이스 예시 — 실제 위젯 API에 맞게 그대로 연결
        try:
            self.repo_table.delete(*self.repo_table.get_children())
        except Exception:
            pass

        for r in rows:
            try:
                self.repo_table.insert(
                    "", "end",
                    values=(
                        r.get("file",""),
                        r.get("client",""),
                        r.get("project",""),
                        r.get("timestamp",""),
                        r.get("lines",0),
                        r.get("main",""),
                        r.get("enclosure",""),
                    )
                )
            except Exception:
                continue

        # 상태 라벨/카운트 갱신(있다면)
        try:
            self.repo_status_var.set(f"loaded: {len(rows)} files")
        except Exception:
            pass


    def _repo_on_select(self,_=None):
        item=self._repo_selected_item()
        if not item: return
        data=load_json(item["path"],{})
        st=data.get("status", {"order":False,"ship":False,"pay":"미입금"})
        self.var_repo_order.set(bool(st.get("order",False)))
        self.var_repo_ship.set(bool(st.get("ship",False)))
        self.var_repo_pay.set(st.get("pay","미입금"))

    def _repo_selected_item(self):
        sel=self.tree_repo.selection()
        if not sel: return None
        v=self.tree_repo.item(sel[0],"values")
        return {"time":v[0], "client":v[1], "path":v[6]}

    def _repo_update_status(self):
        it=self._repo_selected_item()
        if not it: return
        d=load_json(it["path"],{})
        d["status"]={"order": self.var_repo_order.get(), "ship":self.var_repo_ship.get(), "pay":self.var_repo_pay.get()}
        save_json(it["path"], d)
        messagebox.showinfo("저장","상태 저장")
        self._repo_reload()

    def _repo_search_and_focus(self, name: str):
        if not name:
            return
        for iid in self.tree_repo.get_children():
            values = self.tree_repo.item(iid, "values")
            client_name = values[1] if values and len(values) > 1 else ""
            if name in client_name:
                self.tree_repo.selection_set(iid)
                self.tree_repo.focus(iid)
                self.tree_repo.see(iid)
                self._repo_on_select()
                return
        messagebox.showinfo("검색", f"'{name}' 업체를 찾을 수 없습니다.")

    def _repo_add_item(self, data, path):
        st=data.get("status",{"order":False,"ship":False,"pay":"미입금"})
        self.tree_repo.insert("", "end", values=(
            data["timestamp"], data["client"].get("업체명",""),
            "O" if st.get("order") else "",
            "O" if st.get("ship") else "",
            st.get("pay","미입금"),
            data.get("memo",""),
            path
        ))

    def _repo_save_excel(self):
        it=self._repo_selected_item()
        if not it: messagebox.showwarning("선택 없음","저장할 견적 선택"); return
        from openpyxl import Workbook
        d=load_json(it["path"],{})
        wb=Workbook(); ws=wb.active; ws.title="견적"
        ws.append(["No","품명","규격","단위","수량","단가","금액"])
        for L in d.get("lines",[]): ws.append([L["no"],L["품명"],L["규격"],L["단위"],L["수량"],L["단가"],L["금액"]])
        out=filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel",".xlsx")],
                                         initialfile=f"repo_{now_tag()}.xlsx")
        if not out: return
        wb.save(out); messagebox.showinfo("저장", out)

    def _repo_save_csv(self):
        it=self._repo_selected_item()
        if not it:
            messagebox.showwarning("선택 없음","저장할 견적 선택");
            return
        d=load_json(it["path"],{})
        out=filedialog.asksaveasfilename(defaultextension=".csv",
                                         filetypes=[("CSV",".csv")],
                                         initialfile=f"repo_{now_tag()}.csv")
        if not out: return
        import csv
        with open(out,"w",newline="",encoding="utf-8-sig") as f:
            w=csv.writer(f); w.writerow(["No","품명","규격","단위","수량","단가","금액","비고"])
            for L in d.get("lines",[]):
                w.writerow([L.get("no",""),L.get("품명",""),L.get("규격",""),
                            L.get("단위",""),L.get("수량",""),L.get("단가",""),
                            L.get("금액",""),L.get("비고","")])
        messagebox.showinfo("저장", out)

    # ========= [탭5] 단가관리 =========
    def _build_tab_prices(self):
        tab=ttk.Frame(self.nb, padding=8); self.nb.add(tab, text="단가관리")
        bar=ttk.Frame(tab); bar.pack(fill="x")
        ttk.Button(bar,text="외함 단가 첨부", command=lambda:self._price_attach("enclosure")).pack(side="left")
        ttk.Button(bar,text="차단기 단가 첨부", command=lambda:self._price_attach("breaker")).pack(side="left", padx=6)
        ttk.Button(bar,text="부속자재 단가 첨부", command=lambda:self._price_attach("accessory")).pack(side="left")
        ttk.Label(bar,text="  (여기에 저장: data/prices)").pack(side="left", padx=8)

        self.txt_prices=scrolledtext.ScrolledText(tab, wrap="none", height=24);
        self.txt_prices.pack(fill="both",expand=True,pady=(8,0))
        self._load_prices_to_editor()

        info=ttk.Label(tab, text="※ 메커니즘: 이 탭은 파일을 data/prices 폴더에 보관합니다.\n"
                                 "   - 파일명 접두어로 종류를 구분: enclosure_*, breaker_*, accessory_*\n"
                                 "   - 이후 매칭 모듈이 최신 파일을 읽어 단가 조회(RAG/룰) 기반으로 사용합니다.",
                       justify="left")
        info.pack(fill="x", pady=(6,0))

    def _price_attach(self, kind):
        path=filedialog.askopenfilename(title="단가 파일",
                                        filetypes=[("Excel/CSV","*.xlsx;*.csv"),("모든 파일","*.*")])
        if not path: return
        dest=os.path.join(PRICES_DIR, f"{kind}_{os.path.basename(path)}")
        try:
            import shutil; shutil.copy2(path, dest)
        except Exception as e:
            messagebox.showerror("오류", f"복사 실패: {e}");
            return
        messagebox.showinfo("저장", f"{dest} 저장")
        self._load_prices_to_editor()

    def _load_prices_to_editor(self):
        self.txt_prices.delete("1.0",tk.END)
        self.txt_prices.insert(tk.END, "# 단가 파일 목록\n")
        files=sorted(os.listdir(PRICES_DIR)) if os.path.exists(PRICES_DIR) else []
        if not files:
            self.txt_prices.insert(tk.END, "- (없음)\n")
            return
        for fn in files:
            self.txt_prices.insert(tk.END, f"- {fn}\n")

    # ========= [탭6] 거래처관리 (모달로 추가 + 분석 요약) =========
    def _build_tab_clients(self):
        tab=ttk.Frame(self.nb, padding=8); self.nb.add(tab, text="거래처관리")
        top=ttk.Frame(tab); top.pack(fill="x")
        ttk.Button(top, text="거래처 추가", command=self._client_add_modal).pack(side="left")
        ttk.Label(top, text="  검색:").pack(side="left", padx=(10,4))
        self.ent_client_search=ttk.Entry(top, width=24); self.ent_client_search.pack(side="left")
        ttk.Button(top, text="찾기", command=self._clients_reload).pack(side="left", padx=4)

        cols=("name","owner","phone","email","biz","type","stats")
        self.tree_clients=ttk.Treeview(tab, columns=cols, show="headings", height=20)
        headers=["업체명","대표자","연락처","이메일","사업자번호","구분(매출/매입)","분석 요약"]
        widths =[180,80,120,200,120,120,460]
        for c,h,w in zip(cols,headers,widths):
            self.tree_clients.heading(c, text=h)
            self.tree_clients.column(c, width=w)
        self.tree_clients.pack(fill="both", expand=True, pady=(8,0))
        self._clients_reload()

    def _client_add_modal(self):
        win=tk.Toplevel(self); win.title("거래처 추가"); win.grab_set()
        frm=ttk.Frame(win, padding=12); frm.pack(fill="both", expand=True)
        def row(r,lbl,w=26,show=None):
            ttk.Label(frm,text=lbl).grid(row=r,column=0,sticky="e",pady=2)
            e=ttk.Entry(frm,width=w, show=show); e.grid(row=r,column=1,sticky="w",pady=2)
            return e
        e_name=row(0,"업체명*"); e_owner=row(1,"대표자"); e_phone=row(2,"연락처")
        e_email=row(3,"이메일", w=36); e_addr=row(4,"주소", w=36)
        e_fax=row(5,"팩스"); e_biz=row(6,"사업자등록번호")
        ttk.Label(frm,text="구분").grid(row=7,column=0,sticky="e",pady=2)
        cbo_type=ttk.Combobox(frm, values=["매출처","매입처","겸용"], state="readonly", width=10);
        cbo_type.set("매출처"); cbo_type.grid(row=7,column=1,sticky="w",pady=2)
        ttk.Label(frm,text="비고").grid(row=8,column=0,sticky="e",pady=2)
        e_note=ttk.Entry(frm, width=36); e_note.grid(row=8,column=1,sticky="w",pady=2)

        btns=ttk.Frame(frm); btns.grid(row=9,column=0,columnspan=2, pady=(10,0))
        def _ok():
            name=e_name.get().strip()
            if not name:
                messagebox.showwarning("필수","업체명은 필수입니다.")
                return
            db=load_json(CLIENTS_JSON, default={"clients":[]})
            db["clients"].append({
                "name":name, "owner":e_owner.get().strip(), "phone":e_phone.get().strip(),
                "email":e_email.get().strip(), "addr":e_addr.get().strip(), "fax":e_fax.get().strip(),
                "biz":e_biz.get().strip(), "type":cbo_type.get(), "note":e_note.get().strip()
            })
            save_json(CLIENTS_JSON, db)
            self._clients_reload()
            win.destroy()
        ttk.Button(btns,text="확인", command=_ok).pack(side="left", padx=6)
        ttk.Button(btns,text="취소", command=win.destroy).pack(side="left")

    def _clients_reload(self):
        db=load_json(CLIENTS_JSON, default={"clients":[]})
        q=(self.ent_client_search.get().strip() if hasattr(self,"ent_client_search") else "")
        self.tree_clients.delete(*self.tree_clients.get_children())
        for c in db["clients"]:
            if q and q not in c.get("name",""):
                continue
            stat=self._calc_client_stats(c.get("name"))
            summary=(f"견적 {stat['estimate_count']}건, 발주 {stat['order_count']}회, "
                     f"주문율 {stat['order_rate']}%, 최근6M {stat['sales_6m']:,}원, 최근1Y {stat['sales_1y']:,}원, "
                     f"입금상태 {stat['last_pay'] or '미입금'}")
            self.tree_clients.insert("", "end", values=(
                c.get("name",""), c.get("owner",""), c.get("phone",""),
                c.get("email",""), c.get("biz",""), c.get("type",""), summary
            ))

    def _calc_client_stats(self, name=None):
        # 간단 통계 (파일 타임스탬프 기반 최근 매출 집계)
        import datetime
        est_files=[os.path.join(EST_DIR,fn) for fn in os.listdir(EST_DIR) if fn.endswith(".json")]
        ests=[load_json(p,{}) for p in est_files]
        if name: ests=[e for e in ests if (e.get("client",{}).get("업체명","")==name)]
        if not ests and name:
            return {"name":name,"estimate_count":0,"order_count":0,"order_rate":0,"sales_6m":0,"sales_1y":0,"last_pay":None}
        order_count=0; last_pay=None
        sales_6m=sales_1y=0; now=datetime.datetime.now()
        for e in ests:
            st=e.get("status",{})
            if st.get("order"): order_count+=1
            if st.get("pay"): last_pay=st.get("pay")
            # 금액 합계
            subtotal=0
            for L in e.get("lines",[]):
                try: subtotal+=int(float(L.get("금액",0)))
                except: pass
            # 기간 분배 (파일에 timestamp가 문자열로 저장되어 있음)
            ts=e.get("timestamp","")
            try:
                dt=datetime.datetime.strptime(ts,"%Y-%m-%d %H:%M:%S")
            except:
                dt=now
            if (now-dt).days<=183: sales_6m+=subtotal
            if (now-dt).days<=365: sales_1y+=subtotal
        est_count=len(ests)
        order_rate= round((order_count/max(1,est_count))*100)
        return {"name":name or "(전체)","estimate_count":est_count,"order_count":order_count,
                "order_rate":order_rate,"sales_6m":sales_6m,"sales_1y":sales_1y,"last_pay":last_pay}

    # ========= [탭7] 발주·출고 (골격) =========
    def _build_tab_shipping(self):
        tab=ttk.Frame(self.nb, padding=8); self.nb.add(tab, text="발주·출고")
        ttk.Label(tab,text="발주/출고 진행 현황 (개발 중)").pack(anchor="w")

    # ========= [탭8] 보고서 =========
    def _build_tab_reports(self):
        tab=ttk.Frame(self.nb, padding=8); self.nb.add(tab, text="보고서")
        frm=ttk.LabelFrame(tab, text="보고서 생성", padding=8); frm.pack(fill="x")
        ttk.Label(frm,text="기간(예: 2025-08-01~2025-08-31)").grid(row=0,column=0)
        self.ent_report_range=ttk.Entry(frm, width=28); self.ent_report_range.grid(row=0,column=1,padx=6)
        ttk.Button(frm,text="일일", command=lambda:self._report_make("daily")).grid(row=0,column=2,padx=4)
        ttk.Button(frm,text="주간", command=lambda:self._report_make("weekly")).grid(row=0,column=3,padx=4)
        ttk.Button(frm,text="월간", command=lambda:self._report_make("monthly")).grid(row=0,column=4,padx=4)
        self.txt_report=scrolledtext.ScrolledText(tab, wrap="word", height=22); self.txt_report.pack(fill="both",expand=True,pady=(8,0))

    def _report_make(self, kind):
        self.txt_report.delete("1.0",tk.END)
        self.txt_report.insert(tk.END, f"[{kind}] 보고서(모의)\n- 기간: {(self.ent_report_range.get().strip() or '지정 없음')}\n")
        cnt=len([fn for fn in os.listdir(EST_DIR) if fn.endswith('.json')])
        self.txt_report.insert(tk.END, f"- 누적 견적: {cnt} 건\n"); self.txt_report.see(tk.END)

    # ========= [탭9] 설정 =========
    def _build_tab_settings(self):
        tab=ttk.Frame(self.nb, padding=8); self.nb.add(tab, text="설정")

        # SMTP(좌) + 네이버웍스(우)
        top=ttk.Frame(tab); top.pack(fill="x")
        g1=ttk.LabelFrame(top, text="SMTP 메일", padding=8); g1.pack(side="left", fill="x", expand=True, padx=(0,6))
        self.ent_smtp_host=ttk.Entry(g1,width=22); self.ent_smtp_port=ttk.Entry(g1,width=6)
        self.ent_smtp_user=ttk.Entry(g1,width=22); self.ent_smtp_pass=ttk.Entry(g1,width=22, show="*")
        self.ent_smtp_sender=ttk.Entry(g1,width=28)
        ttk.Label(g1,text="Host").grid(row=0,column=0); self.ent_smtp_host.grid(row=0,column=1,padx=4)
        ttk.Label(g1,text="Port").grid(row=0,column=2); self.ent_smtp_port.grid(row=0,column=3,padx=4)
        ttk.Label(g1,text="User").grid(row=1,column=0); self.ent_smtp_user.grid(row=1,column=1,padx=4)
        ttk.Label(g1,text="Pass").grid(row=1,column=2); self.ent_smtp_pass.grid(row=1,column=3,padx=4)
        ttk.Label(g1,text="Sender").grid(row=2,column=0); self.ent_smtp_sender.grid(row=2,column=1,padx=4, columnspan=3, sticky="we")

        gNW=ttk.LabelFrame(top, text="네이버웍스 (ID + API Key)", padding=8); gNW.pack(fill="x", expand=True)
        self.ent_nw_id=ttk.Entry(gNW, width=28); self.ent_nw_key=ttk.Entry(gNW, width=28, show="*")
        ttk.Label(gNW,text="ID").grid(row=0,column=0); self.ent_nw_id.grid(row=0,column=1,padx=4)
        ttk.Label(gNW,text="API Key").grid(row=1,column=0); self.ent_nw_key.grid(row=1,column=1,padx=4)

        gIMAP = ttk.LabelFrame(tab, text="네이버웍스 메일(IMAP 인증)", padding=8); gIMAP.pack(fill="x", pady=(8, 0))
        self.ent_imap_host = ttk.Entry(gIMAP, width=22)
        self.ent_imap_port = ttk.Entry(gIMAP, width=6)
        self.ent_imap_user = ttk.Entry(gIMAP, width=26)
        self.ent_imap_pass = ttk.Entry(gIMAP, width=26, show="*")
        self.var_imap_ssl  = tk.BooleanVar(value=True)
        ttk.Label(gIMAP, text="Host").grid(row=0, column=0, sticky="e"); self.ent_imap_host.grid(row=0, column=1, padx=4, sticky="w")
        ttk.Label(gIMAP, text="Port").grid(row=0, column=2, sticky="e"); self.ent_imap_port.grid(row=0, column=3, padx=4, sticky="w")
        ttk.Label(gIMAP, text="User").grid(row=1, column=0, sticky="e"); self.ent_imap_user.grid(row=1, column=1, padx=4, sticky="w")
        ttk.Label(gIMAP, text="App Password").grid(row=1, column=2, sticky="e"); self.ent_imap_pass.grid(row=1, column=3, padx=4, sticky="w")
        ttk.Checkbutton(gIMAP, text="SSL(권장)", variable=self.var_imap_ssl).grid(row=0, column=4, padx=8)
        ttk.Button(gIMAP, text="IMAP 연결테스트", command=self._test_imap).grid(row=1, column=4, padx=8)

        # 외부 서비스(Only ChatGPT/Claude)
        g2=ttk.LabelFrame(tab, text="외부 서비스 (AI)", padding=8); g2.pack(fill="x", pady=(8,0))
        ttk.Label(g2,text="ChatGPT URL").grid(row=0,column=0,sticky="e");
        self.ent_gpt_url=ttk.Entry(g2,width=48); self.ent_gpt_url.grid(row=0,column=1,padx=4,sticky="w")
        ttk.Label(g2,text="API Key").grid(row=0,column=2,sticky="e");
        self.ent_gpt_key=ttk.Entry(g2,width=36, show="*"); self.ent_gpt_key.grid(row=0,column=3,padx=4,sticky="w")
        ttk.Button(g2,text="연결테스트", command=lambda:self._test_ai("chatgpt")).grid(row=0,column=4,padx=6)

        ttk.Label(g2,text="Claude URL").grid(row=1,column=0,sticky="e");
        self.ent_claude_url=ttk.Entry(g2,width=48); self.ent_claude_url.grid(row=1,column=1,padx=4,sticky="w")
        ttk.Label(g2,text="API Key").grid(row=1,column=2,sticky="e");
        self.ent_claude_key=ttk.Entry(g2,width=36, show="*"); self.ent_claude_key.grid(row=1,column=3,padx=4,sticky="w")
        ttk.Button(g2,text="연결테스트", command=lambda:self._test_ai("claude")).grid(row=1,column=4,padx=6)

        g3=ttk.LabelFrame(tab, text="로컬 도구", padding=8); g3.pack(fill="x", pady=(8,0))
        self.ent_tess=ttk.Entry(g3, width=60)
        ttk.Label(g3,text="Tesseract 경로").grid(row=0,column=0); self.ent_tess.grid(row=0,column=1, padx=4, sticky="we")

        ttk.Button(tab,text="설정 불러오기", command=self._cfg_load).pack(side="left", padx=(0,6), pady=(8,0))
        ttk.Button(tab,text="설정 저장", command=self._cfg_save).pack(side="left", pady=(8,0))
        ttk.Button(tab,text="설정 테스트(모의)", command=lambda: messagebox.showinfo("테스트","일부 모듈 테스트(모의) OK")).pack(side="left", padx=6, pady=(8,0))

        self._cfg_load()

    # ========== 공통 유틸 ==========

    def _http_session_with_retry(self):
        s = requests.Session()
        retry = Retry(total=3, backoff_factor=0.8,
                      status_forcelist=[429, 500, 502, 503, 504],
                      allowed_methods=["POST"])
        adapter = HTTPAdapter(max_retries=retry, pool_connections=4, pool_maxsize=4)
        s.mount("https://", adapter)
        s.mount("http://", adapter)
        return s

    def _call_llm(self, which: str, msg: str):
        try:
            sess = self._http_session_with_retry()
            if which=="chatgpt" and self.cfg.get("chatgpt",{}).get("ok"):
                url=self.cfg["chatgpt"]["url"]; key=self.cfg["chatgpt"]["api_key"]; model=self.cfg["chatgpt"].get("model","gpt-4o")
                headers={"Authorization": f"Bearer {key}","Content-Type":"application/json"}
                body={"model": model, "messages":[{"role":"user","content":msg}],
                      "temperature":0.2, "max_tokens":400}
                r = sess.post(url, headers=headers, json=body, timeout=(10, 60))
                if r.ok:
                    data=r.json()
                    content = data.get("choices",[{}])[0].get("message",{}).get("content")
                    return content or "응답 없음"
                try:    return f"ChatGPT 응답 실패: {r.status_code} {r.json()}"
                except: return f"ChatGPT 응답 실패: {r.status_code} {r.text[:300]}"

            if which=="claude" and self.cfg.get("claude",{}).get("ok"):
                url=self.cfg["claude"]["url"]; key=self.cfg["claude"]["api_key"]; model=self.cfg["claude"].get("model","claude-3-7-sonnet-20250219")
                headers={"x-api-key":key,"anthropic-version":"2023-06-01","content-type":"application/json"}
                body={"model":model,"max_tokens":400,"messages":[{"role":"user","content":msg}]}
                r = sess.post(url, headers=headers, json=body, timeout=(10, 60))
                if r.ok:
                    data=r.json()
                    content = "".join([b.get("text","") for b in data.get("content",[]) if b.get("type")=="text"])
                    return content or "응답 없음"
                try:    return f"Claude 응답 실패: {r.status_code} {r.json()}"
                except: return f"Claude 응답 실패: {r.status_code} {r.text[:400]}"
        except requests.exceptions.ReadTimeout:
            return "네트워크 지연으로 응답 시간이 초과됐습니다. 잠시 후 다시 시도해주세요."
        except Exception as e:
            return f"연결 오류: {e}"
        return None

    def _test_ai(self, which: str):
        if which=="chatgpt":
            url=self.ent_gpt_url.get().strip(); key=self.ent_gpt_key.get().strip()
            if not url or not key:
                messagebox.showwarning("필수","URL과 API Key 입력"); return
            try:
                headers={"Authorization": f"Bearer {key}","Content-Type":"application/json"}
                body={"model": self.cfg["chatgpt"].get("model","gpt-4o"),
                      "messages":[{"role":"user","content":"ping"}], "max_tokens":10}
                r=requests.post(url, headers=headers, json=body, timeout=12)
                self.cfg["chatgpt"]["ok"]=bool(r.ok); save_json(CFG_PATH,self.cfg)
                if r.ok: messagebox.showinfo("성공","ChatGPT 연결 성공")
                else:
                    try: err=r.json()
                    except: err={"text": r.text[:300]}
                    messagebox.showerror("실패", f"HTTP {r.status_code}\n{err}")
            except Exception as e:
                self.cfg["chatgpt"]["ok"]=False; save_json(CFG_PATH,self.cfg)
                messagebox.showerror("오류", f"{e}")
        else:
            url=self.ent_claude_url.get().strip() or "https://api.anthropic.com/v1/messages"
            key=self.ent_claude_key.get().strip()
            if not url or not key:
                messagebox.showwarning("필수","URL과 API Key 입력"); return
            try:
                headers={"x-api-key":key,"anthropic-version":"2023-06-01","content-type":"application/json"}
                body={"model": self.cfg["claude"].get("model","claude-3-7-sonnet-20250219"),
                      "max_tokens":32,"messages":[{"role":"user","content":"ping"}]}
                r=requests.post(url, headers=headers, json=body, timeout=20)
                self.cfg["claude"]["ok"]=bool(r.ok); save_json(CFG_PATH,self.cfg)
                if r.ok: messagebox.showinfo("성공","Claude 연결 성공")
                else:
                    try: err=r.json()
                    except: err={"text": r.text[:500]}
                    messagebox.showerror("실패", f"HTTP {r.status_code}\n{err}")
            except Exception as e:
                self.cfg["claude"]["ok"]=False; save_json(CFG_PATH,self.cfg)
                messagebox.showerror("오류", f"{e}")

    def _test_imap(self):
        host = self.ent_imap_host.get().strip() or "imap.naver.com"
        port = int(self.ent_imap_port.get().strip() or "993")
        user = self.ent_imap_user.get().strip()
        pw   = self.ent_imap_pass.get().strip()
        use_ssl = bool(self.var_imap_ssl.get())
        if not (host and user and pw):
            messagebox.showwarning("IMAP", "Host / User / App Password를 입력하세요.")
            return
        try:
            M = imaplib.IMAP4_SSL(host, port) if use_ssl else imaplib.IMAP4(host, port)
            M.login(user, pw); M.select("INBOX")
            typ, _ = M.search(None, "ALL"); M.logout()
            ok = (typ=="OK")
            self.cfg.setdefault("imap", {})["ok"] = ok
            save_json(CFG_PATH, self.cfg)
            messagebox.showinfo("성공" if ok else "실패",
                                "IMAP 연결 성공 (INBOX 접근 확인)" if ok else "IMAP 검색 실패 (권한/설정 확인)")
        except Exception as e:
            self.cfg.setdefault("imap", {})["ok"] = False
            save_json(CFG_PATH, self.cfg)
            messagebox.showerror("IMAP 오류", str(e))

    def _cfg_load(self):
        cfg=self.cfg
        s=cfg.get("smtp",{})
        self.ent_smtp_host.delete(0,tk.END); self.ent_smtp_host.insert(0,s.get("host",""))
        self.ent_smtp_port.delete(0,tk.END); self.ent_smtp_port.insert(0,s.get("port",587))
        self.ent_smtp_user.delete(0,tk.END); self.ent_smtp_user.insert(0,s.get("user",""))
        self.ent_smtp_pass.delete(0,tk.END); self.ent_smtp_pass.insert(0,s.get("pass",""))
        self.ent_smtp_sender.delete(0,tk.END); self.ent_smtp_sender.insert(0,s.get("sender",""))

        im = cfg.get("imap", {})
        self.ent_imap_host.delete(0, tk.END); self.ent_imap_host.insert(0, im.get("host", "imap.naver.com"))
        self.ent_imap_port.delete(0, tk.END); self.ent_imap_port.insert(0, im.get("port", 993))
        self.ent_imap_user.delete(0, tk.END); self.ent_imap_user.insert(0, im.get("user", ""))
        self.ent_imap_pass.delete(0, tk.END); self.ent_imap_pass.insert(0, im.get("pass", ""))
        self.var_imap_ssl.set(bool(im.get("ssl", True)))

        nw=cfg.get("naver_works",{})
        self.ent_nw_id.delete(0,tk.END); self.ent_nw_id.insert(0,nw.get("id",""))
        self.ent_nw_key.delete(0,tk.END); self.ent_nw_key.insert(0,nw.get("api_key",""))

        # AI
        self.ent_gpt_url.delete(0,tk.END); self.ent_gpt_url.insert(0, cfg.get("chatgpt",{}).get("url",""))
        self.ent_gpt_key.delete(0,tk.END); self.ent_gpt_key.insert(0, cfg.get("chatgpt",{}).get("api_key",""))
        self.ent_claude_url.delete(0,tk.END); self.ent_claude_url.insert(0, cfg.get("claude",{}).get("url",""))
        self.ent_claude_key.delete(0,tk.END); self.ent_claude_key.insert(0, cfg.get("claude",{}).get("api_key",""))

        self.ent_tess.delete(0,tk.END); self.ent_tess.insert(0, cfg.get("tesseract_path",""))

    def _cfg_save(self):
        try:
            self.cfg["smtp"]={"host": self.ent_smtp_host.get().strip(),
                              "port": int(self.ent_smtp_port.get().strip() or "587"),
                              "user": self.ent_smtp_user.get().strip(),
                              "pass": self.ent_smtp_pass.get().strip(),
                              "sender": self.ent_smtp_sender.get().strip()}
            self.cfg["imap"] = {
                "host": self.ent_imap_host.get().strip() or "imap.naver.com",
                "port": int(self.ent_imap_port.get().strip() or "993"),
                "user": self.ent_imap_user.get().strip(),
                "pass": self.ent_imap_pass.get().strip(),
                "ssl":  bool(self.var_imap_ssl.get()),
                "ok":   bool(self.cfg.get("imap", {}).get("ok", False))
            }
            self.cfg["naver_works"]={"id": self.ent_nw_id.get().strip(),
                                     "api_key": self.ent_nw_key.get().strip()}
            self.cfg.setdefault("chatgpt",{})
            self.cfg["chatgpt"]["url"]=self.ent_gpt_url.get().strip()
            self.cfg["chatgpt"]["api_key"]=self.ent_gpt_key.get().strip()
            self.cfg.setdefault("claude",{})
            self.cfg["claude"]["url"]=self.ent_claude_url.get().strip()
            self.cfg["claude"]["api_key"]=self.ent_claude_key.get().strip()
            self.cfg["tesseract_path"]=self.ent_tess.get().strip()
            save_json(CFG_PATH, self.cfg); messagebox.showinfo("저장","설정 저장 완료")
        except Exception as e:
            messagebox.showerror("오류", f"설정 저장 실패: {e}")

# === Multi-panel (분전반 여러 개) 블록 탐지 유틸 ===
from openpyxl import load_workbook

_KEY_HEADERS = {"품명","규격","단위","수량","단가","금액"}
_KW_SUBTOTAL = {"소계","소   계","소  계","SUBTOTAL"}
_KW_TOTAL    = {"합계","합   계","TOTAL"}

def _nz(x:str) -> str:
    return str(x or "").strip().replace("\u3000", " ").replace("\xa0", " ")

def _target_sheets(wb):
    names = wb.sheetnames
    if len(names) >= 3:
        # 규칙: 3개 이상이면 1번+3번만 분석(2번=고압반 무시)
        return [names[0], names[2]]
    return names  # 2개면 1,2 모두

def _detect_header_row(ws, max_rows=200, max_cols=30):
    best_r, best_hits = None, 0
    R = min(ws.max_row, max_rows)
    C = min(ws.max_column, max_cols)
    for r in range(1, R+1):
        vals = {_nz(ws.cell(row=r, column=c).value) for c in range(1, C+1)}
        hits = len(_KEY_HEADERS & vals)
        if hits > best_hits:
            best_r, best_hits = r, hits
        if hits >= 4:
            return r, hits
    return best_r, best_hits

def _row_has_any(ws, r, keywords, max_cols=30):
    C = min(ws.max_column, max_cols)
    for c in range(1, C+1):
        v = _nz(ws.cell(row=r, column=c).value)
        if not v:
            continue
        for k in keywords:
            if k in v:
                return True
    return False

def _is_blank_row(ws, r, max_cols=20):
    C = min(ws.max_column, max_cols)
    for c in range(1, C+1):
        if _nz(ws.cell(row=r, column=c).value):
            return False
    return True

def _next_nonempty_row(ws, from_row, max_seek=6, max_cols=20):
    rmax = min(ws.max_row, from_row + max_seek)
    for r in range(from_row, rmax+1):
        if not _is_blank_row(ws, r, max_cols=max_cols):
            return r
    return None

def detect_panel_blocks_in_sheet(ws, *, max_rows=500, max_cols=40):
    """
    반환: blocks, meta
      blocks = [(start_row, end_row), ...]  # 각 분전반 블록의 데이터 범위(헤더 제외)
      meta   = {"header_row": int|None, "ends": [int,...]}
    규칙:
      - 헤더(품명/규격/단위/수량/단가/금액) 행을 찾고, 그 다음 행부터 데이터로 간주
      - 블록 종료 행: '합계'(우선) 없으면 '소계'
      - 다음 블록 시작: 종료행 이후 공백 1행(±1~2 허용) 다음의 첫 비어있지 않은 행
    """
    R = min(ws.max_row, max_rows)
    C = min(ws.max_column, max_cols)

    header_row, header_hits = _detect_header_row(ws, max_rows=max_rows, max_cols=max_cols)
    if not header_row:
        header_row = 1  # 헤더를 못 찾으면 1행부터 스캔

    # 종료 앵커 수집
    end_rows = []
    for r in range(1, R+1):
        if _row_has_any(ws, r, _KW_TOTAL, max_cols=max_cols):
            end_rows.append(r)
    if not end_rows:
        for r in range(1, R+1):
            if _row_has_any(ws, r, _KW_SUBTOTAL, max_cols=max_cols):
                end_rows.append(r)
    end_rows = sorted(set(end_rows))

    # 시작점: 헤더 다음 줄
    cur_start = header_row + 1
    blocks = []
    for end_r in end_rows:
        if cur_start < end_r:
            blocks.append((cur_start, end_r))
        # 다음 블록 시작 = 종료 후 공백 1행(±2 허용) 뒤 첫 non-empty
        next_start = None
        for off in (1,2,3):
            r = end_r + off
            if r > ws.max_row:
                break
            if _is_blank_row(ws, r):
                nn = _next_nonempty_row(ws, r+1, max_seek=6)
                if nn:
                    next_start = nn
                break
        if not next_start:
            next_start = end_r + 2
        cur_start = next_start

    # 파일 끝 넘어간 블록 제거
    blocks = [(s,e) for (s,e) in blocks if s <= e and s <= R]
    return blocks, {"header_row": header_row, "header_hits": header_hits, "ends": end_rows}

def detect_panel_blocks_in_workbook(xlsx_path, *, row_cap=600, col_cap=40):
    """
    탭 규칙 적용:
      - 시트 2개: 둘 다
      - 시트 3개 이상: 1,3번만(2번=고압반 무시)
    """
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    targets = _target_sheets(wb)
    report = []
    for name in targets:
        ws = wb[name]
        blocks, meta = detect_panel_blocks_in_sheet(ws, max_rows=row_cap, max_cols=col_cap)
        if blocks:
            for idx, (s,e) in enumerate(blocks,  start=1):
                report.append({"sheet": name, "block_idx": idx, "start_row": s, "end_row": e,
                               "header_row": meta["header_row"], "end_rows": meta["ends"]})
        else:
            report.append({"sheet": name, "block_idx": None, "start_row": None, "end_row": None,
                           "header_row": meta["header_row"], "end_rows": meta["ends"]})
    return report
# === /Multi-panel 유틸 ===

# -------- 실행 --------
def main():
    app = None
    try:
        app = App()
        app.mainloop()
    except KeyboardInterrupt:
        print("[UI] 사용자 인터럽트로 종료합니다.")
    except SystemExit:
        pass
    finally:
        # 로그 플러시 추가
        _flush_log("Application exiting")
        try:
            if app is not None and app.winfo_exists():
                app.destroy()
        except Exception:
            pass

    # === [PATCH START] system_estimate_ui.py — 교육용탭(Review & Train) 추가 ===
    # 사용 프레임워크: Tkinter/ttk 가정
    # 기존 App 클래스의 Notebook에 "교육용탭"을 추가하고,
    # 좌: 미니 배치 캔버스 / 우: JSON 편집 + 주석 / 하단: 오토세이브 로그 표시를 구현.
    # 기존 견적 생성 흐름에서 on_estimate_ready_for_review(est_json, source_path) 한 줄만 호출하면 초기 JSON이 로딩됩니다.

    import os, json, time, uuid, threading
    from dataclasses import dataclass, field
    from typing import Any, Dict, List, Optional

    try:
        import tkinter as tk
        from tkinter import ttk, messagebox, filedialog
    except:
        tk = None
        ttk = None

    # ------------------------------------------------------------
    # 경로 유틸
    def _ensure_dir(path: str):
        os.makedirs(path, exist_ok=True)
        return path

    def _now_ts():
        return time.strftime("%Y%m%d_%H%M%S")

    # ------------------------------------------------------------
    # 스냅샷/이벤트 로깅
    @dataclass
    class EditEvent:
        t: float
        who: str
        action: str
        detail: Dict[str, Any] = field(default_factory=dict)

    @dataclass
    class ReviewSession:
        estimate_id: str
        bundle_dir: str
        source_path: Optional[str] = None
        initial_json: Dict[str, Any] = field(default_factory=dict)
        current_json: Dict[str, Any] = field(default_factory=dict)
        notes: List[str] = field(default_factory=list)
        canvas_data: Dict[str, Any] = field(default_factory=lambda: {"shapes": []})  # [{type, x1,y1,x2,y2, text, id}]
        events: List[EditEvent] = field(default_factory=list)
        autosave_interval_sec: int = 8
        _stop_flag: bool = False
        _autosave_thread: Optional[threading.Thread] = None

        def start_autosave(self, log_widget=None):
            if self._autosave_thread and self._autosave_thread.is_alive():
                return
            def _loop():
                while not self._stop_flag:
                    try:
                        self.save_snapshot(kind="autosave", log_widget=log_widget)
                    except Exception as e:
                        if log_widget is not None:
                            log_widget.insert(tk.END, f"[오토세이브 오류] {e}\n")
                    time.sleep(self.autosave_interval_sec)
            self._autosave_thread = threading.Thread(target=_loop, daemon=True)
            self._autosave_thread.start()

        def stop(self):
            self._stop_flag = True

        def _paths(self):
            snaps = _ensure_dir(os.path.join(self.bundle_dir, "snapshots"))
            logs = _ensure_dir(os.path.join(self.bundle_dir, "logs"))
            return snaps, logs

        def save_snapshot(self, kind="manual", log_widget=None):
            snaps, logs = self._paths()
            ts = _now_ts()
            # 파일명: estimate_id_kind_ts.json
            snap_path = os.path.join(snaps, f"{self.estimate_id}_{kind}_{ts}.json")
            data = {
                "estimate_id": self.estimate_id,
                "kind": kind,
                "ts": ts,
                "initial_json": self.initial_json,
                "current_json": self.current_json,
                "notes": self.notes,
                "canvas_data": self.canvas_data,
                "source_path": self.source_path,
            }
            with open(snap_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

            # 이벤트 로그 JSONL
            log_path = os.path.join(logs, f"{self.estimate_id}.jsonl")
            event_record = {
                "t": time.time(),
                "ts": ts,
                "who": "user",
                "action": f"snapshot:{kind}",
                "detail": {
                    "snap_path": snap_path,
                    "counts": {
                        "notes": len(self.notes),
                        "shapes": len(self.canvas_data.get("shapes", [])),
                    }
                }
            }
            with open(log_path, "a", encoding="utf-8") as f:
                f.write(json.dumps(event_record, ensure_ascii=False) + "\n")

            if log_widget is not None:
                log_widget.insert(tk.END, f"[{kind}] 스냅샷 저장: {os.path.basename(snap_path)}\n")
                log_widget.see(tk.END)

        def log_event(self, action: str, detail: Optional[Dict[str, Any]] = None, log_widget=None):
            _, logs = self._paths()
            ts = _now_ts()
            ev = {
                "t": time.time(),
                "ts": ts,
                "who": "user",
                "action": action,
                "detail": detail or {}
            }
            with open(os.path.join(logs, f"{self.estimate_id}.jsonl"), "a", encoding="utf-8") as f:
                f.write(json.dumps(ev, ensure_ascii=False) + "\n")
            if log_widget is not None:
                log_widget.insert(tk.END, f"[이벤트] {action}\n")
                log_widget.see(tk.END)

    # ------------------------------------------------------------
    # 교육용탭 UI 위젯
    class ReviewTrainTab(ttk.Frame):
        def __init__(self, master, app_ref, **kwargs):
            super().__init__(master, **kwargs)
            self.app = app_ref  # App 인스턴스 참조(기존 코드와 연동 포인트)
            self.session: Optional[ReviewSession] = None
            self._json_debounce_ms = 500
            self._json_debounce_after = None
            self._selected_shape_id = None

            # 레이아웃: 좌(캔버스) | 우(JSON/노트/버튼) | 하단(로그)
            self.columnconfigure(0, weight=1)
            self.columnconfigure(1, weight=1)
            self.rowconfigure(0, weight=1)
            self.rowconfigure(1, weight=0)

            # 좌측: 미니 배치 캔버스
            left = ttk.Frame(self)
            left.grid(row=0, column=0, sticky="nsew", padx=6, pady=6)
            left.rowconfigure(1, weight=1)
            ttk.Label(left, text="미니 배치 캔버스 (사각형/텍스트 배치, 드래그 이동)").grid(row=0, column=0, sticky="w")
            self.canvas = tk.Canvas(left, width=520, height=520, bg="#f7f7f9", highlightthickness=1, highlightbackground="#ccc")
            self.canvas.grid(row=1, column=0, sticky="nsew")

            # 캔버스 툴바
            tb = ttk.Frame(left)
            tb.grid(row=2, column=0, sticky="ew", pady=(6,0))
            ttk.Button(tb, text="사각형 추가", command=self._add_rect).pack(side="left")
            ttk.Button(tb, text="텍스트 추가", command=self._add_text).pack(side="left")
            ttk.Button(tb, text="선택 삭제", command=self._delete_selected).pack(side="left")
            ttk.Button(tb, text="정리(격자)", command=self._snap_grid).pack(side="left")

            # 드래그 이동 핸들러
            self.canvas.bind("<Button-1>", self._on_canvas_down)
            self.canvas.bind("<B1-Motion>", self._on_canvas_drag)

            # 우측: JSON + 노트 + 액션
            right = ttk.Frame(self)
            right.grid(row=0, column=1, sticky="nsew", padx=6, pady=6)
            right.rowconfigure(1, weight=1)
            right.rowconfigure(3, weight=1)

            ttk.Label(right, text="견적 JSON (사람/AI가 모두 이해하기 쉬운 스키마)").grid(row=0, column=0, sticky="w")
            self.json_text = tk.Text(right, height=16, wrap="none")
            self.json_text.grid(row=1, column=0, sticky="nsew")
            self.json_text.bind("<<Modified>>", self._on_json_modified)

            ttk.Label(right, text="주석 / 메모").grid(row=2, column=0, sticky="w", pady=(8,0))
            self.notes_text = tk.Text(right, height=8, wrap="word")
            self.notes_text.grid(row=3, column=0, sticky="nsew")

            btns = ttk.Frame(right)
            btns.grid(row=4, column=0, sticky="ew", pady=(8,0))
            ttk.Button(btns, text="가짜 견적 불러오기", command=self._load_fake).pack(side="left")
            ttk.Button(btns, text="스냅샷 저장", command=lambda: self._snapshot("manual")).pack(side="left")
            ttk.Button(btns, text="내보내기(JSON)", command=self._export_json).pack(side="left")
            ttk.Button(btns, text="불러오기(JSON)", command=self._import_json).pack(side="left")

            # 하단: 로그 뷰어
            self.log_text = tk.Text(self, height=6, bg="#111", fg="#ddd")
            self.log_text.grid(row=1, column=0, columnspan=2, sticky="nsew", padx=6, pady=(0,6))
            self._log("교육용탭 준비 완료. 견적 생성 후 on_estimate_ready_for_review(...) 호출 시 초기 JSON이 로딩됩니다.")

        # ---------- 로딩/세션 ----------
        def attach_session(self, session: ReviewSession):
            # 기존 세션 정리
            if self.session:
                self.session.stop()
            self.session = session
            self._log(f"세션 시작: estimate_id={session.estimate_id}")
            # JSON 표시
            self._render_json(session.current_json)
            # 노트 표시
            self.notes_text.delete("1.0", tk.END)
            self.notes_text.insert("1.0", "\n".join(session.notes))
            # 캔버스 표시
            self._render_canvas()
            # 오토세이브 시작
            session.start_autosave(log_widget=self.log_text)

        def on_estimate_ready_for_review(self, est_json: Dict[str, Any], source_path: Optional[str] = None):
            """
            기존 견적 생성 흐름에서 호출:
            ex) self.review_tab.on_estimate_ready_for_review(est_json, source_path=save_path)
            """
            estimate_id = est_json.get("estimate_id") or str(uuid.uuid4())
            bundle_dir = _ensure_dir(os.path.join("data", "training_edits", estimate_id))
            session = ReviewSession(
                estimate_id=estimate_id,
                bundle_dir=bundle_dir,
                source_path=source_path,
                initial_json=json.loads(json.dumps(est_json, ensure_ascii=False)),  # deep copy
                current_json=json.loads(json.dumps(est_json, ensure_ascii=False)),
            )
            self.attach_session(session)
            session.save_snapshot(kind="initial", log_widget=self.log_text)

        # ---------- JSON 편집 ----------
        def _render_json(self, data: Dict[str, Any]):
            self.json_text.delete("1.0", tk.END)
            self.json_text.insert("1.0", json.dumps(data, ensure_ascii=False, indent=2))
            self.json_text.edit_modified(False)

        def _on_json_modified(self, event=None):
            if self.json_text.edit_modified():
                # 디바운스
                if self._json_debounce_after:
                    self.after_cancel(self._json_debounce_after)
                self._json_debounce_after = self.after(self._json_debounce_ms, self._apply_json_from_editor)
                self.json_text.edit_modified(False)

        def _apply_json_from_editor(self):
            if not self.session:
                return
            txt = self.json_text.get("1.0", tk.END).strip()
            try:
                obj = json.loads(txt)
                self.session.current_json = obj
                self.session.log_event("json.update", {"keys": list(obj.keys())}, log_widget=self.log_text)
                # JSON 변경이 캔버스에 영향을 줄 수도 있으므로 필요 시 후처리
            except Exception as e:
                self._log(f"[경고] JSON 파싱 실패: {e}")

        # ---------- 노트/메모 ----------
        def _snapshot(self, kind="manual"):
            if not self.session:
                messagebox.showwarning("알림", "세션이 없습니다. 먼저 견적을 로딩하세요.")
                return
            # 노트 반영
            notes = self.notes_text.get("1.0", tk.END).strip().splitlines()
            self.session.notes = [n for n in notes if n.strip()]
            self.session.save_snapshot(kind=kind, log_widget=self.log_text)

        # ---------- 캔버스 ----------
        def _render_canvas(self):
            self.canvas.delete("all")
            # 격자
            for i in range(0, 520, 20):
                self.canvas.create_line(i, 0, i, 520, fill="#eee")
                self.canvas.create_line(0, i, 520, i, fill="#eee")
            # 기존 도형
            if not self.session:
                return
            for shp in self.session.canvas_data.get("shapes", []):
                if shp.get("type") == "rect":
                    x1,y1,x2,y2 = shp["x1"], shp["y1"], shp["x2"], shp["y2"]
                    rid = self.canvas.create_rectangle(x1,y1,x2,y2, outline="#333", width=2, fill="#dfe9ff")
                    shp["rid"] = rid
                elif shp.get("type") == "text":
                    x,y,tx = shp["x"], shp["y"], shp.get("text","")
                    tid = self.canvas.create_text(x,y, text=tx, anchor="nw", font=("Arial", 11))
                    shp["tid"] = tid

        def _add_rect(self):
            if not self.session:
                return
            shp = {"type":"rect", "x1":40, "y1":40, "x2":140, "y2":100}
            self.session.canvas_data["shapes"].append(shp)
            self.session.log_event("canvas.add_rect", {"pos":[40,40,140,100]}, log_widget=self.log_text)
            self._render_canvas()

        def _add_text(self):
            if not self.session:
                return
            shp = {"type":"text", "x":50, "y":120, "text":"분기 ELB 2P 20A x4"}
            self.session.canvas_data["shapes"].append(shp)
            self.session.log_event("canvas.add_text", {"pos":[50,120], "text": shp["text"]}, log_widget=self.log_text)
            self._render_canvas()

        def _delete_selected(self):
            if not self.session or self._selected_shape_id is None:
                return
            # 선택 로직: 간단히 마지막 클릭 대상 제거
            del_idx = None
            for i, shp in enumerate(self.session.canvas_data.get("shapes", [])):
                if shp.get("_selected"):
                    del_idx = i
                    break
            if del_idx is not None:
                self.session.canvas_data["shapes"].pop(del_idx)
                self.session.log_event("canvas.delete", {"index": del_idx}, log_widget=self.log_text)
                self._selected_shape_id = None
                self._render_canvas()

        def _snap_grid(self):
            if not self.session:
                return
            # 좌표를 10단위 스냅
            for shp in self.session.canvas_data.get("shapes", []):
                if shp["type"] == "rect":
                    shp["x1"] = round(shp["x1"]/10)*10
                    shp["y1"] = round(shp["y1"]/10)*10
                    shp["x2"] = round(shp["x2"]/10)*10
                    shp["y2"] = round(shp["y2"]/10)*10
                elif shp["type"] == "text":
                    shp["x"] = round(shp["x"]/10)*10
                    shp["y"] = round(shp["y"]/10)*10
            self.session.log_event("canvas.snap_grid", {}, log_widget=self.log_text)
            self._render_canvas()

        def _on_canvas_down(self, ev):
            if not self.session:
                return
            # 클릭한 도형 선택/토글
            self._selected_shape_id = None
            hit_idx = None
            for i, shp in enumerate(reversed(self.session.canvas_data.get("shapes", []))):
                idx = len(self.session.canvas_data["shapes"]) - 1 - i
                if shp["type"] == "rect":
                    if shp["x1"] <= ev.x <= shp["x2"] and shp["y1"] <= ev.y <= shp["y2"]:
                        hit_idx = idx; break
                elif shp["type"] == "text":
                    # 텍스트는 근사 박스
                    bx,by = shp["x"], shp["y"]
                    if (bx <= ev.x <= bx+140) and (by <= ev.y <= by+24):
                        hit_idx = idx; break
            # 선택 표시
            for shp in self.session.canvas_data.get("shapes", []):
                shp["_selected"] = False
            if hit_idx is not None:
                self.session.canvas_data["shapes"][hit_idx]["_selected"] = True
                self._selected_shape_id = hit_idx
            self._render_canvas()
            # 선택 윤곽 그리기
            if self._selected_shape_id is not None:
                shp = self.session.canvas_data["shapes"][self._selected_shape_id]
                if shp["type"] == "rect":
                    self.canvas.create_rectangle(shp["x1"], shp["y1"], shp["x2"], shp["y2"], outline="#ff6", width=3)
                else:
                    self.canvas.create_rectangle(shp["x"], shp["y"], shp["x"]+140, shp["y"]+24, outline="#ff6", width=3)

            # 드래그 시작점
            self._drag_off = (ev.x, ev.y)

        def _on_canvas_drag(self, ev):
            if not self.session or self._selected_shape_id is None:
                return
            shp = self.session.canvas_data["shapes"][self._selected_shape_id]
            dx = ev.x - getattr(self, "_drag_off", (ev.x, ev.y))[0]
            dy = ev.y - getattr(self, "_drag_off", (ev.x, ev.y))[1]
            self._drag_off = (ev.x, ev.y)
            if shp["type"] == "rect":
                shp["x1"] += dx; shp["y1"] += dy
                shp["x2"] += dx; shp["y2"] += dy
            else:
                shp["x"] += dx; shp["y"] += dy
            self.session.log_event("canvas.drag", {"idx": self._selected_shape_id, "dx": dx, "dy": dy}, log_widget=self.log_text)
            self._render_canvas()

        # ---------- 파일 입출력 ----------
        def _export_json(self):
            if not self.session:
                return
            path = filedialog.asksaveasfilename(defaultextension=".json", initialfile=f"{self.session.estimate_id}_final.json")
            if not path:
                return
            data = {
                "estimate_id": self.session.estimate_id,
                "current_json": self.session.current_json,
                "notes": self.session.notes,
                "canvas_data": self.session.canvas_data,
                "source_path": self.session.source_path,
            }
            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            self._log(f"내보내기 완료: {os.path.basename(path)}")

        def _import_json(self):
            path = filedialog.askopenfilename(filetypes=[("JSON","*.json")])
            if not path:
                return
            try:
                with open(path, "r", encoding="utf-8") as f:
                    obj = json.load(f)
                if not self.session:
                    # 새 세션 구성
                    estimate_id = obj.get("estimate_id") or str(uuid.uuid4())
                    bundle_dir = _ensure_dir(os.path.join("data", "training_edits", estimate_id))
                    session = ReviewSession(
                        estimate_id=estimate_id,
                        bundle_dir=bundle_dir,
                        source_path=obj.get("source_path"),
                        initial_json=obj.get("current_json", {}),
                        current_json=obj.get("current_json", {}),
                    )
                    session.notes = obj.get("notes", [])
                    session.canvas_data = obj.get("canvas_data", {"shapes":[]})
                    self.attach_session(session)
                else:
                    # 기존 세션에 로드
                    self.session.current_json = obj.get("current_json", {})
                    self.session.notes = obj.get("notes", [])
                    self.session.canvas_data = obj.get("canvas_data", {"shapes":[]})
                    self._render_json(self.session.current_json)
                    self._render_canvas()
                self._log(f"불러오기 완료: {os.path.basename(path)}")
            except Exception as e:
                messagebox.showerror("오류", f"JSON 불러오기 실패: {e}")

        # ---------- 유틸 ----------
        def _log(self, msg: str):
            self.log_text.insert(tk.END, msg + "\n")
            self.log_text.see(tk.END)

        def _load_fake(self):
            # 데모용 가짜 견적을 로딩
            fake = {
                "estimate_id": f"DEMO-{_now_ts()}",
                "client": {"name": "데모고객", "site": "서울", "contact": "010-0000-0000"},
                "enclosure": {"type":"옥내", "material":"SS41", "W":800, "H":1200, "D":300},
                "main": {"type":"MCCB", "poles":"4P", "A":250, "qty":1, "brand":"LS"},
                "branches": [
                    {"type":"ELB", "poles":"2P", "A":20, "qty":4, "brand":"LS"},
                    {"type":"MCCB", "poles":"3P", "A":50, "qty":2, "brand":"ABB"}
                ],
                "accessories":[{"name":"SPD", "grade":"Class II", "qty":1}],
                "labor":{"install":"기본", "days":1},
                "totals":{"sum_items": 1234560, "vat":123456, "grand":1358016}
            }
            self.on_estimate_ready_for_review(fake, source_path=None)

    # ------------------------------------------------------------
    # App 통합 훅: Notebook에 탭 추가 + 외부에서 호출 가능한 브릿지
    def _inject_review_tab_into_app(AppClass):
        """
        기존 App 클래스에 review_tab을 주입하고 탭을 추가합니다.
        App.__init__ 내에서 self._build_tab_review(notebook) 을 호출하거나,
        여기서 바로 Notebook을 찾아 추가합니다.
        """
        orig_init = AppClass.__init__

        def _new_init(self, *args, **kwargs):
            orig_init(self, *args, **kwargs)
            try:
                # NOTE: 기존 코드에서 self.notebook 또는 self.tabs 같은 Notebook 참조명을 사용한다고 가정
                nb = getattr(self, "notebook", None) or getattr(self, "tabs", None)
                if nb is None:
                    # 최상위 위젯에서 Notebook 탐색(비상시)
                    for child in self.winfo_children():
                        if isinstance(child, ttk.Notebook):
                            nb = child; break
                if nb is None:
                    return  # Notebook을 못 찾으면 무시

                # 탭 생성
                self.review_tab = ReviewTrainTab(nb, app_ref=self)
                nb.add(self.review_tab, text="교육용탭")

                # 외부에서 JSON을 넘겨 탭에 로딩하는 브릿지 메서드 바인딩
                def _bridge_on_estimate_ready_for_review(est_json: Dict[str,Any], source_path: Optional[str]=None):
                    self.review_tab.on_estimate_ready_for_review(est_json, source_path=source_path)
                setattr(self, "on_estimate_ready_for_review", _bridge_on_estimate_ready_for_review)

            except Exception as e:
                print("[경고] 교육용탭 초기화 실패:", e)

        AppClass.__init__ = _new_init
        return AppClass

    # ------------------------------------------------------------
    # App 심볼을 찾아 주입 (기존 파일 최하단의 App 클래스 선언 이후에 이 블록이 위치해야 합니다)
    try:
        # 이미 정의돼 있을 App 클래스를 가져와 주입한다.
        # 만약 App 클래스명이 다르면 여기를 수정: 예) class MainApp, KisanApp 등
        _target_app_cls = None
        # 전역 심볼 탐색
        for _name, _obj in list(globals().items()):
            try:
                if isinstance(_obj, type) and _obj.__name__ in ("App","MainApp","KisanApp"):
                    _target_app_cls = _obj
                    break
            except:
                pass
        if _target_app_cls:
            _inject_review_tab_into_app(_target_app_cls)
        else:
            print("[알림] App 클래스를 찾지 못했습니다. 파일에서 App 클래스명을 확인하세요.")
    except Exception as e:
        print("[경고] 교육용탭 패치 주입 실패:", e)

    # === [PATCH END] system_estimate_ui.py — 교육용탭 추가 ===
# ------------------------------ RAG Reindex (append) ------------------------------
import os, io, json, uuid, datetime as dt, glob

def _do_rag_reindex(verbose: bool = True) -> dict:
    """
    data/estimates/*.json 과 RAG 원문(estimate_rag_bundle_v1.0.0.json, accessories_rag_doc.jsonl,
    breaker_rag_doc.jsonl, enclosure_rag_doc.jsonl, costing_rag_doc.jsonl)을 찾아
    data/rag/index.jsonl 로 통합 색인 파일을 만든다. (단순 JSONL 머지; 임베딩/FAISS는 별도 단계)
    """
    # BASE_DIR, DATA_DIR 가 위에서 이미 정의되어 있다는 가정 (경로 수정 완료 상태)
    base_dir = BASE_DIR
    data_dir = os.path.join(base_dir, "data")
    est_dir  = os.path.join(data_dir, "estimates")
    rag_dir  = os.path.join(data_dir, "rag")
    os.makedirs(rag_dir, exist_ok=True)

    # 1) estimates/*.json 수집
    estimate_files = sorted(glob.glob(os.path.join(est_dir, "*.json")))

    # 2) RAG 원문 파일들 탐색(프로젝트 전체를 한 번 훑어서 이름 매칭)
    wanted = {
        "estimate_rag_bundle_v1.0.0.json": None,
        "accessories_rag_doc.jsonl": None,
        "breaker_rag_doc.jsonl": None,
        "enclosure_rag_doc.jsonl": None,
        "costing_rag_doc.jsonl": None,
    }
    # 탐색 루트 후보: 프로젝트 루트, data/, rag/
    search_roots = [
        base_dir,
        data_dir,
        os.path.join(base_dir, "rag"),
        os.path.join(base_dir, "docs"),
    ]
    for root in search_roots:
        for root_dir, _, files in os.walk(root):
            for fn in files:
                if fn in wanted and wanted[fn] is None:
                    wanted[fn] = os.path.join(root_dir, fn)

    # 출력 대상
    out_jsonl = os.path.join(rag_dir, "index.jsonl")
    manifest  = os.path.join(rag_dir, "index.manifest.json")

    written = 0
    doc_list = []

    def write_jsonl(obj: dict, fh):
        nonlocal written
        fh.write(json.dumps(obj, ensure_ascii=False) + "\n")
        written += 1

    # helper: estimate JSON → text 추출(간단 요약)
    def summarize_estimate_json(path: str) -> dict:
        try:
            with io.open(path, "r", encoding="utf-8") as f:
                j = json.load(f)
        except Exception as e:
            return {"text": "", "meta": {"error": str(e)}}

        lines = j.get("lines") or []
        names = []
        for ln in lines:
            nm = (ln.get("raw") or {}).get("품목") or ln.get("item_type") or ln.get("model") or ""
            if not nm:
                # raw dict의 첫 텍스트
                raw = ln.get("raw") or {}
                for v in raw.values():
                    if isinstance(v, str) and v.strip():
                        nm = v.strip(); break
            if nm: names.append(str(nm))
        text = " | ".join(names) if names else json.dumps(j, ensure_ascii=False)
        meta = {
            "file": os.path.basename(path),
            "sheet": j.get("sheet_name"),
            "panel": j.get("panel_name"),
            "block": j.get("block_index"),
            "generated_at": j.get("meta", {}).get("generated_at"),
        }
        return {"text": text, "meta": meta}

    # 작성 시작
    now = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with io.open(out_jsonl, "w", encoding="utf-8") as fh:
        # a) estimates
        for p in estimate_files:
            sid = f"est-{uuid.uuid4().hex[:12]}"
            summ = summarize_estimate_json(p)
            write_jsonl({
                "doc_id": sid,
                "type": "estimate",
                "path": p,
                "text": summ["text"],
                "meta": summ["meta"],
                "created_at": now,
            }, fh)
            doc_list.append({"doc_id": sid, "type": "estimate", "path": p})

        # b) RAG 문서들 (jsonl/ json 구분)
        for fn, fp in wanted.items():
            if not fp or not os.path.exists(fp): 
                continue
            if fp.lower().endswith(".jsonl"):
                with io.open(fp, "r", encoding="utf-8") as rf:
                    for line in rf:
                        line = line.strip()
                        if not line:
                            continue
                        try:
                            obj = json.loads(line)
                        except:
                            obj = {"text": line}
                        sid = f"rag-{uuid.uuid4().hex[:12]}"
                        write_jsonl({
                            "doc_id": sid,
                            "type": "rag",
                            "source": fn,
                            "text": obj.get("text") or json.dumps(obj, ensure_ascii=False),
                            "meta": {k:v for k,v in obj.items() if k!="text"},
                            "created_at": now,
                        }, fh)
                        doc_list.append({"doc_id": sid, "type": "rag", "path": fp})
            elif fp.lower().endswith(".json"):
                try:
                    with io.open(fp, "r", encoding="utf-8") as rf:
                        obj = json.load(rf)
                except Exception as e:
                    obj = {"_error": str(e)}
                sid = f"rag-{uuid.uuid4().hex[:12]}"
                write_jsonl({
                    "doc_id": sid,
                    "type": "rag",
                    "source": fn,
                    "text": json.dumps(obj, ensure_ascii=False),
                    "meta": {"format": "json"},
                    "created_at": now,
                }, fh)
                doc_list.append({"doc_id": sid, "type": "rag", "path": fp})

    # 매니페스트 기록
    out = {
        "created_at": now,
        "project_root": base_dir,
        "index_file": out_jsonl,
        "doc_count": written,
        "docs": doc_list,
        "inputs": {
            "estimates_dir": est_dir,
            "wanted_files": wanted,
        }
    }
    with io.open(manifest, "w", encoding="utf-8") as mf:
        json.dump(out, mf, ensure_ascii=False, indent=2)

    if verbose:
        print(f"[RAG] index written: {out_jsonl}")
        print(f"[RAG] docs: {written}, manifest: {manifest}")
    return out

# === [PATCH START] system_estimate_ui.py — 교육용탭(Review & Train) V2.2 ===
# 추가:
#  - 커리큘럼 생성(10건): 난이도(Easy/Med/Hard) + 학습목표 메타데이터 포함
#  - 리그레션 스위트: 폴더 선택 → JSON 일괄 점검/지표 산출 → CSV 리포트 저장
#  - DXF 내보내기: 캔버스 사각형/텍스트 → DXF(OUTLINE/TEXT 레이어)로 내보냄
# 기존:
#  - 미니 캔버스, JSON 편집 양방향, 오토세이브/스냅샷, Diff/HUD/규칙제안, 무작위 10개 생성

import os, json, time, uuid, threading, copy, random, csv
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple

try:
    import tkinter as tk
    from tkinter import ttk, messagebox, filedialog
except:
    tk = None
    ttk = None

# ---------- 공용 유틸 ----------
def _ensure_dir(path: str):
    os.makedirs(path, exist_ok=True); return path

def _now_ts():
    return time.strftime("%Y%m%d_%H%M%S")

def _json_deepcopy(obj):
    return json.loads(json.dumps(obj, ensure_ascii=False))

# ---------- 이벤트/세션 ----------
@dataclass
class EditEvent:
    t: float
    who: str
    action: str
    detail: Dict[str, Any] = field(default_factory=dict)

@dataclass
class ReviewSession:
    estimate_id: str
    bundle_dir: str
    source_path: Optional[str] = None
    initial_json: Dict[str, Any] = field(default_factory=dict)
    current_json: Dict[str, Any] = field(default_factory=dict)
    notes: List[str] = field(default_factory=list)
    canvas_data: Dict[str, Any] = field(default_factory=lambda: {"shapes": []})
    events: List[EditEvent] = field(default_factory=list)
    autosave_interval_sec: int = 8
    _stop_flag: bool = False
    _autosave_thread: Optional[threading.Thread] = None
    _edit_counters: Dict[str, int] = field(default_factory=dict)

    def start_autosave(self, log_widget=None):
        if self._autosave_thread and self._autosave_thread.is_alive(): return
        def _loop():
            while not self._stop_flag:
                try: self.save_snapshot(kind="autosave", log_widget=log_widget)
                except Exception as e:
                    if log_widget is not None:
                        log_widget.insert(tk.END, f"[오토세이브 오류] {e}\n")
                time.sleep(self.autosave_interval_sec)
        self._autosave_thread = threading.Thread(target=_loop, daemon=True); self._autosave_thread.start()
    def stop(self): self._stop_flag = True

    def _paths(self) -> Tuple[str, str]:
        snaps = _ensure_dir(os.path.join(self.bundle_dir, "snapshots"))
        logs = _ensure_dir(os.path.join(self.bundle_dir, "logs"))
        return snaps, logs

    def log_event(self, action: str, detail: Optional[Dict[str, Any]]=None, log_widget=None):
        _, logs = self._paths()
        ev = {"t": time.time(), "ts": _now_ts(), "who": "user", "action": action, "detail": detail or {}}
        with open(os.path.join(logs, f"{self.estimate_id}.jsonl"), "a", encoding="utf-8") as f:
            f.write(json.dumps(ev, ensure_ascii=False) + "\n")
        key = action.split(":")[0]
        self._edit_counters[key] = self._edit_counters.get(key, 0) + 1
        if log_widget is not None:
            log_widget.insert(tk.END, f"[이벤트] {action}\n"); log_widget.see(tk.END)

    def save_snapshot(self, kind="manual", log_widget=None):
        snaps, logs = self._paths()
        ts = _now_ts()
        snap_path = os.path.join(snaps, f"{self.estimate_id}_{kind}_{ts}.json")
        data = {
            "estimate_id": self.estimate_id, "kind": kind, "ts": ts,
            "initial_json": self.initial_json, "current_json": self.current_json,
            "notes": self.notes, "canvas_data": self.canvas_data, "source_path": self.source_path
        }
        with open(snap_path, "w", encoding="utf-8") as f: json.dump(data, f, ensure_ascii=False, indent=2)
        ev = {"t": time.time(), "ts": ts, "who": "user", "action": f"snapshot:{kind}",
              "detail": {"snap_path": snap_path, "counts": {"notes": len(self.notes),
                       "shapes": len(self.canvas_data.get("shapes", []))}}}
        with open(os.path.join(logs, f"{self.estimate_id}.jsonl"), "a", encoding="utf-8") as f:
            f.write(json.dumps(ev, ensure_ascii=False) + "\n")
        if log_widget is not None:
            log_widget.insert(tk.END, f"[{kind}] 스냅샷 저장: {os.path.basename(snap_path)}\n"); log_widget.see(tk.END)

# ---------- Diff / 품질 ----------
def _dict_diff(a: Any, b: Any, prefix="") -> List[str]:
    diffs = []
    if isinstance(a, dict) and isinstance(b, dict):
        keys = set(a.keys()) | set(b.keys())
        for k in sorted(keys):
            ak, bk = a.get(k, None), b.get(k, None)
            path = f"{prefix}.{k}" if prefix else k
            if type(ak) != type(bk):
                diffs.append(f"{path}: type {type(ak).__name__}->{type(bk).__name__}")
            elif isinstance(ak, (dict, list)):
                diffs.extend(_dict_diff(ak, bk, path))
            else:
                if ak != bk: diffs.append(f"{path}: {ak} -> {bk}")
    elif isinstance(a, list) and isinstance(b, list):
        ln = max(len(a), len(b))
        for i in range(ln):
            ai = a[i] if i < len(a) else None
            bi = b[i] if i < len(b) else None
            path = f"{prefix}[{i}]"
            if type(ai) != type(bi):
                diffs.append(f"{path}: type {type(ai).__name__}->{type(bi).__name__}")
            elif isinstance(ai, (dict, list)):
                diffs.extend(_dict_diff(ai, bi, path))
            else:
                if ai != bi: diffs.append(f"{path}: {ai} -> {bi}")
    else:
        if a != b: diffs.append(f"{prefix}: {a} -> {b}")
    return diffs

def _quality_checks(data: Dict[str, Any]) -> List[str]:
    warn = []
    req = ["enclosure", "main", "branches", "totals"]
    for k in req:
        if k not in data: warn.append(f"[구조] '{k}' 누락")

    enc = data.get("enclosure", {})
    for f in ["W","H","D"]:
        v = enc.get(f)
        if v is None or (isinstance(v, (int, float)) and v <= 0):
            warn.append(f"[외함] {f} 값 비정상: {v}")

    main = data.get("main", {})
    if not main.get("type"): warn.append("[메인] 종류 누락")
    if not main.get("poles"): warn.append("[메인] 극수 누락")
    if main.get("A") in (None, 0): warn.append("[메인] 용량(A) 누락/0")

    branches = data.get("branches", [])
    if not isinstance(branches, list): warn.append("[분기] 리스트 아님")
    else:
        if len(branches) == 0: warn.append("[분기] 비어 있음")
        total_qty = 0
        for i, br in enumerate(branches):
            q = br.get("qty", 0); total_qty += (q if isinstance(q, int) else 0)
            if br.get("type") is None: warn.append(f"[분기#{i}] type 누락")
            if br.get("poles") is None: warn.append(f"[분기#{i}] poles 누락")
            if br.get("A") in (None, 0): warn.append(f"[분기#{i}] A 누락/0")
        if total_qty > 40: warn.append(f"[분기] 수량 과다(={total_qty}). 과밀 가능성 확인 필요")

    totals = data.get("totals", {})
    if not totals: warn.append("[합계] totals 누락")
    else:
        g = totals.get("grand", 0)
        if isinstance(g, (int, float)) and g <= 0:
            warn.append(f"[합계] grand 비정상: {g}")
    return warn

# ---------- UI ----------
class ReviewTrainTab(ttk.Frame):
    def __init__(self, master, app_ref, **kwargs):
        super().__init__(master, **kwargs)
        self.app = app_ref
        self.session: Optional[ReviewSession] = None
        self._json_debounce_ms = 500
        self._json_debounce_after = None
        self._selected_shape_idx: Optional[int] = None
        self._drag_off = (0,0)

        # 레이아웃
        self.columnconfigure(0, weight=1); self.columnconfigure(1, weight=1)
        self.rowconfigure(0, weight=1); self.rowconfigure(1, weight=0); self.rowconfigure(2, weight=0)

        # 좌: 캔버스
        left = ttk.Frame(self); left.grid(row=0, column=0, sticky="nsew", padx=6, pady=6)
        left.rowconfigure(1, weight=1)
        ttk.Label(left, text="미니 배치 캔버스").grid(row=0, column=0, sticky="w")
        self.canvas = tk.Canvas(left, width=520, height=520, bg="#f7f7f9",
                                highlightthickness=1, highlightbackground="#ccc")
        self.canvas.grid(row=1, column=0, sticky="nsew")
        tbar = ttk.Frame(left); tbar.grid(row=2, column=0, sticky="ew", pady=(6,0))
        ttk.Button(tbar, text="사각형", command=self._add_rect).pack(side="left")
        ttk.Button(tbar, text="텍스트", command=self._add_text).pack(side="left")
        ttk.Button(tbar, text="삭제", command=self._delete_selected).pack(side="left")
        ttk.Button(tbar, text="격자스냅", command=self._snap_grid).pack(side="left")
        ttk.Button(tbar, text="DXF 내보내기", command=self._export_dxf).pack(side="left")
        self.canvas.bind("<Button-1>", self._on_canvas_down)
        self.canvas.bind("<B1-Motion>", self._on_canvas_drag)

        # 우: JSON/노트/버튼
        right = ttk.Frame(self); right.grid(row=0, column=1, sticky="nsew", padx=6, pady=6)
        right.rowconfigure(1, weight=1); right.rowconfigure(3, weight=1)
        ttk.Label(right, text="견적 JSON").grid(row=0, column=0, sticky="w")
        self.json_text = tk.Text(right, height=16, wrap="none"); self.json_text.grid(row=1, column=0, sticky="nsew")
        self.json_text.bind("<<Modified>>", self._on_json_modified)

        ttk.Label(right, text="주석/메모").grid(row=2, column=0, sticky="w", pady=(8,0))
        self.notes_text = tk.Text(right, height=8, wrap="word"); self.notes_text.grid(row=3, column=0, sticky="nsew")

        btns = ttk.Frame(right); btns.grid(row=4, column=0, sticky="ew", pady=(8,0))
        ttk.Button(btns, text="가짜 견적", command=self._load_fake).pack(side="left")
        ttk.Button(btns, text="스냅샷", command=lambda: self._snapshot("manual")).pack(side="left")
        ttk.Button(btns, text="JSON 내보내기", command=self._export_json).pack(side="left")
        ttk.Button(btns, text="JSON 불러오기", command=self._import_json).pack(side="left")
        ttk.Button(btns, text="이벤트 내보내기", command=self._export_events).pack(side="left")

        # 中: HUD/규칙/프리셋/생성기/커리큘럼/리그레션
        mid = ttk.Frame(self); mid.grid(row=1, column=0, columnspan=2, sticky="ew", padx=6)
        mid.columnconfigure(0, weight=1); mid.columnconfigure(1, weight=1); mid.columnconfigure(2, weight=1); mid.columnconfigure(3, weight=1)

        self.hud_lbl = ttk.Label(mid, text="품질 HUD: -"); self.hud_lbl.grid(row=0, column=0, sticky="w")
        self.rule_lbl = ttk.Label(mid, text="규칙 제안: -"); self.rule_lbl.grid(row=0, column=1, sticky="w")

        # 프리셋 + 무작위 10개(요청/견적) + 커리큘럼 + 리그레션
        bar = ttk.Frame(mid); bar.grid(row=0, column=2, columnspan=2, sticky="e")
        ttk.Label(bar, text="프리셋:").pack(side="left")
        ttk.Button(bar, text="ELB 2P 20A×4", command=lambda: self._preset_branch("ELB","2P",20,4,"LS")).pack(side="left")
        ttk.Button(bar, text="MCCB 3P 50A×2", command=lambda: self._preset_branch("MCCB","3P",50,2,"ABB")).pack(side="left")
        ttk.Button(bar, text="요청 10개 생성", command=self._generate_synthetic_requests).pack(side="left", padx=(8,0))
        ttk.Button(bar, text="견적 10개 생성", command=self._generate_synthetic_quotes).pack(side="left")
        ttk.Button(bar, text="커리큘럼 10개", command=self._generate_curriculum).pack(side="left", padx=(8,0))
        ttk.Button(bar, text="리그레션 실행", command=self._run_regression_suite).pack(side="left")

        # 下: 로그 + 스냅샷 히스토리 + Diff
        bottom = ttk.Frame(self); bottom.grid(row=2, column=0, columnspan=2, sticky="nsew", padx=6, pady=(0,6))
        bottom.columnconfigure(0, weight=1); bottom.columnconfigure(1, weight=0); bottom.columnconfigure(2, weight=1)
        ttk.Label(bottom, text="이벤트 로그").grid(row=0, column=0, sticky="w")
        ttk.Label(bottom, text="스냅샷").grid(row=0, column=1, sticky="w")
        ttk.Label(bottom, text="Diff(초기↔현재)").grid(row=0, column=2, sticky="w")

        self.log_text = tk.Text(bottom, height=8, bg="#111", fg="#ddd"); self.log_text.grid(row=1, column=0, sticky="nsew")
        hist_frame = ttk.Frame(bottom); hist_frame.grid(row=1, column=1, sticky="ns", padx=6)
        self.hist_list = tk.Listbox(hist_frame, height=8, width=28); self.hist_list.pack(side="left", fill="y")
        ttk.Button(hist_frame, text="불러오기", command=self._load_snapshot_from_list).pack(side="left", padx=(4,0))
        self.diff_text = tk.Text(bottom, height=8, bg="#0b1", fg="#021", wrap="word"); self.diff_text.grid(row=1, column=2, sticky="nsew")

        self._log("교육용탭 V2.2 로딩 완료. 커리큘럼/리그레션/DXF 버튼을 사용할 수 있습니다.")

    # ---------- 세션/로드 ----------
    def attach_session(self, session: ReviewSession):
        if self.session: self.session.stop()
        self.session = session
        self._log(f"세션 시작: {session.estimate_id}")
        self._render_json(session.current_json)
        self._render_canvas()
        self._render_diff()
        self._refresh_hud_and_rules()
        self._refresh_history()
        self.notes_text.delete("1.0", tk.END)
        self.notes_text.insert("1.0", "\n".join(session.notes))
        session.start_autosave(log_widget=self.log_text)

    def on_estimate_ready_for_review(self, est_json: Dict[str, Any], source_path: Optional[str]=None):
        est_id = est_json.get("estimate_id") or f"EST-{_now_ts()}"
        bundle_dir = _ensure_dir(os.path.join("data", "training_edits", est_id))
        session = ReviewSession(
            estimate_id=est_id, bundle_dir=bundle_dir, source_path=source_path,
            initial_json=_json_deepcopy(est_json), current_json=_json_deepcopy(est_json)
        )
        self.attach_session(session)
        session.save_snapshot(kind="initial", log_widget=self.log_text)

    # ---------- JSON 편집 ----------
    def _render_json(self, data: Dict[str, Any]):
        self.json_text.delete("1.0", tk.END)
        self.json_text.insert("1.0", json.dumps(data, ensure_ascii=False, indent=2))
        self.json_text.edit_modified(False)

    def _on_json_modified(self, event=None):
        if self.json_text.edit_modified():
            if self._json_debounce_after: self.after_cancel(self._json_debounce_after)
            self._json_debounce_after = self.after(self._json_debounce_ms, self._apply_json_from_editor)
            self.json_text.edit_modified(False)

    def _apply_json_from_editor(self):
        if not self.session: return
        txt = self.json_text.get("1.0", tk.END).strip()
        try:
            obj = json.loads(txt)
            self.session.current_json = obj
            self.session.log_event("json.update", {"keys": list(obj.keys())}, self.log_text)
            self._render_diff(); self._refresh_hud_and_rules()
        except Exception as e:
            self._log(f"[경고] JSON 파싱 실패: {e}")

    # ---------- 노트/스냅샷 ----------
    def _snapshot(self, kind="manual"):
        if not self.session:
            messagebox.showwarning("알림","세션 없음"); return
        self.session.notes = [n for n in self.notes_text.get("1.0", tk.END).splitlines() if n.strip()]
        self.session.save_snapshot(kind=kind, log_widget=self.log_text)
        self._refresh_history(); self._render_diff(); self._refresh_hud_and_rules()

    def _refresh_history(self):
        self.hist_list.delete(0, tk.END)
        if not self.session: return
        snaps_dir = os.path.join(self.session.bundle_dir, "snapshots")
        if not os.path.isdir(snaps_dir): return
        files = sorted(os.listdir(snaps_dir))
        for fn in files:
            if fn.endswith(".json"): self.hist_list.insert(tk.END, fn)

    def _load_snapshot_from_list(self):
        if not self.session: return
        sel = self.hist_list.curselection()
        if not sel: return
        fn = self.hist_list.get(sel[0])
        path = os.path.join(self.session.bundle_dir, "snapshots", fn)
        try:
            with open(path, "r", encoding="utf-8") as f: obj = json.load(f)
            self.session.current_json = obj.get("current_json", {})
            self.session.notes = obj.get("notes", [])
            self.session.canvas_data = obj.get("canvas_data", {"shapes":[]})
            self._render_json(self.session.current_json); self._render_canvas()
            self._render_diff(); self._refresh_hud_and_rules()
            self._log(f"[히스토리] 로드: {fn}")
        except Exception as e:
            messagebox.showerror("오류", f"스냅샷 로드 실패: {e}")

    # ---------- Diff/HUD/규칙 ----------
    def _render_diff(self):
        self.diff_text.delete("1.0", tk.END)
        if not self.session: return
        diffs = _dict_diff(self.session.initial_json, self.session.current_json)
        if not diffs: self.diff_text.insert(tk.END, "변경 없음")
        else:
            for d in diffs: self.diff_text.insert(tk.END, "• " + d + "\n")

    def _refresh_hud_and_rules(self):
        if not self.session: return
        warns = _quality_checks(self.session.current_json)
        hud = "품질 OK" if not warns else " | ".join(warns[:6]) + (" ..." if len(warns)>6 else "")
        self.hud_lbl.config(text=f"품질 HUD: {hud}")

        suggestions = []
        cnt = self.session._edit_counters
        if cnt.get("json",0) + cnt.get("canvas",0) > 10:
            suggestions.append("반복 편집 多 → 프리셋/룰로 승격 제안")
        if cnt.get("canvas.add_rect",0)>=3 and cnt.get("canvas.add_text",0)>=3:
            suggestions.append("배치 스케치 패턴 감지 → CAD 자동 배치 룰 후보")
        if cnt.get("json.update",0)>=5:
            suggestions.append("브랜드/극수/용량 정규화 룰 보강")
        self.rule_lbl.config(text="규칙 제안: " + (" , ".join(suggestions) if suggestions else "-"))

    # ---------- 캔버스 ----------
    def _render_canvas(self):
        self.canvas.delete("all")
        for i in range(0, 520, 20):
            self.canvas.create_line(i, 0, i, 520, fill="#eee")
            self.canvas.create_line(0, i, 520, i, fill="#eee")
        if not self.session: return
        for shp in self.session.canvas_data.get("shapes", []):
            if shp.get("type")=="rect":
                rid = self.canvas.create_rectangle(shp["x1"],shp["y1"],shp["x2"],shp["y2"],
                                                   outline="#333",width=2,fill="#dfe9ff"); shp["rid"]=rid
            elif shp.get("type")=="text":
                tid = self.canvas.create_text(shp["x"],shp["y"], text=shp.get("text",""),
                                              anchor="nw", font=("Arial",11)); shp["tid"]=tid

    def _hit_test(self, x,y)->Optional[int]:
        if not self.session: return None
        for i in range(len(self.session.canvas_data["shapes"])-1, -1, -1):
            shp = self.session.canvas_data["shapes"][i]
            if shp["type"]=="rect":
                if shp["x1"]<=x<=shp["x2"] and shp["y1"]<=y<=shp["y2"]: return i
            else:
                bx,by = shp["x"], shp["y"]
                if bx<=x<=bx+140 and by<=y<=by+24: return i
        return None

    def _on_canvas_down(self, ev):
        if not self.session: return
        idx = self._hit_test(ev.x, ev.y)
        self._selected_shape_idx = idx
        self._drag_off = (ev.x, ev.y)
        self._render_canvas()
        if idx is not None:
            shp = self.session.canvas_data["shapes"][idx]
            if shp["type"]=="rect":
                self.canvas.create_rectangle(shp["x1"],shp["y1"],shp["x2"],shp["y2"], outline="#ff6", width=3)
            else:
                self.canvas.create_rectangle(shp["x"],shp["y"],shp["x"]+140,shp["y"]+24, outline="#ff6", width=3)

    def _on_canvas_drag(self, ev):
        if not self.session or self._selected_shape_idx is None: return
        shp = self.session.canvas_data["shapes"][self._selected_shape_idx]
        dx, dy = ev.x - self._drag_off[0], ev.y - self._drag_off[1]
        self._drag_off = (ev.x, ev.y)
        if shp["type"]=="rect":
            shp["x1"]+=dx; shp["y1"]+=dy; shp["x2"]+=dx; shp["y2"]+=dy
        else:
            shp["x"]+=dx; shp["y"]+=dy
        self.session.log_event("canvas.drag", {"idx": self._selected_shape_idx, "dx":dx,"dy":dy}, self.log_text)
        self._render_canvas()

    def _add_rect(self):
        if not self.session: return
        shp = {"type":"rect","x1":40,"y1":40,"x2":140,"y2":100}
        self.session.canvas_data["shapes"].append(shp)
        self.session.log_event("canvas.add_rect", {"pos":[40,40,140,100]}, self.log_text)
        self._render_canvas()

    def _add_text(self):
        if not self.session: return
        shp = {"type":"text","x":50,"y":120,"text":"분기 ELB 2P 20A x4"}
        self.session.canvas_data["shapes"].append(shp)
        self.session.log_event("canvas.add_text", {"pos":[50,120],"text":shp["text"]}, self.log_text)
        self._render_canvas()

    def _delete_selected(self):
        if not self.session or self._selected_shape_idx is None: return
        self.session.canvas_data["shapes"].pop(self._selected_shape_idx)
        self.session.log_event("canvas.delete", {"idx": self._selected_shape_idx}, self.log_text)
        self._selected_shape_idx = None; self._render_canvas()

    def _snap_grid(self):
        if not self.session: return
        for shp in self.session.canvas_data.get("shapes", []):
            if shp["type"]=="rect":
                for k in ("x1","y1","x2","y2"): shp[k] = round(shp[k]/10)*10
            else:
                shp["x"] = round(shp["x"]/10)*10; shp["y"] = round(shp["y"]/10)*10
        self.session.log_event("canvas.snap_grid", {}, self.log_text)
        self._render_canvas()

    # ---------- 프리셋/입출력 ----------
    def _preset_branch(self, t, p, a, q, brand):
        if not self.session: return
        data = self.session.current_json
        data.setdefault("branches", []).append({"type":t,"poles":p,"A":a,"qty":q,"brand":brand})
        self._render_json(data); self._apply_json_from_editor()
        self.session.log_event("json.preset_branch", {"type":t,"poles":p,"A":a,"qty":q,"brand":brand}, self.log_text)

    def _export_json(self):
        if not self.session: return
        path = filedialog.asksaveasfilename(defaultextension=".json",
                    initialfile=f"{self.session.estimate_id}_final.json")
        if not path: return
        out = {"estimate_id": self.session.estimate_id,
               "current_json": self.session.current_json,
               "notes": self.session.notes,
               "canvas_data": self.session.canvas_data,
               "source_path": self.session.source_path}
        with open(path,"w",encoding="utf-8") as f: json.dump(out,f,ensure_ascii=False,indent=2)
        self._log(f"JSON 내보내기: {os.path.basename(path)}")

    def _import_json(self):
        path = filedialog.askopenfilename(filetypes=[("JSON","*.json")])
        if not path: return
        try:
            with open(path,"r",encoding="utf-8") as f: obj = json.load(f)
            if not self.session:
                est_id = obj.get("estimate_id") or f"EST-{_now_ts()}"
                bundle_dir = _ensure_dir(os.path.join("data","training_edits", est_id))
                session = ReviewSession(estimate_id=est_id, bundle_dir=bundle_dir,
                                        source_path=obj.get("source_path"),
                                        initial_json=obj.get("current_json",{}),
                                        current_json=obj.get("current_json",{}))
                session.notes = obj.get("notes",[]); session.canvas_data = obj.get("canvas_data",{"shapes":[]})
                self.attach_session(session)
            else:
                self.session.current_json = obj.get("current_json",{})
                self.session.notes = obj.get("notes",[])
                self.session.canvas_data = obj.get("canvas_data",{"shapes":[]})
                self._render_json(self.session.current_json); self._render_canvas()
            self._render_diff(); self._refresh_hud_and_rules()
            self._log(f"JSON 불러오기: {os.path.basename(path)}")
        except Exception as e:
            messagebox.showerror("오류", f"불러오기 실패: {e}")

    def _export_events(self):
        if not self.session: return
        _, logs = self.session._paths()
        src = os.path.join(logs, f"{self.session.estimate_id}.jsonl")
        if not os.path.exists(src):
            messagebox.showinfo("알림","아직 이벤트가 없습니다."); return
        path = filedialog.asksaveasfilename(defaultextension=".jsonl",
                                            initialfile=f"{self.session.estimate_id}_events.jsonl")
        if not path: return
        with open(src,"r",encoding="utf-8") as f: data = f.read()
        with open(path,"w",encoding="utf-8") as f: f.write(data)
        self._log(f"이벤트 내보내기: {os.path.basename(path)}")

    # ---------- 무작위 샘플 생성기(요청/견적) ----------
    def _rand_client(self):
        names = ["코리아ENG","에이플랜","스마트전기","한빛산전","라이트테크"]
        sites = ["서울","인천","수원","대전","부산","광주","울산"]
        return {"name": random.choice(names), "site": random.choice(sites),
                "contact": f"010-{random.randint(1000,9999)}-{random.randint(1000,9999)}"}

    def _rand_enclosure(self):
        etype = random.choice(["옥내","옥외자립"])
        material = random.choice(["SS41","SUS304","AL"])
        W = random.choice([600,700,800,900,1000])
        H = random.choice([1000,1100,1200,1300,1400])
        D = random.choice([250,300,350,400])
        return {"type": etype, "material": material, "W": W, "H": H, "D": D}

    def _rand_main(self):
        t = random.choice(["MCCB","ACB"])
        poles = random.choice(["3P","4P"])
        A = random.choice([125,160,225,250,400,630])
        brand = random.choice(["LS","ABB","Schneider","Hyundai"])
        return {"type": t, "poles": poles, "A": A, "qty": 1, "brand": brand}

    def _rand_branch(self):
        t = random.choice(["ELB","MCCB"])
        poles = random.choice(["2P","3P"])
        A = random.choice([15,20,30,40,50,60,75,100])
        qty = random.randint(1,6)
        brand = random.choice(["LS","ABB","Schneider","Hyundai"])
        return {"type": t, "poles": poles, "A": A, "qty": qty, "brand": brand}

    def _rand_accessories(self):
        acc = []
        if random.random() < 0.5: acc.append({"name":"SPD","grade":random.choice(["Class I","Class II"]),"qty":1})
        if random.random() < 0.3: acc.append({"name":"FAN","size":random.choice([92,120]),"qty":1})
        return acc

    def _calc_totals_mock(self, branches_count:int):
        base = 800000
        per = 120000 * max(1, branches_count)
        grand = base + per
        vat = int(grand * 0.1)
        return {"sum_items": grand, "vat": vat, "grand": grand+vat}

    def _generate_synthetic_requests(self):
        ts = _now_ts()
        out_dir = _ensure_dir(os.path.join("data","synthetic","requests", ts))
        made = []
        for i in range(10):
            req = {
                "request_id": f"REQ-{ts}-{i+1}",
                "client": self._rand_client(),
                "requirements": {
                    "enclosure": self._rand_enclosure(),
                    "main_hint": {"type": random.choice(["MCCB","ACB"]),
                                  "poles": random.choice(["3P","4P"]),
                                  "A": random.choice([160,225,250,400])},
                    "branch_hints": [self._rand_branch() for _ in range(random.randint(2,5))],
                    "options": {"outdoor": random.random()<0.4, "base": random.random()<0.5}
                },
                "notes": ["무작위 생성 요청서", "학습용 샘플"]
            }
            path = os.path.join(out_dir, f"{req['request_id']}.json")
            with open(path,"w",encoding="utf-8") as f: json.dump(req,f,ensure_ascii=False,indent=2)
            made.append(path)
        self._log(f"[생성완료] 견적요청서 10개 → {out_dir}")
        if hasattr(self.app, "rag_reindex_path"):
            try: self.app.rag_reindex_path(out_dir)
            except Exception as e: self._log(f"[참고] 재색인 훅 실패: {e}")
        sample = self._request_to_estimate_seed(made[0])
        self.on_estimate_ready_for_review(sample, source_path=made[0])

    def _request_to_estimate_seed(self, req_path:str)->Dict[str,Any]:
        with open(req_path,"r",encoding="utf-8") as f: req = json.load(f)
        enc = req["requirements"]["enclosure"]; main_hint = req["requirements"]["main_hint"]
        branches = req["requirements"]["branch_hints"]
        seed = {
            "estimate_id": f"EST-SEED-{_now_ts()}",
            "client": req.get("client",{}),
            "enclosure": enc,
            "main": {"type": main_hint.get("type","MCCB"),
                     "poles": main_hint.get("poles","3P"),
                     "A": main_hint.get("A",160),
                     "qty": 1, "brand": random.choice(["LS","ABB","Schneider","Hyundai"])},
            "branches": branches,
            "accessories": self._rand_accessories(),
            "labor": {"install":"기본","days":1},
            "totals": self._calc_totals_mock(sum(b.get("qty",0) for b in branches))
        }
        return seed

    def _generate_synthetic_quotes(self):
        ts = _now_ts()
        out_dir = _ensure_dir(os.path.join("data","synthetic","quotes", ts))
        made = []
        for i in range(10):
            branches = [self._rand_branch() for _ in range(random.randint(3,8))]
            est = {
                "estimate_id": f"EST-{ts}-{i+1}",
                "client": self._rand_client(),
                "enclosure": self._rand_enclosure(),
                "main": self._rand_main(),
                "branches": branches,
                "accessories": self._rand_accessories(),
                "labor": {"install":"기본","days":random.choice([1,2])},
                "totals": self._calc_totals_mock(sum(b.get("qty",0) for b in branches))
            }
            path = os.path.join(out_dir, f"{est['estimate_id']}.json")
            with open(path,"w",encoding="utf-8") as f: json.dump(est,f,ensure_ascii=False,indent=2)
            made.append(path)
        self._log(f"[생성완료] 견적서 10개 → {out_dir}")
        if hasattr(self.app, "rag_reindex_path"):
            try: self.app.rag_reindex_path(out_dir)
            except Exception as e: self._log(f"[참고] 재색인 훅 실패: {e}")
        with open(made[0],"r",encoding="utf-8") as f: sample = json.load(f)
        self.on_estimate_ready_for_review(sample, source_path=made[0])

    # ---------- 커리큘럼 생성(10건) ----------
    def _generate_curriculum(self):
        """
        Easy(3) → Medium(4) → Hard(3) = 총 10건
        각 아이템에 'objectives' 메타데이터 포함: 예) ELB 연쇄, 옥외자립, 과밀 경고, 브랜드 교체, 스냅룰(용량 스냅) 등
        """
        ts = _now_ts()
        out_dir = _ensure_dir(os.path.join("data","curriculum", ts))
        plan = (["Easy"]*3) + (["Medium"]*4) + (["Hard"]*3)
        made = []

        def _mk_item(level:str, idx:int)->Dict[str,Any]:
            obj = []
            enc = self._rand_enclosure()
            main = self._rand_main()
            branches = [self._rand_branch() for _ in range(3 if level=="Easy" else (5 if level=="Medium" else 8))]

            # 레벨별 학습목표/변수 주입
            if level in ("Medium","Hard"):
                # 옥외자립/베이스 옵션 강조
                if random.random() < 0.6:
                    enc["type"] = "옥외자립"; obj.append("옥외자립 옵션/여유공차")
                # 브랜드 혼합 → 정규화 필요
                obj.append("브랜드 혼합 정규화")
            if level == "Hard":
                # 과밀 유도
                for _ in range(2): branches.append(self._rand_branch())
                obj.append("분기 과밀 경고 처리")
                # 용량 스냅 유도(예: 53A 같은 애매값)
                if branches:
                    branches[0]["A"] = random.choice([53,57,109]); obj.append("용량 스냅 규칙")

            est = {
                "estimate_id": f"CUR-{level}-{ts}-{idx}",
                "client": self._rand_client(),
                "enclosure": enc,
                "main": main,
                "branches": branches,
                "accessories": self._rand_accessories(),
                "labor": {"install":"기본","days":random.choice([1,2])},
                "totals": self._calc_totals_mock(sum(b.get("qty",0) for b in branches)),
                "objectives": obj,
                "difficulty": level
            }
            return est

        for i, lv in enumerate(plan, start=1):
            est = _mk_item(lv, i)
            path = os.path.join(out_dir, f"{est['estimate_id']}.json")
            with open(path,"w",encoding="utf-8") as f: json.dump(est,f,ensure_ascii=False,indent=2)
            made.append(path)

        self._log(f"[커리큘럼] {len(made)}건 생성 → {out_dir}")
        # RAG 재색인 훅
        if hasattr(self.app, "rag_reindex_path"):
            try: self.app.rag_reindex_path(out_dir)
            except Exception as e: self._log(f"[참고] 재색인 훅 실패: {e}")
        # 첫 건 로딩
        with open(made[0],"r",encoding="utf-8") as f: sample = json.load(f)
        self.on_estimate_ready_for_review(sample, source_path=made[0])

    # ---------- 리그레션 스위트 ----------
    def _run_regression_suite(self):
        """
        폴더 선택 → 폴더 내 .json 전부 스캔
        - 객체 형태가 두 가지일 수 있음:
          (A) 견적 객체 자체(estimate_id, enclosure, main, branches, totals...)
          (B) 교육용탭 내보내기/스냅샷 객체({current_json, notes, ...})
        - 품질 경고 수, 분기 총량, grand 값, (있다면) diff 개수 산출
        - CSV 리포트 저장: data/reports/regression_<ts>.csv
        """
        root = filedialog.askdirectory(title="리그레션 실행할 폴더 선택")
        if not root: return
        rows = []
        files_scanned = 0

        def _normalize(obj:Dict[str,Any])->Tuple[Dict[str,Any], int]:
            """return (estimate_json, diff_count)"""
            if "current_json" in obj:
                cur = obj.get("current_json",{})
                init = obj.get("initial_json",{})
                diff = len(_dict_diff(init, cur)) if init else -1
                return (cur, diff)
            else:
                return (obj, -1)

        for dirpath, _, filenames in os.walk(root):
            for fn in filenames:
                if not fn.lower().endswith(".json"): continue
                p = os.path.join(dirpath, fn)
                try:
                    with open(p,"r",encoding="utf-8") as f: obj = json.load(f)
                    est, diff_cnt = _normalize(obj)
                    warns = _quality_checks(est)
                    branches = est.get("branches",[])
                    grand = (est.get("totals") or {}).get("grand", None)
                    rows.append({
                        "path": p,
                        "warn_cnt": len(warns),
                        "branches_total_qty": sum(b.get("qty",0) for b in branches if isinstance(b.get("qty",0), int)),
                        "diff_cnt": diff_cnt,
                        "grand": grand
                    })
                    files_scanned += 1
                except Exception as e:
                    rows.append({"path": p, "warn_cnt": -1, "branches_total_qty": -1, "diff_cnt": -1, "grand": None})
                    self._log(f"[경고] 파싱 실패: {p} / {e}")

        rep_dir = _ensure_dir(os.path.join("data","reports"))
        rep_path = os.path.join(rep_dir, f"regression_{_now_ts()}.csv")
        with open(rep_path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=["path","warn_cnt","branches_total_qty","diff_cnt","grand"])
            w.writeheader()
            for r in rows: w.writerow(r)

        # 요약
        ok = sum(1 for r in rows if r["warn_cnt"]==0 and r["warn_cnt"]!=-1)
        bad = sum(1 for r in rows if r["warn_cnt"]>0)
        fail = sum(1 for r in rows if r["warn_cnt"]==-1)
        self._log(f"[리그레션] 스캔:{files_scanned} / OK:{ok} / WARN:{bad} / FAIL:{fail} → {rep_path}")
        messagebox.showinfo("리그레션 완료", f"총 {files_scanned}건\nOK:{ok} WARN:{bad} FAIL:{fail}\n리포트: {rep_path}")

    # ---------- DXF 내보내기 ----------
    def _export_dxf(self):
        """
        캔버스 도형 → DXF (OUTLINE/TEXT 레이어)
        스케일: 1 캔버스 픽셀 = 1 단위(임시). 필요시 스케일 팩터로 조정.
        """
        if not self.session:
            messagebox.showwarning("알림","세션 없음"); return
        path = filedialog.asksaveasfilename(defaultextension=".dxf",
                                            initialfile=f"{self.session.estimate_id}.dxf")
        if not path: return

        def _dxf_header():
            return "0\nSECTION\n2\nHEADER\n0\nENDSEC\n"

        def _dxf_entities(shapes):
            out = ["0","SECTION","2","ENTITIES"]
            # RECT → LWPOLYLINE (닫힘), TEXT → TEXT
            for shp in shapes:
                if shp.get("type")=="rect":
                    x1,y1,x2,y2 = shp["x1"],shp["y1"],shp["x2"],shp["y2"]
                    pts = [(x1,y1),(x2,y1),(x2,y2),(x1,y2)]
                    out += ["0","LWPOLYLINE","8","OUTLINE","90","4","70","1"]
                    for (x,y) in pts:
                        out += ["10",str(x),"20",str(y)]
                elif shp.get("type")=="text":
                    x,y,tx = shp["x"],shp["y"], shp.get("text","")
                    out += ["0","TEXT","8","TEXT","10",str(x),"20",str(y),"40","12","1",tx]
            out += ["0","ENDSEC"]
            return "\n".join(out)+"\n"

        def _dxf_eof(): return "0\nEOF\n"

        shapes = self.session.canvas_data.get("shapes",[])
        dxf = _dxf_header() + _dxf_entities(shapes) + _dxf_eof()
        with open(path,"w",encoding="utf-8") as f: f.write(dxf)
        self._log(f"[DXF] 내보내기 완료: {path}")
        messagebox.showinfo("DXF 내보내기", f"저장 완료:\n{path}")

    # ---------- 가짜 견적 로드 ----------
    def _load_fake(self):
        fake = {
            "estimate_id": f"DEMO-{_now_ts()}",
            "client": {"name":"데모고객","site":"서울","contact":"010-0000-0000"},
            "enclosure": {"type":"옥내","material":"SS41","W":800,"H":1200,"D":300},
            "main": {"type":"MCCB","poles":"4P","A":250,"qty":1,"brand":"LS"},
            "branches":[
                {"type":"ELB","poles":"2P","A":20,"qty":4,"brand":"LS"},
                {"type":"MCCB","poles":"3P","A":50,"qty":2,"brand":"ABB"}
            ],
            "accessories":[{"name":"SPD","grade":"Class II","qty":1}],
            "labor":{"install":"기본","days":1},
            "totals":{"sum_items":1234560,"vat":123456,"grand":1358016}
        }
        self.on_estimate_ready_for_review(fake, source_path=None)

    # ---------- 메시지 ----------
    def _log(self, msg:str):
        self.log_text.insert(tk.END, msg+"\n"); self.log_text.see(tk.END)

# ---------- App에 주입 ----------
def _inject_review_tab_into_app(AppClass):
    orig_init = AppClass.__init__
    def _new_init(self, *args, **kwargs):
        orig_init(self, *args, **kwargs)
        try:
            nb = getattr(self, "notebook", None) or getattr(self, "tabs", None)
            if nb is None:
                for child in self.winfo_children():
                    if isinstance(child, ttk.Notebook): nb = child; break
            if nb is None: return
            self.review_tab = ReviewTrainTab(nb, app_ref=self)
            nb.add(self.review_tab, text="교육용탭")
            def _bridge_on_estimate_ready_for_review(est_json: Dict[str,Any], source_path: Optional[str]=None):
                self.review_tab.on_estimate_ready_for_review(est_json, source_path=source_path)
            setattr(self, "on_estimate_ready_for_review", _bridge_on_estimate_ready_for_review)
        except Exception as e:
            print("[경고] 교육용탭 초기화 실패:", e)
    AppClass.__init__ = _new_init; return AppClass

try:
    _target = None
    for _n,_o in list(globals().items()):
        if isinstance(_o, type) and _o.__name__ in ("App","MainApp","KisanApp"):
            _target = _o; break
    if _target: _inject_review_tab_into_app(_target)
    else: print("[알림] App 클래스를 찾지 못했습니다. 파일에서 App 클래스명을 확인하세요.")
except Exception as e:
    print("[경고] 교육용탭 패치 주입 실패:", e)

# === [PATCH END] system_estimate_ui.py — 교육용탭(Review & Train) V2.2 ===

# === KISAN-AI PATCH: UI Smoke Hooks (append-only) ===
import json as _json_ui
import time as _time_ui
from pathlib import Path as _Path_ui

def _ui_log(event: str, meta: dict | None = None):
    """
    공용 UI 로그 훅. 어느 핸들러든 한 줄만 호출하면 out/ui_smoke.log에 JSONL로 누적 기록됨.
    event: 'E1_TAB_OPEN' | 'E2_BTN_CLICK1' | 'E3_BTN_CLICK2' | 'E4_BTN_CLICK3' | 'E5_TABLE_OK'
    """
    try:
        _out = _Path_ui(__file__).resolve().parent / "out"
        _out.mkdir(parents=True, exist_ok=True)
        _p = _out / "ui_smoke.log"
        rec = {"ts": int(_time_ui.time()), "event": event, "meta": meta or {}}
        with _p.open("a", encoding="utf-8") as f:
            f.write(_json_ui.dumps(rec, ensure_ascii=False) + "\n")
    except Exception:
        pass

# 아래 5개 함수는 실제 UI 코드에서 해당 시점에 '한 줄'로 호출하세요.
# 예) 견적 탭을 여는 핸들러 마지막 줄에:  _ui_log("E1_TAB_OPEN")
def ui_smoke_tab_open(meta: dict | None = None):
    _ui_log("E1_TAB_OPEN", meta)

def ui_smoke_btn1(meta: dict | None = None):
    _ui_log("E2_BTN_CLICK1", meta)

def ui_smoke_btn2(meta: dict | None = None):
    _ui_log("E3_BTN_CLICK2", meta)

def ui_smoke_btn3(meta: dict | None = None):
    _ui_log("E4_BTN_CLICK3", meta)

def ui_smoke_table_ok(meta: dict | None = None):
    _ui_log("E5_TABLE_OK", meta)
# === /KISAN-AI PATCH ===


if __name__=="__main__":
    main()